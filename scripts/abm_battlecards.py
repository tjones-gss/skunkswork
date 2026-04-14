#!/usr/bin/env python3
"""
ABM Competitive Battlecards
NAM Intelligence Pipeline

Generates: data/exports/GSS_Competitive_Battlecards.xlsx
  Sheet 1: Battlecard Summary    — one row per competitor with detection counts
  Sheet 2: Detailed Battlecards  — per-competitor positioning, win themes, objections
  Sheet 3: Competitor-Account Matrix — companies where a competitor was detected
"""

import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from scripts.abm_shared import (
    BASE_DIR,
    EXPORTS_DIR,
    HEADER_FILL,
    HEADER_FONT,
    HEADER_ALIGNMENT,
    THIN_BORDER,
    THREAT_FILLS,
    THREAT_RESPONSE,
    WHITE_FONT,
    SECTION_FILL,
    SECTION_FONT,
    SUBTITLE_FONT,
    COMPETITOR_ALIASES,
    load_and_merge_data,
    load_events,
    load_competitors,
    get_email_provider,
    get_spf_list,
    get_tech_stack,
    get_associations_list,
    get_primary_contact,
    get_contacts,
    detect_competitor,
    compute_icp_score,
    assign_tier,
    style_header_row,
    write_section_header,
)

OUTPUT_PATH = EXPORTS_DIR / "GSS_Competitive_Battlecards.xlsx"

# ---------------------------------------------------------------------------
# Battlecard content — embedded positioning intel for each tracked competitor
# ---------------------------------------------------------------------------

BATTLECARD_CONTENT = {
    "Epicor": {
        "positioning": "GSS delivers faster ROI with manufacturing-first design, while Epicor's broad portfolio dilutes manufacturing focus.",
        "win_themes": [
            "Purpose-built for manufacturing vs. Epicor's multi-industry approach",
            "Faster implementation (90 days avg vs. Epicor's 6-12 months)",
            "Lower total cost of ownership — no per-module licensing surprises",
            "Real-time shop floor visibility included, not an add-on",
            "In-house support team — no outsourced partner network required",
        ],
        "objections": [
            (
                "Epicor is the industry leader",
                "Market share doesn't equal best fit. GSS specializes exclusively in manufacturing — every feature is built for the shop floor.",
            ),
            (
                "We need Epicor's industry-specific modules",
                "GSS covers 95% of manufacturing workflows natively. The remaining 5% is typically custom reporting, which GSS handles with built-in customization tools.",
            ),
            (
                "Our integrator recommended Epicor",
                "Integrators earn margin on complex implementations. GSS's simpler architecture means faster go-live and lower consulting fees.",
            ),
        ],
        "differentiators": [
            ("Shop Floor Control", "Real-time tracking built-in vs. Epicor's add-on MES module"),
            ("Implementation Speed", "90-day average vs. 6-12 months for Epicor Kinetic"),
            ("Pricing Model", "All-inclusive licensing vs. per-module pricing"),
            ("Support", "Direct in-house support vs. partner network"),
        ],
        "displacement_triggers": [
            "Epicor Kinetic migration deadline — companies forced to upgrade from Prophet 21/Eclipse",
            "Sticker shock at Epicor renewal pricing — 20-30% increases common",
        ],
        "talk_track": (
            "If you're evaluating Epicor, I'd love 15 minutes to show you how GSS delivers the same "
            "manufacturing capabilities in half the implementation time at a fraction of the cost. "
            "We're purpose-built for manufacturers like you."
        ),
    },
    "Plex": {
        "positioning": "GSS offers comparable cloud manufacturing ERP without Rockwell's industrial automation lock-in.",
        "win_themes": [
            "No vendor lock-in to Rockwell automation ecosystem",
            "More flexible for job shops and mixed-mode manufacturers",
            "Lower entry cost — no mandatory Rockwell hardware",
            "Stronger in discrete manufacturing vs. Plex's process focus",
        ],
        "objections": [
            (
                "Plex is pure cloud",
                "GSS offers cloud deployment too, with the flexibility to run on-premise if needed for air-gapped or low-connectivity environments.",
            ),
            (
                "We use Rockwell automation",
                "GSS integrates with any automation platform. Plex's tight Rockwell coupling limits your future automation choices.",
            ),
        ],
        "differentiators": [
            ("Deployment Flexibility", "Cloud + on-premise options vs. cloud-only"),
            ("Automation Agnostic", "Works with any PLC/automation vs. Rockwell-only optimization"),
            ("Job Shop Strength", "Built for job shops vs. Plex's repetitive manufacturing focus"),
        ],
        "displacement_triggers": [
            "Rockwell acquisition integration pain — Plex customers facing product roadmap uncertainty",
            "Companies wanting automation vendor diversity",
        ],
        "talk_track": (
            "Plex is solid for cloud manufacturing, but you shouldn't have to buy into the entire Rockwell "
            "ecosystem. GSS gives you the same cloud ERP with freedom to choose your automation partners."
        ),
    },
    "Syspro": {
        "positioning": "GSS matches SYSPRO's mid-market manufacturing focus with stronger US support and faster implementation.",
        "win_themes": [
            "US-based development and support vs. SYSPRO's South African HQ",
            "Faster implementation with manufacturing templates",
            "Stronger shop floor and real-time production tracking",
            "Better value — comparable features at lower TCO",
        ],
        "objections": [
            (
                "SYSPRO has strong manufacturing features",
                "SYSPRO is solid, but GSS was built exclusively for US manufacturers. Our workflows match American manufacturing practices out of the box.",
            ),
            (
                "SYSPRO is a NAM member",
                "So are many of our customers. NAM membership doesn't make an ERP better — manufacturing expertise does.",
            ),
        ],
        "differentiators": [
            ("US Manufacturing Focus", "Built by US manufacturers for US manufacturers"),
            ("Real-time Shop Floor", "Native real-time tracking vs. SYSPRO's batch updates"),
            ("Implementation", "90-day avg with in-house team vs. partner-dependent"),
        ],
        "displacement_triggers": [
            "SYSPRO version upgrade complexity — customers on older versions face costly migrations",
            "Support responsiveness issues due to timezone differences",
        ],
        "talk_track": (
            "If SYSPRO is on your shortlist, compare our implementation timeline and TCO side by side. "
            "GSS typically goes live 3x faster with dedicated US-based support."
        ),
    },
    "Acumatica": {
        "positioning": "GSS provides deeper manufacturing functionality than Acumatica's general-purpose cloud ERP.",
        "win_themes": [
            "Manufacturing-specific vs. Acumatica's horizontal platform",
            "Deeper shop floor control and production scheduling",
            "No consumption-based pricing surprises",
            "Purpose-built MES included, not third-party add-on",
        ],
        "objections": [
            (
                "Acumatica is modern cloud-native",
                "Cloud architecture doesn't equal manufacturing depth. GSS delivers cloud with 30+ years of manufacturing-specific development.",
            ),
            (
                "Acumatica's unlimited user licensing is attractive",
                "Unlimited users on a shallow manufacturing platform still requires add-ons. GSS includes everything manufacturers need.",
            ),
        ],
        "differentiators": [
            ("Manufacturing Depth", "30+ years of mfg-specific development vs. horizontal platform"),
            ("Production Scheduling", "Advanced finite scheduling built-in vs. basic MRP"),
            ("Pricing Clarity", "Fixed pricing vs. consumption-based model"),
        ],
        "displacement_triggers": [
            "Manufacturers outgrowing Acumatica's basic manufacturing module",
            "Hidden costs from required ISV add-ons for manufacturing-specific features",
        ],
        "talk_track": (
            "Acumatica is great for distribution and services, but manufacturers need purpose-built tools. "
            "Let me show you the manufacturing features that Acumatica requires add-ons to match."
        ),
    },
    "Infor": {
        "positioning": "GSS delivers simpler, faster manufacturing ERP without Infor's enterprise complexity and cost.",
        "win_themes": [
            "Simpler implementation — no army of consultants required",
            "Manufacturing-focused vs. Infor's multi-industry complexity",
            "Transparent pricing vs. Infor's enterprise sales model",
            "Faster time-to-value for mid-market manufacturers",
        ],
        "objections": [
            (
                "Infor has industry-specific CloudSuites",
                "Industry-specific marketing doesn't mean industry-specific software. GSS is actually built for manufacturing, not repackaged.",
            ),
            (
                "We need Infor's scale",
                "If you're evaluating both, you're likely mid-market — exactly where GSS excels and Infor over-serves.",
            ),
        ],
        "differentiators": [
            ("Complexity", "Right-sized for mid-market vs. enterprise overkill"),
            ("Implementation", "Weeks not months, in-house team not consultants"),
            ("Cost", "Transparent per-user pricing vs. enterprise negotiation"),
        ],
        "displacement_triggers": [
            "Infor CloudSuite migration forcing expensive re-implementations",
            "Mid-market customers tired of being treated as small fish by Infor sales",
        ],
        "talk_track": (
            "Infor serves the Fortune 500. If you're a mid-market manufacturer, you deserve an ERP partner "
            "who treats you as their primary customer, not an afterthought."
        ),
    },
    "Microsoft Dynamics": {
        "positioning": "GSS provides manufacturing-native ERP that Dynamics 365 requires extensive customization to match.",
        "win_themes": [
            "Manufacturing-native vs. Dynamics' horizontal platform requiring ISV add-ons",
            "Shop floor control included vs. requiring third-party MES",
            "Simpler licensing — no confusing per-app pricing tiers",
            "Manufacturing expertise in support team vs. generalist Microsoft partners",
        ],
        "objections": [
            (
                "We're a Microsoft shop",
                "GSS integrates fully with Microsoft 365, Azure, and Power Platform. You get the Microsoft ecosystem benefits without the ERP compromises.",
            ),
            (
                "Dynamics is the safe choice",
                "Safe for IT, expensive for manufacturing. Ask your shop floor team if they'd rather have purpose-built or customized-to-fit.",
            ),
        ],
        "differentiators": [
            ("Manufacturing Depth", "Native shop floor control vs. ISV add-ons"),
            ("Licensing Simplicity", "All-inclusive vs. per-app module pricing"),
            ("Implementation Risk", "Pre-built manufacturing workflows vs. custom development"),
        ],
        "displacement_triggers": [
            "Dynamics AX/NAV end-of-life forcing migration to D365",
            "Sticker shock when adding manufacturing-specific ISV modules",
        ],
        "talk_track": (
            "Love Microsoft? So do we — GSS runs beautifully alongside M365. But for manufacturing ERP, "
            "you need something built for the shop floor, not adapted from a financial system."
        ),
    },
    "Sap": {
        "positioning": "GSS delivers mid-market manufacturing ERP at a fraction of SAP's cost and complexity.",
        "win_themes": [
            "Right-sized for mid-market vs. SAP's enterprise bloat",
            "10x faster implementation timeline",
            "No mandatory consulting army",
            "Total cost typically 50-70% less than SAP Business One",
        ],
        "objections": [
            (
                "SAP is the gold standard",
                "For the Fortune 500, yes. For a mid-market manufacturer, SAP is expensive overengineering.",
            ),
        ],
        "differentiators": [
            ("Total Cost", "50-70% lower TCO for comparable manufacturing features"),
            ("Time to Value", "Weeks vs. months/years"),
            ("Complexity", "Built for mid-market, not scaled down from enterprise"),
        ],
        "displacement_triggers": [
            "SAP licensing audit triggering unexpected costs",
            "S/4HANA migration deadline pressure",
        ],
        "talk_track": (
            "If SAP quoted you, compare the total 5-year cost including implementation, consulting, and licensing. "
            "GSS typically delivers comparable manufacturing capability at 50% of the investment."
        ),
    },
    "Netsuite": {
        "positioning": "GSS provides deeper manufacturing functionality than NetSuite's finance-first platform.",
        "win_themes": [
            "Manufacturing-first vs. NetSuite's accounting-first design",
            "Real-time shop floor tracking included",
            "No Oracle upsell pressure after acquisition",
            "Better fit for complex manufacturing workflows",
        ],
        "objections": [
            (
                "NetSuite is cloud-native",
                "Cloud architecture is table stakes. Manufacturing depth is what separates GSS.",
            ),
        ],
        "differentiators": [
            ("Manufacturing Focus", "Built for shop floor vs. adapted from financial system"),
            ("Production Tracking", "Real-time vs. batch-updated inventory"),
        ],
        "displacement_triggers": [
            "Oracle price increases post-acquisition",
            "Manufacturers outgrowing NetSuite's basic manufacturing module",
        ],
        "talk_track": (
            "NetSuite is fantastic for finance and distribution. But if you're running a shop floor, "
            "you need an ERP that speaks manufacturing, not accounting."
        ),
    },
    "Sage": {
        "positioning": "GSS offers modern manufacturing ERP to replace aging Sage installations.",
        "win_themes": [
            "Modern architecture vs. Sage's legacy codebase",
            "Manufacturing-specific vs. Sage's general business focus",
            "Cloud-ready vs. Sage's on-premise heritage",
        ],
        "objections": [
            (
                "We've used Sage for 15 years",
                "That loyalty is admirable, but technology has evolved. Let us show you what modern manufacturing ERP looks like.",
            ),
        ],
        "differentiators": [
            ("Architecture", "Modern cloud-ready platform vs. legacy desktop"),
            ("Manufacturing Features", "Full MES + shop floor vs. basic inventory/order mgmt"),
        ],
        "displacement_triggers": [
            "Sage 100/300 end-of-support timelines",
            "Need for cloud/remote access post-COVID",
        ],
        "talk_track": (
            "If you've been on Sage for years, you know the limitations. GSS gives you a modern manufacturing "
            "ERP without starting from scratch — we migrate your data and train your team."
        ),
    },
    "Qad": {
        "positioning": "GSS matches QAD's manufacturing focus with simpler deployment and better mid-market fit.",
        "win_themes": [
            "Simpler implementation for mid-market",
            "Broader manufacturing mode support (job shop + repetitive + mixed)",
            "US-based support and development",
        ],
        "objections": [
            (
                "QAD specializes in manufacturing",
                "So does GSS — with 30+ years of exclusive manufacturing focus and a simpler deployment model.",
            ),
        ],
        "differentiators": [
            ("Deployment", "Faster implementation without enterprise complexity"),
            ("Mixed Mode", "Handles job shop, repetitive, and mixed-mode natively"),
        ],
        "displacement_triggers": [
            "QAD Adaptive ERP migration forcing re-implementation",
            "Mid-market customers seeking simpler alternative",
        ],
        "talk_track": (
            "QAD knows manufacturing, and so do we. The difference is GSS gets you live faster at a lower cost, "
            "with direct US-based support."
        ),
    },
    "Aptean": {
        "positioning": "GSS provides unified manufacturing ERP vs. Aptean's fragmented acquisition portfolio.",
        "win_themes": [
            "Single unified platform vs. Aptean's collection of acquired products",
            "Consistent user experience across all modules",
            "No product roadmap uncertainty from acquisitions",
        ],
        "objections": [
            (
                "Aptean has industry-specific solutions",
                "Aptean acquires solutions; GSS builds them. A unified platform means fewer integration headaches.",
            ),
        ],
        "differentiators": [
            ("Platform Unity", "Single codebase vs. acquired product portfolio"),
            ("Roadmap Clarity", "Single product roadmap vs. acquisition-driven direction"),
        ],
        "displacement_triggers": [
            "Aptean product consolidation forcing migrations",
            "Integration fatigue from multiple acquired products",
        ],
        "talk_track": (
            "Aptean has grown by acquisition, which means their customers often run on different platforms under "
            "one brand. GSS is one platform, one experience, one support team."
        ),
    },
    "Odoo": {
        "positioning": "GSS provides enterprise-grade manufacturing ERP that Odoo's open-source model can't match for production environments.",
        "win_themes": [
            "Enterprise-grade reliability for production manufacturing",
            "Dedicated manufacturing support team vs. community forums",
            "Pre-built manufacturing workflows vs. configuration-heavy",
            "Proven in regulated manufacturing environments",
        ],
        "objections": [
            (
                "Odoo is much cheaper",
                "The license is cheaper; the total cost isn't. Factor in customization, hosting, support, and the cost of downtime from a less mature manufacturing module.",
            ),
        ],
        "differentiators": [
            ("Reliability", "Enterprise SLA vs. community-supported"),
            ("Manufacturing Maturity", "30+ years vs. recent addition to Odoo's module library"),
        ],
        "displacement_triggers": [
            "Odoo community support inadequacy for production issues",
            "Manufacturing complexity outgrowing Odoo's module",
        ],
        "talk_track": (
            "Odoo is great for starting out, but when your manufacturing operations demand reliability and depth, "
            "GSS is where serious manufacturers graduate to."
        ),
    },
    "Delmiaworks": {
        "positioning": "GSS provides broader manufacturing ERP coverage beyond DELMIAWorks' plastics/injection-molding niche.",
        "win_themes": [
            "Broader industry coverage — job shops, metal fab, electronics, and more",
            "Not locked into a single vertical (74% plastics dependency is a risk)",
            "Larger installed base and longer track record",
            "Full ERP suite without vertical-specific pricing",
        ],
        "objections": [
            (
                "DELMIAWorks dominates our segment",
                "Dominating plastics doesn't mean they're the best fit for mixed-mode or diversifying manufacturers. GSS serves the full shop.",
            ),
        ],
        "differentiators": [
            ("Industry Breadth", "Serves all discrete manufacturers vs. plastics/process niche"),
            ("Roadmap Risk", "Independent roadmap vs. DELMIAWorks' Dassault/3DS acquisition layer"),
        ],
        "displacement_triggers": [
            "Manufacturers diversifying beyond plastics needing broader ERP coverage",
            "DELMIAWorks customers frustrated by Dassault parent company's enterprise focus",
        ],
        "talk_track": (
            "DELMIAWorks owns the plastics niche, but if your shop does more than injection molding, "
            "you deserve an ERP designed for the full range of manufacturing — that's GSS."
        ),
    },
    "Fulcrum": {
        "positioning": "GSS offers proven enterprise-grade reliability where Fulcrum is still building market credibility.",
        "win_themes": [
            "30+ years of manufacturing ERP track record vs. Fulcrum's emerging status",
            "Larger customer base and broader industry references",
            "More complete feature set for complex manufacturing environments",
            "Established partner ecosystem and integrations",
        ],
        "objections": [
            (
                "Fulcrum is newer and more modern",
                "Modern UI doesn't equal manufacturing depth. GSS combines modern architecture with three decades of manufacturing-specific functionality.",
            ),
        ],
        "differentiators": [
            ("Track Record", "30+ years live in manufacturing vs. emerging competitor"),
            ("Feature Completeness", "Full MES + ERP breadth vs. growing module library"),
        ],
        "displacement_triggers": [
            "Fulcrum customers hitting feature gaps as operations scale",
            "Risk-averse buyers seeking proven implementations",
        ],
        "talk_track": (
            "Fulcrum is growing fast, but when you're betting your shop floor on an ERP, you want "
            "a vendor with 30 years of references and a proven go-live track record. That's GSS."
        ),
    },
    "Eci (Jobboss/M1)": {
        "positioning": "GSS provides a complete manufacturing ERP platform where ECI's job shop tools are limited in scope and scale.",
        "win_themes": [
            "Full ERP suite vs. ECI's job shop-focused, point-solution roots",
            "Better suited for manufacturers growing beyond pure job shop complexity",
            "Unified platform — not a collection of acquired products (JobBOSS, M1, E2)",
            "Stronger reporting and business intelligence built in",
        ],
        "objections": [
            (
                "JobBOSS is built for job shops",
                "GSS is too — and also handles mixed-mode, repetitive, and project manufacturing. You won't outgrow it.",
            ),
            (
                "ECI is a NAM Gold member",
                "Sponsorship level doesn't equal product quality. GSS earns customer loyalty through results, not association fees.",
            ),
        ],
        "differentiators": [
            ("Platform Scope", "Full ERP vs. job-shop-focused point tools"),
            ("Scalability", "Handles growth from job shop to mid-enterprise"),
            ("Platform Unity", "Single product vs. ECI's acquired product family (JobBOSS, M1, E2)"),
        ],
        "displacement_triggers": [
            "Job shops growing to mixed-mode manufacturing needing broader ERP capabilities",
            "ECI product consolidation (JobBOSS/M1/E2) causing version confusion and migration pressure",
        ],
        "talk_track": (
            "If you're on JobBOSS or M1, you're likely bumping into limits as your business grows. "
            "GSS handles everything a job shop needs today and scales with you as complexity increases."
        ),
    },
    "Ifs": {
        "positioning": "GSS delivers comparable manufacturing ERP to IFS at significantly lower cost and complexity for mid-market manufacturers.",
        "win_themes": [
            "Right-sized for mid-market vs. IFS's defense/aerospace enterprise focus",
            "Faster implementation — IFS deployments typically take 12-18 months",
            "No requirement for large consulting teams",
            "Strong in discrete manufacturing without aerospace-grade complexity",
        ],
        "objections": [
            (
                "IFS specializes in our industry",
                "IFS specializes in aerospace, defense, and utilities. For general discrete manufacturing, GSS is purpose-built and far simpler to deploy.",
            ),
        ],
        "differentiators": [
            ("Market Focus", "Mid-market discrete manufacturing vs. IFS's defense/aerospace/energy niche"),
            ("Implementation", "90-day average vs. IFS's 12-18 month enterprise deployments"),
            ("Cost", "Mid-market pricing vs. enterprise IFS licensing"),
        ],
        "displacement_triggers": [
            "Mid-market manufacturers over-sold IFS and struggling with complexity",
            "IFS license costs exceeding budget for non-aerospace manufacturers",
        ],
        "talk_track": (
            "IFS is built for defense contractors and utilities. If you're a discrete manufacturer, "
            "you're paying enterprise prices for complexity you don't need. GSS is built for you."
        ),
    },
}

# Threat-level sort order
THREAT_ORDER = {"HIGH": 0, "MEDIUM": 1, "EMERGING": 2, "LOW": 3}


def _col(idx: int) -> str:
    return get_column_letter(idx)


def _set_cell(ws, row: int, col: int, value, font=None, fill=None, alignment=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell


def _wrap(ws, row: int, col: int, value, font=None, fill=None):
    """Write a cell with wrap-text alignment."""
    alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    return _set_cell(ws, row, col, value, font=font, fill=fill, alignment=alignment, border=THIN_BORDER)


# ---------------------------------------------------------------------------
# Detection helpers
# ---------------------------------------------------------------------------

def _detect_source(rec: dict, competitor_name: str) -> str:
    """Return which field first matched the competitor: erp_system, tech_stack, or spf_services."""
    from scripts.abm_shared import get_tech_stack, get_spf_list, COMPETITOR_ALIASES

    # Normalise competitor key to look up aliases
    comp_lower = competitor_name.lower()
    aliases = COMPETITOR_ALIASES.get(comp_lower, [])
    if not aliases:
        # Try partial match on alias keys
        for key in COMPETITOR_ALIASES:
            if comp_lower in key or key in comp_lower:
                aliases = COMPETITOR_ALIASES[key]
                break

    erp = (rec.get("erp_system") or "").lower()
    ts_list = [t.lower() for t in get_tech_stack(rec)]
    spf_list = [s.lower() for s in get_spf_list(rec)]

    for alias in aliases:
        if alias in erp:
            return "erp_system"
    for alias in aliases:
        for ts in ts_list:
            if alias in ts:
                return "tech_stack"
    for alias in aliases:
        for spf in spf_list:
            if alias in spf:
                return "spf_services"
    return "tech_stack"


def _build_competitor_map(companies: list[dict]) -> dict[str, list[dict]]:
    """Build {competitor_title: [company_records]} map using detect_competitor()."""
    result = defaultdict(list)
    for rec in companies:
        comp = detect_competitor(rec)
        if comp:
            result[comp].append(rec)
    return result


# ---------------------------------------------------------------------------
# Event-to-competitor mapping
# ---------------------------------------------------------------------------

def _events_for_competitor(presence_text: str, events: list[dict]) -> str:
    """Find event names mentioned in a competitor's presence field."""
    presence_lower = presence_text.lower()
    matched = []
    for ev in events:
        name = ev.get("event_name", "")
        # Check if any word token from the event name appears in presence
        keywords = [w.strip().lower() for w in name.replace("/", " ").split() if len(w.strip()) > 2]
        if any(kw in presence_lower for kw in keywords):
            matched.append(name)
    return "; ".join(matched) if matched else ""


# ---------------------------------------------------------------------------
# Sheet 1 — Battlecard Summary
# ---------------------------------------------------------------------------

SUMMARY_COLUMNS = [
    ("Competitor", 20),
    ("Threat Level", 12),
    ("Show Presence", 40),
    ("Strategy Notes", 40),
    ("GSS Response", 15),
    ("Companies Detected", 15),
    ("Primary Associations", 20),
    ("Key Events", 25),
]


def _build_summary_sheet(ws, competitors: list[dict], competitor_map: dict[str, list[dict]], events: list[dict]):
    ws.title = "Battlecard Summary"

    # Title row
    title_cell = ws.cell(row=1, column=1, value="GSS Competitive Intelligence — Battlecard Summary")
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(SUMMARY_COLUMNS))

    # Subtitle
    sub_cell = ws.cell(row=2, column=1, value=f"15 tracked ERP competitors | NAM Intelligence Pipeline")
    sub_cell.font = SUBTITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(SUMMARY_COLUMNS))
    ws.row_dimensions[2].height = 16

    # Header row
    style_header_row(ws, SUMMARY_COLUMNS, row=3)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = None  # style_header_row sets this; we'll redo after data

    # Sort competitors: HIGH, MEDIUM, EMERGING, LOW
    sorted_comps = sorted(
        competitors,
        key=lambda c: THREAT_ORDER.get(c.get("threat_level", "LOW").upper(), 3),
    )

    for row_idx, comp_row in enumerate(sorted_comps, start=4):
        name = comp_row.get("competitor", "")
        threat = (comp_row.get("threat_level") or "LOW").upper()
        presence = comp_row.get("presence", "")
        notes = comp_row.get("strategy_notes", "")
        gss_response = THREAT_RESPONSE.get(threat, "Track Only")

        # Detected companies
        # Try exact title match first, then fuzzy
        detected_list = competitor_map.get(name.title(), [])
        if not detected_list:
            # Try matching by normalised name
            for key, recs in competitor_map.items():
                if key.lower() in name.lower() or name.lower() in key.lower():
                    detected_list = recs
                    break
        count = len(detected_list)

        # Primary associations of detected companies
        assoc_counter: Counter = Counter()
        for rec in detected_list:
            for a in [x.strip() for x in (rec.get("associations") or "").split(";") if x.strip()]:
                assoc_counter[a] += 1
        primary_assoc = "; ".join(a for a, _ in assoc_counter.most_common(3))

        key_events = _events_for_competitor(presence, events)

        ws.row_dimensions[row_idx].height = 30

        values = [name, threat, presence, notes, gss_response, count, primary_assoc, key_events]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

            # Color-code threat level column
            if col_idx == 2:
                fill = THREAT_FILLS.get(threat, THREAT_FILLS["LOW"])
                cell.fill = fill
                cell.font = WHITE_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Re-apply auto-filter over data range
    last_row = 3 + len(sorted_comps)
    ws.auto_filter.ref = f"A3:{_col(len(SUMMARY_COLUMNS))}{last_row}"


# ---------------------------------------------------------------------------
# Sheet 2 — Detailed Battlecards
# ---------------------------------------------------------------------------

_DETAIL_COLS = 2  # A = label/section, B = content


def _threat_fill_for(name: str, competitors: list[dict]) -> PatternFill:
    """Find threat level fill for a competitor by name."""
    for c in competitors:
        if c.get("competitor", "").lower() == name.lower():
            threat = (c.get("threat_level") or "LOW").upper()
            return THREAT_FILLS.get(threat, THREAT_FILLS["LOW"])
    return THREAT_FILLS["LOW"]


def _generic_battlecard(comp_name: str) -> dict:
    return {
        "positioning": f"GSS provides purpose-built manufacturing ERP as a stronger mid-market alternative to {comp_name}.",
        "win_themes": [
            "30+ years of exclusive manufacturing focus",
            "Faster implementation with dedicated in-house team",
            "All-inclusive pricing — no per-module surprises",
            "Real-time shop floor control included",
        ],
        "objections": [
            (
                f"{comp_name} is already in our budget",
                "Budget comfort doesn't always equal best fit. Let us show you a side-by-side comparison.",
            ),
        ],
        "differentiators": [
            ("Manufacturing Focus", f"Purpose-built vs. {comp_name}'s broader scope"),
            ("Implementation", "90-day average go-live with in-house team"),
            ("Support", "Direct US-based manufacturing experts"),
        ],
        "displacement_triggers": [
            f"{comp_name} version upgrade pain or rising renewal costs",
            "Desire for a more manufacturing-focused ERP partner",
        ],
        "talk_track": (
            f"If {comp_name} is on your shortlist, we'd love 15 minutes to walk you through how GSS "
            "stacks up on the metrics that matter most to manufacturers."
        ),
    }


LABEL_FONT = Font(bold=True, size=10, color="1F4E79")
BODY_FONT = Font(size=10)
OBJECTION_Q_FONT = Font(bold=True, italic=True, size=10, color="555555")
TRIGGER_FONT = Font(bold=True, size=10, color="C65911")
TALK_TRACK_FONT = Font(italic=True, size=10, color="1F4E79")
BULLET_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")


def _write_battlecard(ws, row: int, comp_name: str, content: dict, threat_fill: PatternFill) -> int:
    """Write a single competitor's battlecard section. Returns next available row."""

    # ── Section header ────────────────────────────────────────────────
    header_text = f"  {comp_name.upper()}  |  Threat Level: {_threat_label(comp_name, threat_fill)}"
    header_cell = ws.cell(row=row, column=1, value=f"  {comp_name}")
    header_cell.fill = threat_fill
    header_cell.font = Font(bold=True, size=13, color="FFFFFF")
    header_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row, column=2).fill = threat_fill
    ws.row_dimensions[row].height = 24
    row += 1

    # ── Positioning ───────────────────────────────────────────────────
    _write_label_row(ws, row, "POSITIONING", content["positioning"], LABEL_FONT, BODY_FONT)
    ws.row_dimensions[row].height = 42
    row += 1

    # ── Win Themes ────────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="WIN THEMES").font = LABEL_FONT
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="top")
    for i, theme in enumerate(content["win_themes"]):
        fill = BULLET_FILL if i % 2 == 0 else ALT_ROW_FILL
        ws.cell(row=row, column=2, value=f"  {i + 1}. {theme}").font = BODY_FONT
        ws.cell(row=row, column=2).fill = fill
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.row_dimensions[row].height = 18
        row += 1

    # ── Objections / Rebuttals ────────────────────────────────────────
    obj_header = ws.cell(row=row, column=1, value="OBJECTIONS & REBUTTALS")
    obj_header.font = LABEL_FONT
    obj_header.alignment = Alignment(horizontal="left", vertical="top")
    obj_header.border = THIN_BORDER
    ws.cell(row=row, column=2, value="Objection  |  GSS Rebuttal").font = Font(bold=True, size=10, color="888888")
    ws.cell(row=row, column=2).border = THIN_BORDER
    ws.row_dimensions[row].height = 16
    row += 1
    for i, (objection, rebuttal) in enumerate(content["objections"]):
        fill = BULLET_FILL if i % 2 == 0 else ALT_ROW_FILL
        ws.cell(row=row, column=1, value=f'  Q: "{objection}"').font = OBJECTION_Q_FONT
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.cell(row=row, column=2, value=f"  A: {rebuttal}").font = BODY_FONT
        ws.cell(row=row, column=2).fill = fill
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 36
        row += 1

    # ── Technical Differentiators ─────────────────────────────────────
    diff_header = ws.cell(row=row, column=1, value="TECHNICAL DIFFERENTIATORS")
    diff_header.font = LABEL_FONT
    diff_header.border = THIN_BORDER
    ws.cell(row=row, column=2, value="Feature Area  |  GSS Advantage").font = Font(bold=True, size=10, color="888888")
    ws.cell(row=row, column=2).border = THIN_BORDER
    ws.row_dimensions[row].height = 16
    row += 1
    for i, (feature, advantage) in enumerate(content["differentiators"]):
        fill = BULLET_FILL if i % 2 == 0 else ALT_ROW_FILL
        ws.cell(row=row, column=1, value=f"  {feature}").font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row, column=2, value=f"  {advantage}").font = BODY_FONT
        ws.cell(row=row, column=2).fill = fill
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 18
        row += 1

    # ── Displacement Triggers ─────────────────────────────────────────
    trig_label = ws.cell(row=row, column=1, value="DISPLACEMENT TRIGGERS")
    trig_label.font = LABEL_FONT
    trig_label.border = THIN_BORDER
    trig_label.alignment = Alignment(horizontal="left", vertical="top")
    for i, trigger in enumerate(content["displacement_triggers"]):
        fill = BULLET_FILL if i % 2 == 0 else ALT_ROW_FILL
        if i > 0:
            ws.cell(row=row, column=1).border = THIN_BORDER
            ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=2, value=f"  \u26a1 {trigger}").font = TRIGGER_FONT
        ws.cell(row=row, column=2).fill = fill
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 18
        row += 1

    # ── Talk Track ────────────────────────────────────────────────────
    _write_label_row(ws, row, "TALK TRACK", content["talk_track"], LABEL_FONT, TALK_TRACK_FONT,
                     fill=PatternFill(start_color="EAF3FB", end_color="EAF3FB", fill_type="solid"))
    ws.row_dimensions[row].height = 52
    row += 1

    # Spacer
    ws.row_dimensions[row].height = 6
    row += 1

    return row


def _write_label_row(ws, row: int, label: str, value: str, label_font, value_font, fill=None):
    lc = ws.cell(row=row, column=1, value=f"  {label}")
    lc.font = label_font
    lc.alignment = Alignment(horizontal="left", vertical="top")
    lc.border = THIN_BORDER
    if fill:
        lc.fill = fill
    vc = ws.cell(row=row, column=2, value=value)
    vc.font = value_font
    vc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    vc.border = THIN_BORDER
    if fill:
        vc.fill = fill


def _threat_label(comp_name: str, threat_fill: PatternFill) -> str:
    """Not actually used in header text; kept for future use."""
    return ""


def _lookup_content(name: str) -> dict:
    """
    Look up battlecard content for a competitor name, trying progressively looser matches.
    Falls back to _generic_battlecard() if no match found.
    """
    # Try exact key first
    if name in BATTLECARD_CONTENT:
        return BATTLECARD_CONTENT[name]
    # Title-case (e.g. "SYSPRO" -> "Syspro")
    if name.title() in BATTLECARD_CONTENT:
        return BATTLECARD_CONTENT[name.title()]
    # First token before "/" or "(" (e.g. "Plex/Rockwell" -> "Plex", "ECI (JobBOSS/M1)" -> "Eci (Jobboss/M1)")
    first_token = name.split("/")[0].strip().split("(")[0].strip().title()
    if first_token in BATTLECARD_CONTENT:
        return BATTLECARD_CONTENT[first_token]
    # Full title() of the whole name (catches "ECI (JobBOSS/M1)" -> "Eci (Jobboss/M1)")
    full_title = name.title()
    # Normalise parens for ECI-style names
    import re as _re
    full_norm = _re.sub(r"[^A-Za-z0-9 ]", "", full_title).strip()
    for key in BATTLECARD_CONTENT:
        key_norm = _re.sub(r"[^A-Za-z0-9 ]", "", key).strip()
        if key_norm.lower() == full_norm.lower():
            return BATTLECARD_CONTENT[key]
    # Partial substring match
    name_lower = name.lower()
    for key in BATTLECARD_CONTENT:
        if key.lower() in name_lower or name_lower.split("/")[0].strip() in key.lower():
            return BATTLECARD_CONTENT[key]
    return _generic_battlecard(name)


def _build_detail_sheet(ws, competitors: list[dict]):
    ws.title = "Detailed Battlecards"
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 85

    # Title
    title = ws.cell(row=1, column=1, value="GSS Competitive Battlecards — Sales Enablement")
    title.font = Font(bold=True, size=14, color="1F4E79")
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws.row_dimensions[1].height = 28

    sub = ws.cell(row=2, column=1, value="Use these cards before every competitive sales call")
    sub.font = SUBTITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    ws.row_dimensions[2].height = 16

    current_row = 4

    # Sort by threat level
    sorted_comps = sorted(
        competitors,
        key=lambda c: THREAT_ORDER.get(c.get("threat_level", "LOW").upper(), 3),
    )

    for comp_row in sorted_comps:
        name = comp_row.get("competitor", "")
        threat = (comp_row.get("threat_level") or "LOW").upper()
        threat_fill = THREAT_FILLS.get(threat, THREAT_FILLS["LOW"])

        content = _lookup_content(name)

        current_row = _write_battlecard(ws, current_row, name, content, threat_fill)


# ---------------------------------------------------------------------------
# Sheet 3 — Competitor-Account Matrix
# ---------------------------------------------------------------------------

MATRIX_COLUMNS = [
    ("Company Name", 35),
    ("Website", 30),
    ("State", 8),
    ("Association(s)", 12),
    ("Detected Competitor", 20),
    ("Detection Source", 15),
    ("Quality Score", 12),
    ("ICP Score", 12),
    ("Primary Contact", 25),
    ("Contact Email", 30),
]


def _build_matrix_sheet(ws, companies: list[dict]):
    ws.title = "Competitor-Account Matrix"

    # Title
    title = ws.cell(row=1, column=1, value="Competitive Displacement Targets — Detected Competitor Usage")
    title.font = Font(bold=True, size=14, color="1F4E79")
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(MATRIX_COLUMNS))
    ws.row_dimensions[1].height = 24

    sub = ws.cell(row=2, column=1, value="Companies where a competing ERP product was detected in tech stack, SPF services, or erp_system field")
    sub.font = SUBTITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(MATRIX_COLUMNS))
    ws.row_dimensions[2].height = 16

    style_header_row(ws, MATRIX_COLUMNS, row=3)
    ws.freeze_panes = "A4"

    # Build rows: only companies with a detected competitor
    matrix_rows = []
    for rec in companies:
        comp = detect_competitor(rec)
        if not comp:
            continue
        icp = compute_icp_score(rec)
        icp_score = icp["icp_score"]
        quality_score = rec.get("quality_score", 0)
        try:
            quality_score = int(quality_score)
        except (ValueError, TypeError):
            quality_score = 0
        contact_name, contact_email = get_primary_contact(rec)
        assocs = "; ".join(get_associations_list(rec))
        detection_source = _detect_source(rec, comp)
        matrix_rows.append({
            "company_name": rec.get("company_name", ""),
            "website": rec.get("website") or rec.get("domain", ""),
            "state": rec.get("state", ""),
            "associations": assocs,
            "competitor": comp,
            "detection_source": detection_source,
            "quality_score": quality_score,
            "icp_score": icp_score,
            "contact_name": contact_name,
            "contact_email": contact_email,
        })

    # Sort by competitor, then ICP score descending
    matrix_rows.sort(key=lambda r: (r["competitor"], -r["icp_score"]))

    for row_idx, mr in enumerate(matrix_rows, start=4):
        alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid") if row_idx % 2 == 0 else None
        values = [
            mr["company_name"],
            mr["website"],
            mr["state"],
            mr["associations"],
            mr["competitor"],
            mr["detection_source"],
            mr["quality_score"],
            mr["icp_score"],
            mr["contact_name"],
            mr["contact_email"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            cell.border = THIN_BORDER
            if alt_fill:
                cell.fill = alt_fill
        ws.row_dimensions[row_idx].height = 16

    # Auto-filter
    last_row = max(4, 3 + len(matrix_rows))
    ws.auto_filter.ref = f"A3:{_col(len(MATRIX_COLUMNS))}{last_row}"

    return len(matrix_rows)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Loading data...")
    companies = load_and_merge_data()
    events = load_events()
    competitors = load_competitors()

    print(f"  {len(companies):,} companies loaded")
    print(f"  {len(events)} events loaded")
    print(f"  {len(competitors)} competitors loaded")

    competitor_map = _build_competitor_map(companies)
    total_detected = sum(len(v) for v in competitor_map.values())
    print(f"  {total_detected} companies with a detected competitor ({len(competitor_map)} competitors found)")

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Sheet 1
    print("Building Sheet 1: Battlecard Summary...")
    ws1 = wb.create_sheet("Battlecard Summary")
    _build_summary_sheet(ws1, competitors, competitor_map, events)

    # Sheet 2
    print("Building Sheet 2: Detailed Battlecards...")
    ws2 = wb.create_sheet("Detailed Battlecards")
    _build_detail_sheet(ws2, competitors)

    # Sheet 3
    print("Building Sheet 3: Competitor-Account Matrix...")
    ws3 = wb.create_sheet("Competitor-Account Matrix")
    matrix_count = _build_matrix_sheet(ws3, companies)

    # Save
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"\nSaved: {OUTPUT_PATH}")
    print(f"  Sheet 1: {len(competitors)} competitors summarised")
    print(f"  Sheet 2: {len(competitors)} detailed battlecards")
    print(f"  Sheet 3: {matrix_count} competitor-detected accounts")


if __name__ == "__main__":
    main()
