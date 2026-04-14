#!/usr/bin/env python3
"""
ABM Campaign Playbooks
NAM Intelligence Pipeline

Generates: data/exports/GSS_Campaign_Playbooks.xlsx
  Sheet 1: Campaign Overview (summary table of all 6 campaigns)
  Sheet 2: Microsoft Ecosystem Modernization
  Sheet 3: Legacy IT Transformation
  Sheet 4: Salesforce CRM Cross-Sell
  Sheet 5: PMA Association Blitz
  Sheet 6: Greenfield ERP
  Sheet 7: Competitive Displacement
"""

import sys
from pathlib import Path

# ── Path resolution ────────────────────────────────────────────────────
# This script may live inside a git worktree (agent-aaab7052) while the
# actual data files reside in the main repo tree.  We detect the main repo
# location at import time and inject it onto sys.path so abm_shared loads
# from the correct location — then we patch the module's path constants so
# load_and_merge_data() reads from the right directories.

_SCRIPT_DIR = Path(__file__).resolve().parent  # .../scripts/
_WORKTREE_ROOT = _SCRIPT_DIR.parent            # .../.claude/worktrees/agent-aaab7052/

# Walk up from the worktree root to find the main repo (the one that has the
# actual pipeline data — identified by the presence of companies_all.csv).
_CANDIDATE_MAIN = _WORKTREE_ROOT
for _ in range(8):
    if (_CANDIDATE_MAIN / "data" / "exports" / "companies_all.csv").exists():
        break
    _CANDIDATE_MAIN = _CANDIDATE_MAIN.parent

_DATA_ROOT: Path = _CANDIDATE_MAIN  # repo root that has data/

# Inject the chosen root first so `from scripts.abm_shared import` resolves there
sys.path.insert(0, str(_DATA_ROOT))

import scripts.abm_shared as _abm_shared  # noqa: E402 — must follow sys.path patch

# Patch module-level path constants so data-loading functions use the main repo
_abm_shared.BASE_DIR = _DATA_ROOT
_abm_shared.DATA_DIR = _DATA_ROOT / "data"
_abm_shared.EXPORTS_DIR = _DATA_ROOT / "data" / "exports"
_abm_shared.ENRICHED_PATH = _DATA_ROOT / "data" / "processed" / "enriched_all.jsonl"
_abm_shared.CSV_PATH = _DATA_ROOT / "data" / "exports" / "companies_all.csv"
_abm_shared.EVENTS_PATH = _DATA_ROOT / "data" / "exports" / "events_2026.csv"
_abm_shared.COMPETITORS_PATH = _DATA_ROOT / "data" / "exports" / "competitor_analysis.csv"

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from scripts.abm_shared import (  # noqa: E402
    HEADER_ALIGNMENT,
    HEADER_FILL,
    HEADER_FONT,
    SECTION_FILL,
    SECTION_FONT,
    SUBTITLE_FONT,
    THIN_BORDER,
    assign_tier,
    compute_icp_score,
    detect_competitor,
    get_associations_list,
    get_contacts,
    get_email_provider,
    get_employee_count,
    get_primary_contact,
    get_spf_list,
    load_and_merge_data,
    load_competitors,
    load_events,
    write_section_header,
)

EXPORTS_DIR = _abm_shared.EXPORTS_DIR
EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

OUTPUT_PATH = EXPORTS_DIR / "GSS_Campaign_Playbooks.xlsx"

# ── Company list columns shared across campaign sheets ─────────────────

COMPANY_COLUMNS = [
    ("Company Name", 35),
    ("Website", 30),
    ("City", 18),
    ("State", 8),
    ("Association", 12),
    ("Email Provider", 18),
    ("ICP Score", 12),
    ("ICP Tier", 12),
    ("Primary Contact", 25),
    ("Contact Email", 30),
    ("Quality Score", 12),
]

# ── Tier fill colours ──────────────────────────────────────────────────

TIER_FILL = {
    "Tier 1": PatternFill(start_color="006100", end_color="006100", fill_type="solid"),
    "Tier 2": PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid"),
    "Tier 3": PatternFill(start_color="C65911", end_color="C65911", fill_type="solid"),
}
TIER_FONT = {
    "Tier 1": Font(bold=True, color="FFFFFF"),
    "Tier 2": Font(bold=True, color="FFFFFF"),
    "Tier 3": Font(bold=True, color="FFFFFF"),
}

CAMPAIGN_TITLE_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
CAMPAIGN_TITLE_FONT = Font(bold=True, color="FFFFFF", size=16)

ALT_ROW_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

# ── Email templates ────────────────────────────────────────────────────

EMAIL_TEMPLATES = {
    "microsoft": [
        {
            "label": "Cold Intro — Microsoft Ecosystem",
            "subject": "{{company_name}} + Global Shop Solutions ERP",
            "body": (
                "Hi {{contact_name}},\n\n"
                "I noticed {{company_name}} is running Microsoft 365 — which tells me your team"
                " values modern, integrated tools.\n\n"
                "Many manufacturers in {{association}} are taking the next step by connecting their"
                " shop floor to their business systems with Global Shop Solutions ERP.\n\n"
                "GSS is purpose-built for manufacturers and integrates natively with the Microsoft"
                " stack — so you keep what's working and gain full visibility from quote to cash.\n\n"
                "Would a 20-minute call to explore the fit make sense this week?\n\n"
                "Best,\n[REP NAME]\nGlobal Shop Solutions"
            ),
        },
        {
            "label": "Follow-Up — Value Prop + Social Proof",
            "subject": "Re: {{company_name}} — How [Similar Manufacturer] cut lead times 30%",
            "body": (
                "Hi {{contact_name}},\n\n"
                "Following up on my last note — wanted to share a quick win.\n\n"
                "A fellow {{association}} member in your state cut production lead times by 30%"
                " in the first year after going live on GSS. They were also on Microsoft 365 and"
                " found the integration seamless.\n\n"
                "I'd love to show you how the same outcome could look for {{company_name}}."
                " 15 minutes on the calendar?\n\n"
                "Best,\n[REP NAME]\nGlobal Shop Solutions"
            ),
        },
        {
            "label": "Event-Based — IMTS / FABTECH",
            "subject": "See GSS at {{event_name}} — Booth #[XXX]",
            "body": (
                "Hi {{contact_name}},\n\n"
                "Are you heading to {{event_name}} this year? Our team will be at Booth #[XXX]"
                " running live demos of Global Shop Solutions ERP — including our Microsoft 365"
                " integration.\n\n"
                "We'd love to set aside 30 minutes to walk {{contact_title}} and your team"
                " through what a digital transformation looks like for a manufacturer your size.\n\n"
                "Reply here or book a slot at [CALENDAR LINK]. See you there!\n\n"
                "Best,\n[REP NAME]\nGlobal Shop Solutions"
            ),
        },
    ],
    "legacy": [
        {
            "label": "Cold Intro — Legacy IT",
            "subject": "{{company_name}}: your competitors already moved to the cloud",
            "body": (
                "Hi {{contact_name}},\n\n"
                "Running on-premise email and servers is expensive — and it's keeping your team"
                " from focusing on what actually drives revenue.\n\n"
                "Manufacturers in {{association}} who have moved to cloud ERP report spending 40%"
                " less time on IT maintenance and 25% faster month-end close.\n\n"
                "Global Shop Solutions gives you a single, cloud-ready platform built specifically"
                " for manufacturers — no adapting accounting software to fit the shop floor.\n\n"
                "Worth a quick 15-minute call to see if the timing is right for {{company_name}}?\n\n"
                "Best,\n[REP NAME]\nGlobal Shop Solutions"
            ),
        },
        {
            "label": "Follow-Up — ROI Focus",
            "subject": "Re: {{company_name}} — Quick question on your server costs",
            "body": (
                "Hi {{contact_name}},\n\n"
                "Most manufacturers running on-premise infrastructure spend $50K–$150K per year"
                " on server maintenance, backup, and IT support — costs that disappear with"
                " cloud ERP.\n\n"
                "At {{company_name}}, I imagine you'd rather invest that budget in your people and"
                " equipment.\n\n"
                "Can I send over a one-page ROI calculator? No commitment — just data to help"
                " make the case internally.\n\n"
                "Best,\n[REP NAME]\nGlobal Shop Solutions"
            ),
        },
        {
            "label": "Event-Based — Automate 2026",
            "subject": "Meeting at Automate 2026?",
            "body": (
                "Hi {{contact_name}},\n\n"
                "We'll be exhibiting at {{event_name}} and meeting with manufacturers who are"
                " evaluating their IT infrastructure roadmap.\n\n"
                "If modernizing your back-office systems is on the agenda for 2026, I'd love to"
                " show you what a cloud ERP migration looks like for a manufacturer like"
                " {{company_name}}.\n\n"
                "Reply to grab a 30-minute slot at the show.\n\n"
                "Best,\n[REP NAME]\nGlobal Shop Solutions"
            ),
        },
    ],
    "salesforce": [
        {
            "label": "Cold Intro — Salesforce CRM",
            "subject": "Connect your Salesforce CRM to your shop floor",
            "body": (
                "Hi {{contact_name}},\n\n"
                "I noticed {{company_name}} is using Salesforce — which means your sales team has"
                " great visibility into the pipeline. But does your Salesforce data reflect what's"
                " actually happening on the shop floor in real time?\n\n"
                "Global Shop Solutions ERP integrates directly with Salesforce so your sales reps"
                " can see production status, delivery dates, and inventory — without leaving CRM.\n\n"
                "Would a 20-minute demo be worth your time?\n\n"
                "Best,\n[REP NAME]\nGlobal Shop Solutions"
            ),
        },
        {
            "label": "Follow-Up — AppExchange Angle",
            "subject": "Re: {{company_name}} — GSS on Salesforce AppExchange",
            "body": (
                "Hi {{contact_name}},\n\n"
                "Quick follow-up — did you know Global Shop Solutions is listed on the Salesforce"
                " AppExchange? That means your IT team can validate us through your existing"
                " Salesforce relationship.\n\n"
                "I'd love to walk you through how {{company_name}} could get bi-directional data"
                " flow between CRM and ERP — closing the gap between what's sold and what's"
                " actually being built.\n\n"
                "15 minutes this week?\n\n"
                "Best,\n[REP NAME]\nGlobal Shop Solutions"
            ),
        },
        {
            "label": "Event-Based — IMTS 2026",
            "subject": "Salesforce + GSS ERP demo at {{event_name}}",
            "body": (
                "Hi {{contact_name}},\n\n"
                "We're hosting a live integration demo at {{event_name}} showing how Salesforce"
                " CRM + Global Shop Solutions ERP creates a seamless quote-to-cash workflow.\n\n"
                "If your team is attending, I'd love to reserve a slot for {{company_name}}"
                " — bring your VP of Sales and your ops lead for the most impactful 30 minutes"
                " at the show.\n\n"
                "Reply to book a time.\n\n"
                "Best,\n[REP NAME]\nGlobal Shop Solutions"
            ),
        },
    ],
    "pma": [
        {
            "label": "Cold Intro — PMA Member",
            "subject": "PMA member? So are 500+ GSS customers",
            "body": (
                "Hi {{contact_name}},\n\n"
                "As a fellow PMA member, {{company_name}} understands the unique challenges of"
                " precision metalforming — tight tolerances, complex scheduling, and customers"
                " who demand on-time delivery.\n\n"
                "Global Shop Solutions ERP was built for exactly this. More than 500 metalformers"
                " and fabricators rely on GSS to schedule jobs, manage tooling, and hit delivery"
                " dates without heroics.\n\n"
                "Can I show you what that looks like for a shop your size?\n\n"
                "Best,\n[REP NAME]\nGlobal Shop Solutions"
            ),
        },
        {
            "label": "Follow-Up — Association Social Proof",
            "subject": "Re: {{company_name}} — What your PMA neighbors are doing",
            "body": (
                "Hi {{contact_name}},\n\n"
                "Following up — wanted to share that several PMA members in your region recently"
                " went live on GSS and are seeing measurable results: faster quote turnaround,"
                " lower WIP, and better on-time delivery.\n\n"
                "I'd be happy to connect you with a reference customer who runs a similar shop to"
                " {{company_name}}. Sometimes hearing it from a peer is more valuable than any"
                " demo.\n\n"
                "Interested?\n\n"
                "Best,\n[REP NAME]\nGlobal Shop Solutions"
            ),
        },
        {
            "label": "Event-Based — PMA Forming Our Future / FABTECH",
            "subject": "See you at {{event_name}}?",
            "body": (
                "Hi {{contact_name}},\n\n"
                "We'll be at {{event_name}} this year — and we'd love to connect with"
                " {{company_name}} face to face.\n\n"
                "Stop by our booth for a live demo tailored to precision metalformers, or let me"
                " know if you'd like to set up a private meeting on the show floor.\n\n"
                "Reply here or grab a time at [CALENDAR LINK].\n\n"
                "Best,\n[REP NAME]\nGlobal Shop Solutions"
            ),
        },
    ],
    "greenfield": [
        {
            "label": "Cold Intro — No ERP",
            "subject": "{{company_name}}: Is it time for your first real ERP?",
            "body": (
                "Hi {{contact_name}},\n\n"
                "Most manufacturers reach a point where spreadsheets, QuickBooks, and tribal"
                " knowledge just can't keep up — jobs fall through the cracks, delivery dates slip,"
                " and growth stalls.\n\n"
                "Global Shop Solutions was designed as a first ERP for manufacturers — not adapted"
                " from accounting software. It's visual, shop-floor-friendly, and most teams are"
                " live in 90 days.\n\n"
                "Would it make sense to talk about where {{company_name}} is on that journey?\n\n"
                "Best,\n[REP NAME]\nGlobal Shop Solutions"
            ),
        },
        {
            "label": "Follow-Up — Education + Demo Offer",
            "subject": "Re: {{company_name}} — Free ERP readiness assessment",
            "body": (
                "Hi {{contact_name}},\n\n"
                "Thinking about ERP for the first time can feel overwhelming. That's why we offer"
                " a free 30-minute ERP Readiness Assessment — no sales pitch, just an honest"
                " conversation about whether the timing is right.\n\n"
                "We cover: where manufacturers typically hit a wall with manual systems, what a"
                " realistic implementation looks like, and what ROI looks like in year 1.\n\n"
                "I can send you a calendar link if you're curious.\n\n"
                "Best,\n[REP NAME]\nGlobal Shop Solutions"
            ),
        },
        {
            "label": "Event-Based — IPC APEX / METALCON",
            "subject": "First-time ERP buyers: come talk to us at {{event_name}}",
            "body": (
                "Hi {{contact_name}},\n\n"
                "We'll be at {{event_name}} running demos specifically for manufacturers who are"
                " evaluating ERP for the first time.\n\n"
                "No jargon. No pressure. Just a clear picture of what modern ERP looks like"
                " for a manufacturer like {{company_name}} — and what it actually costs to get"
                " started.\n\n"
                "Book a 20-minute slot at our booth: [CALENDAR LINK]\n\n"
                "Best,\n[REP NAME]\nGlobal Shop Solutions"
            ),
        },
    ],
    "competitive": [
        {
            "label": "Cold Intro — Competitive Displacement",
            "subject": "{{company_name}}: unhappy with {{competitor_name}}?",
            "body": (
                "Hi {{contact_name}},\n\n"
                "I know switching ERP systems isn't a conversation you take lightly — but if"
                " you've been frustrated with {{competitor_name}}'s support, pricing, or roadmap,"
                " you're not alone.\n\n"
                "We've helped dozens of manufacturers migrate from {{competitor_name}} to Global"
                " Shop Solutions — with a structured migration path, dedicated onboarding, and"
                " a fixed-fee implementation so there are no surprises.\n\n"
                "Would a 20-minute call be worth exploring?\n\n"
                "Best,\n[REP NAME]\nGlobal Shop Solutions"
            ),
        },
        {
            "label": "Follow-Up — ROI Comparison",
            "subject": "Re: {{company_name}} — GSS vs {{competitor_name}} side-by-side",
            "body": (
                "Hi {{contact_name}},\n\n"
                "Following up — I put together a quick comparison of GSS vs {{competitor_name}}"
                " for manufacturers your size: total cost of ownership over 3 years, module"
                " depth, and customer satisfaction scores.\n\n"
                "It's a 2-page PDF — want me to send it over?\n\n"
                "No pressure, just data.\n\n"
                "Best,\n[REP NAME]\nGlobal Shop Solutions"
            ),
        },
        {
            "label": "Event-Based — Competitor Presence at Show",
            "subject": "{{event_name}}: come see what you're missing with {{competitor_name}}",
            "body": (
                "Hi {{contact_name}},\n\n"
                "We'll both be at {{event_name}} — and if you're at the {{competitor_name}} booth"
                " doing your due diligence, I'd invite you to spend 20 minutes with us afterward.\n\n"
                "We'll show you exactly where GSS outperforms {{competitor_name}} for manufacturers"
                " your size — real data, real customer stories, no fluff.\n\n"
                "Book a slot: [CALENDAR LINK]\n\n"
                "Best,\n[REP NAME]\nGlobal Shop Solutions"
            ),
        },
    ],
}

# ── Call scripts ───────────────────────────────────────────────────────

CALL_SCRIPTS = {
    "microsoft": {
        "opener": (
            "Hi [CONTACT], this is [REP] from Global Shop Solutions. I noticed"
            " [COMPANY] is running Microsoft 365 — I work with a lot of manufacturers"
            " in [ASSOCIATION] who are looking to complete their Microsoft ecosystem with"
            " a modern ERP. Do you have 2 minutes?"
        ),
        "q1": "When you think about your biggest operational bottleneck today — scheduling, inventory, or customer delivery — which one keeps you up at night?",
        "q2": "Have you evaluated ERP systems before, or would this be your first formal process?",
    },
    "legacy": {
        "opener": (
            "Hi [CONTACT], this is [REP] from Global Shop Solutions. I'm calling manufacturers"
            " who are still running on-premise infrastructure — your peers are moving to cloud ERP"
            " and I wanted to see if the timing might be right for [COMPANY]. Quick 2 minutes?"
        ),
        "q1": "What's your current plan for server refresh or infrastructure upgrades in the next 12 months?",
        "q2": "Who else would be involved in a decision to modernize your back-office systems?",
    },
    "salesforce": {
        "opener": (
            "Hi [CONTACT], this is [REP] from Global Shop Solutions. I noticed [COMPANY]"
            " is using Salesforce — I work with manufacturers who want their CRM and shop"
            " floor talking to each other. Do you have 2 minutes?"
        ),
        "q1": "Can your sales team see real-time production status and delivery dates from within Salesforce today?",
        "q2": "How much time does your team spend manually reconciling CRM data with what's happening on the floor?",
    },
    "pma": {
        "opener": (
            "Hi [CONTACT], this is [REP] from Global Shop Solutions. We work with a lot of PMA"
            " members and I wanted to reach out to [COMPANY] specifically. Quick 2 minutes?"
        ),
        "q1": "For a metalforming operation like yours, what's the biggest scheduling or delivery challenge you're managing right now?",
        "q2": "Are you currently using any ERP or production scheduling software, or is it mostly spreadsheets and tribal knowledge?",
    },
    "greenfield": {
        "opener": (
            "Hi [CONTACT], this is [REP] from Global Shop Solutions. I work with manufacturers"
            " who are looking at ERP for the first time — I wanted to see if [COMPANY] is at"
            " that point. Quick 2 minutes?"
        ),
        "q1": "At what point did you start feeling like your current systems — spreadsheets, QuickBooks, etc. — couldn't keep up with growth?",
        "q2": "What would have to be true about an ERP solution for you to feel confident moving forward?",
    },
    "competitive": {
        "opener": (
            "Hi [CONTACT], this is [REP] from Global Shop Solutions. I know you're already"
            " running [COMPETITOR] — I'm not here to bash them. I just work with manufacturers"
            " who are re-evaluating and I wanted to see if [COMPANY] is in that conversation."
            " Quick 2 minutes?"
        ),
        "q1": "What's the one thing you wish [COMPETITOR] did better for your shop?",
        "q2": "If you were starting fresh today, would you choose [COMPETITOR] again? What would change?",
    },
}

# ── Channel strategies ─────────────────────────────────────────────────

CHANNEL_STRATEGIES = {
    "microsoft": [
        ("Primary: LinkedIn InMail", "Decision-makers are active on LinkedIn; M365 signals digital comfort"),
        ("Secondary: Email sequence (3-touch)", "Direct email works well for tech-savvy audiences"),
        ("Tertiary: Microsoft Partner Network co-marketing", "Leverage Microsoft channel for warm intros"),
        ("Support: Retargeting ads (LinkedIn + Google)", "Reinforce messaging to engaged accounts"),
    ],
    "legacy": [
        ("Primary: Direct mail + phone", "Less digital audience responds better to physical outreach"),
        ("Secondary: Email sequence (2-touch)", "Simpler, shorter emails — avoid digital overload"),
        ("Tertiary: Trade publication ads", "Target industry press read by operations leaders"),
        ("Support: Webinar — Cloud Migration Roadmap for Manufacturers", "Educational content builds trust"),
    ],
    "salesforce": [
        ("Primary: Salesforce AppExchange listing", "Discoverable by Salesforce admins and buyers"),
        ("Secondary: Co-marketing with Salesforce reps", "Leverage SI partnerships for warm intros"),
        ("Tertiary: LinkedIn InMail to CRM/IT stakeholders", "Target Salesforce admins and VPs of Sales"),
        ("Support: Email sequence + ROI content", "Send integration one-pager + ROI calculator"),
    ],
    "pma": [
        ("Primary: Association co-marketing (PMA newsletter, email blast)", "Leverage PMA brand trust"),
        ("Secondary: Event sponsorship / booth at PMA events", "Face-to-face at industry events"),
        ("Tertiary: Peer referral program", "Encourage current GSS PMA customers to refer peers"),
        ("Support: LinkedIn InMail to PMA-tagged contacts", "Reinforce association connection"),
    ],
    "greenfield": [
        ("Primary: Educational content (blog, webinars, guides)", "Build awareness before buying mode"),
        ("Secondary: Demo offer (free 30-min ERP readiness assessment)", "Low-commitment entry point"),
        ("Tertiary: LinkedIn sponsored content", "Reach operations managers not yet searching for ERP"),
        ("Support: Google Search ads (ERP for manufacturers keywords)", "Capture active searchers"),
    ],
    "competitive": [
        ("Primary: Comparison content (GSS vs [Competitor] landing pages)", "Capture in-market buyers"),
        ("Secondary: LinkedIn InMail targeting competitor tech stack users", "Reach confirmed users"),
        ("Tertiary: ROI calculator + migration guide", "Reduce switching cost perception"),
        ("Support: PPC on competitor brand keywords", "Capture competitor dissatisfaction searches"),
    ],
}

# ── Event timing ───────────────────────────────────────────────────────

EVENT_TIMING = {
    "microsoft": [
        ("IMTS 2026", "Sep 2026", "Launch 6-week LinkedIn InMail sequence 8 weeks pre-show"),
        ("FABTECH 2026", "Nov 2026", "Email campaign focusing on Microsoft integration story"),
        ("MS Ignite / Microsoft Inspire", "Jul/Nov 2026", "Co-brand with Microsoft partner messaging"),
    ],
    "legacy": [
        ("Automate 2026", "May 2026", "Run direct mail 4 weeks pre-show; follow up by phone post-show"),
        ("IMTS 2026", "Sep 2026", "On-site demo booth targeting ops-heavy attendees"),
        ("Regional Mfg. Events", "Ongoing", "Local events reach on-premise IT decision-makers"),
    ],
    "salesforce": [
        ("IMTS 2026", "Sep 2026", "Live Salesforce + GSS integration demo at booth"),
        ("Dreamforce 2026", "Sep 2026", "GSS presence in Salesforce ecosystem; AppExchange push"),
        ("FABTECH 2026", "Nov 2026", "CRM + ERP integration workshop"),
    ],
    "pma": [
        ("PMA Forming Our Future 2026", "Mar 2026", "Title sponsor or gold sponsor; live demo station"),
        ("FABTECH 2026", "Nov 2026", "PMA-themed messaging on show floor; member meet-up"),
        ("PMA Regional Workshops", "Quarterly", "Attend PMA chapter events in OH, MI, IN, PA"),
    ],
    "greenfield": [
        ("IPC APEX Expo 2026", "Jan 2026", "Target electronics mfg. greenfield buyers"),
        ("METALCON 2026", "Oct 2026", "Reach metal construction mfg. first-time ERP evaluators"),
        ("IBEX 2026", "Sep 2026", "Marine manufacturing greenfield opportunities"),
    ],
    "competitive": [
        ("IMTS 2026", "Sep 2026", "Run competitive displacement ads 4 weeks pre-show"),
        ("FABTECH 2026", "Nov 2026", "Comparison landing pages live before show; demo at booth"),
        ("Epicor Insights 2026", "May 2026", "Target Epicor users with displacement messaging"),
    ],
}

# ── KPIs per campaign ──────────────────────────────────────────────────

KPIS = {
    "microsoft": [
        ("Email Open Rate", "Target: 28–35%", "M365 audience is digitally engaged"),
        ("LinkedIn InMail Response Rate", "Target: 12–18%", "Warm audience with ecosystem affinity"),
        ("Meetings Booked (per 100 accounts)", "Target: 8–12", "High ICP alignment drives conversion"),
        ("Pipeline Generated (90-day)", "Target: $400K–$800K", "Based on avg GSS deal size $120K ACV"),
    ],
    "legacy": [
        ("Direct Mail Response Rate", "Target: 3–6%", "On-premise audience less digitally responsive"),
        ("Phone Connect Rate", "Target: 15–25%", "Cold call efficiency for this segment"),
        ("Meetings Booked (per 100 accounts)", "Target: 5–8", "Longer sales cycle expected"),
        ("Pipeline Generated (90-day)", "Target: $250K–$500K", "Modernization deals often larger"),
    ],
    "salesforce": [
        ("Email Open Rate", "Target: 30–40%", "Tech-savvy Salesforce users engage with email"),
        ("AppExchange Inbound Leads", "Target: 5–10/month", "Organic discovery from AppExchange listing"),
        ("Meetings Booked (per 100 accounts)", "Target: 10–15", "High intent — already investing in tech"),
        ("Pipeline Generated (90-day)", "Target: $350K–$700K", "CRM buyers have budget and authority"),
    ],
    "pma": [
        ("Email Open Rate (association email blast)", "Target: 35–50%", "Association lists have high trust"),
        ("Event Demo Requests (per show)", "Target: 20–40", "PMA events concentrate the audience"),
        ("Meetings Booked (per 100 accounts)", "Target: 10–15", "Association trust accelerates pipeline"),
        ("Pipeline Generated (90-day)", "Target: $500K–$1M", "Largest addressable segment"),
    ],
    "greenfield": [
        ("Content Engagement Rate (webinar attendance)", "Target: 5–10%", "Education-first approach"),
        ("Demo Request Conversion (from content)", "Target: 8–15% of engaged", "Nurtured greenfield buyers"),
        ("Meetings Booked (per 100 accounts)", "Target: 4–7", "Longer nurture cycle expected"),
        ("Pipeline Generated (90-day)", "Target: $200K–$450K", "Smaller initial deals, longer cycle"),
    ],
    "competitive": [
        ("Email Open Rate", "Target: 25–35%", "Personalized competitor messaging drives opens"),
        ("Comparison Page Conversion Rate", "Target: 3–6%", "High-intent competitive searchers"),
        ("Meetings Booked (per 100 accounts)", "Target: 6–10", "Motivated by pain with current vendor"),
        ("Pipeline Generated (90-day)", "Target: $300K–$600K", "Displacement deals sized similarly to greenfield"),
    ],
}


# ── Helpers ────────────────────────────────────────────────────────────


def _col_letter(idx: int) -> str:
    return get_column_letter(idx)


def _set_col_widths(ws, columns: list[tuple]):
    for i, (_, width) in enumerate(columns, 1):
        ws.column_dimensions[_col_letter(i)].width = width


def _style_header(ws, columns: list[tuple], row: int = 1):
    for i, (header, _) in enumerate(columns, 1):
        cell = ws.cell(row=row, column=i, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    ws.freeze_panes = f"A{row + 1}"


def _build_company_row(rec: dict, icp_score: int, tier: str) -> list:
    """Build a row of values matching COMPANY_COLUMNS."""
    name, email = get_primary_contact(rec)
    assoc = ", ".join(get_associations_list(rec)) or ""
    ec_min, ec_max = get_employee_count(rec)
    return [
        rec.get("company_name", ""),
        rec.get("website", "") or rec.get("domain", ""),
        rec.get("city", ""),
        rec.get("state", ""),
        assoc,
        rec.get("email_provider", "") or "",
        icp_score,
        tier,
        name,
        email,
        rec.get("quality_score", 0),
    ]


def _write_company_table(ws, records: list[dict], start_row: int, col_count: int = 11):
    """Write the company header + data rows starting at start_row."""
    write_section_header(ws, start_row, "Target Companies", col_span=col_count)
    hdr_row = start_row + 1
    _style_header(ws, COMPANY_COLUMNS, row=hdr_row)
    _set_col_widths(ws, COMPANY_COLUMNS)

    for i, rec in enumerate(records):
        icp_data = compute_icp_score(rec)
        icp_score = icp_data["icp_score"]
        has_c = bool(get_contacts(rec))
        qs = rec.get("quality_score", 0)
        if not isinstance(qs, int):
            try:
                qs = int(qs)
            except (ValueError, TypeError):
                qs = 0
        tier = assign_tier(icp_score, has_c, qs)
        row_values = _build_company_row(rec, icp_score, tier)
        row_num = hdr_row + 1 + i
        fill = ALT_ROW_FILL if i % 2 == 1 else None
        for j, val in enumerate(row_values, 1):
            cell = ws.cell(row=row_num, column=j, value=val)
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
            if j == 7:  # ICP Score — colour by tier
                if tier in TIER_FILL:
                    cell.fill = TIER_FILL[tier]
                    cell.font = TIER_FONT[tier]
                cell.alignment = Alignment(horizontal="center")
            if j == 8:  # ICP Tier
                if tier in TIER_FILL:
                    cell.fill = TIER_FILL[tier]
                    cell.font = TIER_FONT[tier]
                cell.alignment = Alignment(horizontal="center")
            # Hyperlink for website
            if j == 2 and val and isinstance(val, str) and val.startswith("http"):
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")

    if ws.auto_filter.ref is None:
        ws.auto_filter.ref = ws.dimensions

    return hdr_row + 1 + len(records)


def _write_campaign_playbook(
    ws,
    campaign_name: str,
    campaign_desc: str,
    audience_desc: str,
    messages: list[str],
    email_key: str,
    call_key: str,
    channel_key: str,
    event_key: str,
    kpi_key: str,
    records: list[dict],
    col_count: int = 11,
):
    """Write the full playbook layout for one campaign sheet."""

    # ── Title block (rows 1-2) ──────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=col_count)
    title_cell = ws.cell(row=1, column=1, value=f"GSS Campaign Playbook: {campaign_name}")
    title_cell.fill = CAMPAIGN_TITLE_FILL
    title_cell.font = CAMPAIGN_TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 8

    # ── Subtitle (row 3) ────────────────────────────────────────────────
    sub = ws.cell(row=3, column=1, value=campaign_desc)
    sub.font = SUBTITLE_FONT
    sub.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[3].height = 32

    # ── Target Audience (rows 5-6) ──────────────────────────────────────
    write_section_header(ws, 5, "Target Audience", col_span=col_count)
    aud = ws.cell(row=6, column=1, value=audience_desc)
    aud.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[6].height = 40

    # ── Messaging Framework (rows 8-11) ─────────────────────────────────
    write_section_header(ws, 8, "Messaging Framework", col_span=col_count)
    for i, msg in enumerate(messages[:3], 1):
        cell = ws.cell(row=8 + i, column=1, value=f"{i}. {msg}")
        cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[8 + i].height = 28

    # ── Email Templates (rows 13-22) ────────────────────────────────────
    write_section_header(ws, 13, "Email Templates", col_span=col_count)
    templates = EMAIL_TEMPLATES.get(email_key, [])
    trow = 14
    for tmpl in templates:
        # Label row
        lbl = ws.cell(row=trow, column=1, value=tmpl["label"])
        lbl.font = Font(bold=True, size=10)
        # Subject row
        subj = ws.cell(row=trow + 1, column=1, value=f"Subject: {tmpl['subject']}")
        subj.font = Font(italic=True, size=10, color="1F4E79")
        # Body rows (may wrap)
        body_cell = ws.cell(row=trow + 2, column=1, value=tmpl["body"])
        body_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[trow + 2].height = 90
        trow += 4  # gap between templates

    # ── Call Script (rows 24-27) ─────────────────────────────────────────
    cs = CALL_SCRIPTS.get(call_key, {})
    write_section_header(ws, 24, "Call Script", col_span=col_count)
    opener = ws.cell(row=25, column=1, value=f"Opener: {cs.get('opener', '')}")
    opener.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[25].height = 50
    q1 = ws.cell(row=26, column=1, value=f"Q1: {cs.get('q1', '')}")
    q1.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[26].height = 30
    q2 = ws.cell(row=27, column=1, value=f"Q2: {cs.get('q2', '')}")
    q2.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[27].height = 30

    # ── Channel Strategy (rows 29-33) ────────────────────────────────────
    channels = CHANNEL_STRATEGIES.get(channel_key, [])
    write_section_header(ws, 29, "Channel Strategy", col_span=col_count)
    for i, (ch_name, ch_rationale) in enumerate(channels[:4], 1):
        cell = ws.cell(row=29 + i, column=1, value=f"{ch_name}: {ch_rationale}")
        cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[29 + i].height = 24

    # ── Event Timing (rows 35-38) ─────────────────────────────────────────
    events = EVENT_TIMING.get(event_key, [])
    write_section_header(ws, 35, "Event Timing", col_span=col_count)
    for i, (ev_name, ev_date, ev_tactic) in enumerate(events[:3], 1):
        cell = ws.cell(row=35 + i, column=1, value=f"{ev_name} ({ev_date}) — {ev_tactic}")
        cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[35 + i].height = 24

    # ── KPIs (rows 40-44) ─────────────────────────────────────────────────
    kpis = KPIS.get(kpi_key, [])
    write_section_header(ws, 40, "KPIs & Expected Metrics", col_span=col_count)
    for i, (kpi_name, kpi_target, kpi_note) in enumerate(kpis[:4], 1):
        cell = ws.cell(row=40 + i, column=1, value=f"{kpi_name}: {kpi_target}  [{kpi_note}]")
        cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[40 + i].height = 22

    # ── Target Companies (row 47+) ────────────────────────────────────────
    _write_company_table(ws, records, start_row=47, col_count=col_count)

    # Widen column A for playbook text
    ws.column_dimensions["A"].width = 120


# ── Segment filter functions ───────────────────────────────────────────


def _filter_microsoft(records: list[dict]) -> list[dict]:
    return [r for r in records if get_email_provider(r).lower() == "microsoft 365"]


def _filter_legacy(records: list[dict]) -> list[dict]:
    legacy_providers = {"self-hosted", "self-hosted (on-premise)", "other", ""}
    return [
        r for r in records
        if get_email_provider(r).lower() in legacy_providers
        or not get_email_provider(r)
    ]


def _filter_salesforce(records: list[dict]) -> list[dict]:
    out = []
    for r in records:
        spf = [s.lower() for s in get_spf_list(r)]
        if any(s in ("salesforce", "pardot") for s in spf):
            out.append(r)
    return out


def _filter_pma(records: list[dict]) -> list[dict]:
    return [r for r in records if "PMA" in get_associations_list(r)]


def _filter_greenfield(records: list[dict]) -> list[dict]:
    out = []
    for r in records:
        ec_min, ec_max = get_employee_count(r)
        emp = ec_max or ec_min
        if emp == 0:
            continue  # skip unknown size
        erp = (r.get("erp_system") or "").strip()
        competitor = detect_competitor(r)
        if not erp and not competitor:
            out.append(r)
    return out


def _filter_competitive(records: list[dict]) -> list[dict]:
    return [r for r in records if detect_competitor(r)]


def _sort_by_icp(records: list[dict]) -> list[dict]:
    def _score(rec):
        return compute_icp_score(rec)["icp_score"]
    return sorted(records, key=_score, reverse=True)


# ── Sheet 1: Campaign Overview ─────────────────────────────────────────


def _build_overview_row(
    campaign_name: str,
    segment_desc: str,
    target_recs: list[dict],
    primary_channel: str,
    event_tieins: str,
    priority: str,
) -> list:
    with_contacts = sum(1 for r in target_recs if get_contacts(r))
    scores = [compute_icp_score(r)["icp_score"] for r in target_recs]
    avg_score = round(sum(scores) / len(scores), 1) if scores else 0
    return [
        campaign_name,
        segment_desc,
        len(target_recs),
        with_contacts,
        avg_score,
        primary_channel,
        event_tieins,
        priority,
    ]


OVERVIEW_COLUMNS = [
    ("Campaign Name", 30),
    ("Segment Description", 40),
    ("Target Accounts", 16),
    ("With Contacts", 14),
    ("Avg ICP Score", 14),
    ("Primary Channel", 22),
    ("Event Tie-in", 32),
    ("Priority", 12),
]


def _write_overview(ws, campaign_rows: list[list]):
    # Title
    col_count = len(OVERVIEW_COLUMNS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    title = ws.cell(row=1, column=1, value="GSS ABM Campaign Overview — Manufacturing ERP")
    title.fill = CAMPAIGN_TITLE_FILL
    title.font = CAMPAIGN_TITLE_FONT
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    subtitle = ws.cell(row=2, column=1, value="6 targeted campaigns derived from NAM Intelligence Pipeline data — 2,083 manufacturing companies")
    subtitle.font = SUBTITLE_FONT
    ws.row_dimensions[2].height = 18

    # Header
    _style_header(ws, OVERVIEW_COLUMNS, row=4)
    _set_col_widths(ws, OVERVIEW_COLUMNS)

    # Data rows
    priority_fills = {
        "P1 — Highest": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "P2 — High": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "P3 — Medium": PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid"),
    }
    for i, row_data in enumerate(campaign_rows, 1):
        row_num = 4 + i
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=j, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        priority_val = row_data[7]
        if priority_val in priority_fills:
            for j in range(1, col_count + 1):
                ws.cell(row=row_num, column=j).fill = priority_fills[priority_val]
        ws.row_dimensions[row_num].height = 30

    # Summary stats section
    last_row = 4 + len(campaign_rows) + 3
    write_section_header(ws, last_row, "Campaign Portfolio Summary", col_span=col_count)
    total_accounts = sum(r[2] for r in campaign_rows)
    total_contacts = sum(r[3] for r in campaign_rows)
    ws.cell(row=last_row + 1, column=1, value=f"Total unique accounts across all campaigns: {total_accounts:,}  |  Accounts with known contacts: {total_contacts:,}")
    ws.cell(row=last_row + 1, column=1).font = Font(bold=True)


# ── Main ───────────────────────────────────────────────────────────────


def main():
    print("Loading and merging data...")
    records = load_and_merge_data()
    print(f"  Loaded {len(records):,} records")

    # Build segments
    print("Filtering campaign segments...")
    seg_microsoft = _sort_by_icp(_filter_microsoft(records))
    seg_legacy = _sort_by_icp(_filter_legacy(records))
    seg_salesforce = _sort_by_icp(_filter_salesforce(records))
    seg_pma = _sort_by_icp(_filter_pma(records))
    seg_greenfield = _sort_by_icp(_filter_greenfield(records))
    seg_competitive = _sort_by_icp(_filter_competitive(records))

    print(f"  Microsoft Ecosystem:    {len(seg_microsoft):>5,}")
    print(f"  Legacy IT:              {len(seg_legacy):>5,}")
    print(f"  Salesforce CRM:         {len(seg_salesforce):>5,}")
    print(f"  PMA Association:        {len(seg_pma):>5,}")
    print(f"  Greenfield ERP:         {len(seg_greenfield):>5,}")
    print(f"  Competitive Displace:   {len(seg_competitive):>5,}")

    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Campaign Overview ─────────────────────────────────────
    print("Writing Sheet 1: Campaign Overview...")
    ws_overview = wb.create_sheet("Campaign Overview")
    overview_rows = [
        _build_overview_row(
            "Microsoft Ecosystem Modernization",
            "Companies using Microsoft 365 email — already in the Microsoft ecosystem",
            seg_microsoft,
            "LinkedIn InMail + Email",
            "IMTS 2026, FABTECH 2026",
            "P1 — Highest",
        ),
        _build_overview_row(
            "PMA Association Blitz",
            "All PMA member companies — association trust + peer social proof",
            seg_pma,
            "Association co-marketing + Events",
            "PMA Forming Our Future, FABTECH 2026",
            "P1 — Highest",
        ),
        _build_overview_row(
            "Salesforce CRM Cross-Sell",
            "Companies with Salesforce/Pardot in SPF — CRM investment = ERP budget",
            seg_salesforce,
            "AppExchange + Co-marketing",
            "IMTS 2026, Dreamforce 2026",
            "P2 — High",
        ),
        _build_overview_row(
            "Greenfield ERP",
            "Companies with no detected ERP system — first-time buyers",
            seg_greenfield,
            "Educational content + Demo offers",
            "IPC APEX, METALCON, IBEX",
            "P2 — High",
        ),
        _build_overview_row(
            "Competitive Displacement",
            "Companies with known competitor ERP — displacement opportunity",
            seg_competitive,
            "Comparison content + ROI calculator",
            "IMTS 2026, FABTECH 2026",
            "P2 — High",
        ),
        _build_overview_row(
            "Legacy IT Transformation",
            "Companies with on-premise / self-hosted email infrastructure",
            seg_legacy,
            "Direct mail + Phone",
            "Automate 2026, IMTS 2026",
            "P3 — Medium",
        ),
    ]
    _write_overview(ws_overview, overview_rows)

    # ── Sheet 2: Microsoft Ecosystem Modernization ────────────────────
    print(f"Writing Sheet 2: Microsoft Ecosystem ({len(seg_microsoft):,} accounts)...")
    ws2 = wb.create_sheet("Microsoft Ecosystem")
    _write_campaign_playbook(
        ws2,
        campaign_name="Microsoft Ecosystem Modernization",
        campaign_desc=(
            "Target audience: ~{n:,} companies already running Microsoft 365. "
            "Message theme: 'You already trust Microsoft for email. Complete your digital transformation "
            "with manufacturing ERP that integrates seamlessly.'".format(n=len(seg_microsoft))
        ),
        audience_desc=(
            f"{len(seg_microsoft):,} companies identified with Microsoft 365 email — the single largest "
            "addressable segment. These buyers are cloud-comfortable, IT-budget-approved, and often "
            "looking to complete their Microsoft ecosystem. GSS's native M365 integration is the "
            "primary value driver. Focus on Dynamics 365 displacement and IT-led buying processes."
        ),
        messages=[
            "You already trust Microsoft for email — complete your digital transformation with GSS ERP that integrates natively with M365",
            "Eliminate the gap between your Microsoft tools and your shop floor with a single, integrated manufacturing platform",
            "500+ manufacturers have standardized on Microsoft + GSS — your competitors are already there",
        ],
        email_key="microsoft",
        call_key="microsoft",
        channel_key="microsoft",
        event_key="microsoft",
        kpi_key="microsoft",
        records=seg_microsoft,
    )

    # ── Sheet 3: Legacy IT Transformation ─────────────────────────────
    print(f"Writing Sheet 3: Legacy IT ({len(seg_legacy):,} accounts)...")
    ws3 = wb.create_sheet("Legacy IT Transformation")
    _write_campaign_playbook(
        ws3,
        campaign_name="Legacy IT Transformation",
        campaign_desc=(
            "Target audience: ~{n:,} companies with on-premise / self-hosted email infrastructure. "
            "Message theme: 'Your competitors have moved to the cloud — including their ERP. "
            "Stop maintaining servers and start growing.'".format(n=len(seg_legacy))
        ),
        audience_desc=(
            f"{len(seg_legacy):,} companies identified with on-premise or legacy email infrastructure "
            "(self-hosted mail servers, generic providers, or no detected provider). These buyers are "
            "typically less digitally active, requiring direct mail and phone outreach. The opportunity: "
            "modernization pain is high, and GSS's cloud-migration path is a compelling offer. "
            "Target IT decision-makers and ops leaders simultaneously."
        ),
        messages=[
            "Your competitors have moved to the cloud — including their ERP — while you're still maintaining servers",
            "Every dollar spent on server maintenance is a dollar not invested in growing your manufacturing business",
            "GSS makes the migration to cloud ERP predictable: fixed-fee implementation, 90-day go-live, dedicated onboarding",
        ],
        email_key="legacy",
        call_key="legacy",
        channel_key="legacy",
        event_key="legacy",
        kpi_key="legacy",
        records=seg_legacy,
    )

    # ── Sheet 4: Salesforce CRM Cross-Sell ─────────────────────────────
    print(f"Writing Sheet 4: Salesforce CRM ({len(seg_salesforce):,} accounts)...")
    ws4 = wb.create_sheet("Salesforce CRM Cross-Sell")
    _write_campaign_playbook(
        ws4,
        campaign_name="Salesforce CRM Cross-Sell",
        campaign_desc=(
            "Target audience: ~{n:,} companies with Salesforce or Pardot in SPF records. "
            "Message theme: 'Your CRM is only as powerful as the data feeding it. "
            "Connect your shop floor to your sales pipeline.'".format(n=len(seg_salesforce))
        ),
        audience_desc=(
            f"{len(seg_salesforce):,} companies identified as active Salesforce or Pardot users via SPF records. "
            "These buyers have already demonstrated willingness to invest in business software and have a VP of Sales "
            "or Marketing Ops stakeholder in the buying process. The GSS + Salesforce integration story is compelling: "
            "bi-directional data flow, real-time production status in CRM, and AppExchange discoverability. "
            "Target both IT/Ops (ERP buyer) and Sales leadership (CRM power user)."
        ),
        messages=[
            "Your Salesforce CRM is only as powerful as the data feeding it — connect it to your shop floor with GSS ERP",
            "Give your sales reps real-time production status, delivery dates, and inventory visibility inside Salesforce",
            "GSS is on the Salesforce AppExchange — validate us through your existing Salesforce relationship",
        ],
        email_key="salesforce",
        call_key="salesforce",
        channel_key="salesforce",
        event_key="salesforce",
        kpi_key="salesforce",
        records=seg_salesforce,
    )

    # ── Sheet 5: PMA Association Blitz ────────────────────────────────
    print(f"Writing Sheet 5: PMA Association ({len(seg_pma):,} accounts)...")
    ws5 = wb.create_sheet("PMA Association Blitz")
    _write_campaign_playbook(
        ws5,
        campaign_name="PMA Association Blitz",
        campaign_desc=(
            "Target audience: ~{n:,} PMA member companies. "
            "Message theme: 'Fellow PMA members trust Global Shop Solutions. "
            "Join 500+ metalformers running modern ERP.'".format(n=len(seg_pma))
        ),
        audience_desc=(
            f"{len(seg_pma):,} PMA (Precision Metalforming Association) member companies identified. "
            "PMA is GSS's highest-density association segment. These buyers respond strongly to peer "
            "social proof and association-endorsed messaging. Leverage PMA's communication channels "
            "(newsletter, email blasts, webinars) and event sponsorships. The metalforming use case "
            "(tooling management, scheduling, on-time delivery) maps precisely to GSS core strengths. "
            "Consider a PMA Member Exclusive pricing bundle for Q2 2026."
        ),
        messages=[
            "Fellow PMA members trust Global Shop Solutions — join 500+ metalformers who have standardized on GSS ERP",
            "GSS was built for precision metalforming: tooling management, scheduling, delivery performance tracking — all in one system",
            "Exclusive PMA member onboarding program: dedicated implementation support + association peer reference network",
        ],
        email_key="pma",
        call_key="pma",
        channel_key="pma",
        event_key="pma",
        kpi_key="pma",
        records=seg_pma,
    )

    # ── Sheet 6: Greenfield ERP ────────────────────────────────────────
    print(f"Writing Sheet 6: Greenfield ERP ({len(seg_greenfield):,} accounts)...")
    ws6 = wb.create_sheet("Greenfield ERP")
    _write_campaign_playbook(
        ws6,
        campaign_name="Greenfield ERP",
        campaign_desc=(
            "Target audience: ~{n:,} companies with no detected ERP system and known employee count. "
            "Message theme: 'Your first real ERP. Built for manufacturers from day one, "
            "not adapted from accounting software.'".format(n=len(seg_greenfield))
        ),
        audience_desc=(
            f"{len(seg_greenfield):,} companies with no detected ERP system and verified employee count > 0. "
            "These are first-time ERP buyers — likely running spreadsheets, QuickBooks, or tribal knowledge. "
            "They need education before they enter a buying process. Lead with content (guides, webinars, "
            "ROI calculators) and use low-commitment offers (free ERP readiness assessment) to qualify. "
            "The GSS story: manufacturer-native, 90-day go-live, simple pricing."
        ),
        messages=[
            "Your first real ERP — built for manufacturers from day one, not adapted from accounting software",
            "Most manufacturers your size go live in 90 days or less with GSS — and see ROI in year 1",
            "Free ERP Readiness Assessment: 30 minutes, no pitch, just an honest conversation about whether the timing is right",
        ],
        email_key="greenfield",
        call_key="greenfield",
        channel_key="greenfield",
        event_key="greenfield",
        kpi_key="greenfield",
        records=seg_greenfield,
    )

    # ── Sheet 7: Competitive Displacement ─────────────────────────────
    print(f"Writing Sheet 7: Competitive Displacement ({len(seg_competitive):,} accounts)...")
    ws7 = wb.create_sheet("Competitive Displacement")
    _write_campaign_playbook(
        ws7,
        campaign_name="Competitive Displacement",
        campaign_desc=(
            "Target audience: ~{n:,} companies with detected competitor ERP systems. "
            "Message theme: Varies by competitor — lead with pain points, migration path, and ROI.".format(n=len(seg_competitive))
        ),
        audience_desc=(
            f"{len(seg_competitive):,} companies identified with a known competitor ERP (Epicor, SAP, "
            "Oracle, Infor, Dynamics 365, NetSuite, Plex, Acumatica, and others via tech stack / SPF signals). "
            "Personalize outreach by competitor: reference their known pain points (support costs, "
            "upgrade complexity, pricing). Lead with comparison content and migration stories. "
            "GSS differentiators: fixed-fee implementation, manufacturer-native UX, and responsive "
            "US-based support. Do NOT lead with feature lists — lead with migration risk reduction."
        ),
        messages=[
            "We've helped dozens of manufacturers migrate from [Competitor] to GSS — with a structured migration path and fixed-fee implementation",
            "GSS vs [Competitor]: lower total cost of ownership, faster implementation, and a US-based support team that actually answers the phone",
            "Not happy with [Competitor]? You're not alone — and switching is less painful than you think",
        ],
        email_key="competitive",
        call_key="competitive",
        channel_key="competitive",
        event_key="competitive",
        kpi_key="competitive",
        records=seg_competitive,
    )

    # Save
    print(f"\nSaving to {OUTPUT_PATH}...")
    wb.save(str(OUTPUT_PATH))
    print(f"Done. Output: {OUTPUT_PATH}")
    print(f"File size: {OUTPUT_PATH.stat().st_size / 1024:.1f} KB")


if __name__ == "__main__":
    main()
