#!/usr/bin/env python3
"""
ABM Executive Briefing
NAM Intelligence Pipeline

Generates: data/exports/GSS_Executive_Briefing.xlsx
  Sheet 1: Title & Overview (KPI dashboard)
  Sheet 2: Market Opportunity (TAM + association matrix)
  Sheet 3: Data Quality Scorecard (grade + enrichment coverage)
  Sheet 4: Top 25 Target Accounts (money slide)
  Sheet 5: Competitive Threats (competitor overview + insights)
  Sheet 6: Investment Recommendations (30/60/90-day action plan)
"""

import sys
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from scripts.abm_shared import (
    BASE_DIR,
    EXPORTS_DIR,
    GRADE_FILLS,
    HEADER_ALIGNMENT,
    HEADER_FILL,
    HEADER_FONT,
    METRIC_FONT,
    METRIC_LARGE_FONT,
    SECTION_FILL,
    SECTION_FONT,
    SUBTITLE_FONT,
    THIN_BORDER,
    TIER_FILLS,
    TIER_FONTS,
    THREAT_FILLS,
    THREAT_RESPONSE,
    WHITE_FONT,
    assign_tier,
    compute_icp_score,
    detect_competitor,
    get_associations_list,
    get_contacts,
    get_email_provider,
    get_employee_count,
    get_primary_contact,
    get_spf_list,
    get_tech_stack,
    has_website,
    load_and_merge_data,
    load_associations_config,
    load_competitors,
    load_events,
    style_header_row,
    write_section_header,
)

OUTPUT_PATH = EXPORTS_DIR / "GSS_Executive_Briefing.xlsx"

# ── Shared styles ─────────────────────────────────────────────────────

_TITLE_FONT = Font(bold=True, size=28, color="1F4E79")
_KPI_LABEL_FONT = Font(bold=True, size=12, color="1F4E79")
_KPI_VALUE_FONT = Font(bold=True, size=22, color="1F4E79")
_KPI_DESC_FONT = Font(italic=True, size=10, color="555555")
_BULLET_FONT = Font(size=11, color="333333")
_SEPARATOR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_GRAY_FILL = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
_BOLD_GREEN_FONT = Font(bold=True, color="006100")
_BOLD_YELLOW_FONT = Font(bold=True, color="7D6608")
_BOLD_RED_FONT = Font(bold=True, color="9C0006")
_BOLD_GRAY_FONT = Font(bold=True, color="555555")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_LEFT_NOWRAP = Alignment(horizontal="left", vertical="center", wrap_text=False)
_IMPACT_FILLS = {
    "High": _GREEN_FILL,
    "Med": _YELLOW_FILL,
    "Low": _GRAY_FILL,
}
_IMPACT_FONTS = {
    "High": _BOLD_GREEN_FONT,
    "Med": _BOLD_YELLOW_FONT,
    "Low": _BOLD_GRAY_FONT,
}
_EFFORT_FILLS = {
    "Low": _GREEN_FILL,
    "Med": _YELLOW_FILL,
    "High": _RED_FILL,
}
_EFFORT_FONTS = {
    "Low": _BOLD_GREEN_FONT,
    "Med": _BOLD_YELLOW_FONT,
    "High": _BOLD_RED_FONT,
}


# ── Utility helpers ───────────────────────────────────────────────────

def _col(idx: int) -> str:
    return get_column_letter(idx)


def _set_row_height(ws, row: int, height: float):
    ws.row_dimensions[row].height = height


def _emp_display(rec: dict) -> str:
    ec_min, ec_max = get_employee_count(rec)
    if ec_min and ec_max and ec_min != ec_max:
        return f"{ec_min:,}-{ec_max:,}"
    if ec_max:
        return f"{ec_max:,}"
    if ec_min:
        return f"{ec_min:,}+"
    return ""


def _recommend_campaign(rec: dict, icp: dict) -> str:
    """Recommend a campaign approach for this company."""
    competitor = detect_competitor(rec)
    ep = get_email_provider(rec).lower()
    dims = [
        ("tech_maturity", icp["tech_maturity"], 25),
        ("size_fit", icp["size_fit"], 20),
        ("geo_fit", icp["geo_fit"], 15),
        ("tech_gap", icp["tech_gap"], 15),
        ("assoc_engagement", icp["assoc_engagement"], 15),
    ]
    dims.sort(key=lambda d: d[1] / d[2], reverse=True)
    top = dims[0][0]

    if competitor:
        return f"Displacement — migrate from {competitor.title()}"
    if top == "tech_gap" and icp["tech_gap"] >= 12:
        return "Greenfield — first ERP value story"
    if top == "tech_maturity" and "microsoft" in ep:
        return "ABM 1:1 — Microsoft ecosystem modernization"
    if top == "tech_maturity" and "self-hosted" in ep:
        return "Digital transformation — legacy modernization"
    if top == "size_fit" and icp["size_fit"] >= 18:
        return "ABM 1:1 — personalized executive outreach"
    if top == "assoc_engagement" and icp["assoc_engagement"] >= 12:
        return "Association-centric — peer trust + event presence"
    if top == "geo_fit" and icp["geo_fit"] >= 12:
        return "Regional cluster campaign"
    return "ABM 1:few — multi-touch nurture"


def _write_separator(ws, row: int, col_count: int = 10):
    """Write a solid dark-blue separator row."""
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _SEPARATOR_FILL
    _set_row_height(ws, row, 6)


def _write_kpi_row(ws, row: int, label: str, value: str, description: str):
    """Write a KPI row: label (A), value (B), description (C)."""
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = _KPI_LABEL_FONT
    lc.alignment = _LEFT_NOWRAP

    vc = ws.cell(row=row, column=2, value=value)
    vc.font = _KPI_VALUE_FONT
    vc.alignment = _LEFT_NOWRAP

    dc = ws.cell(row=row, column=3, value=description)
    dc.font = _KPI_DESC_FONT
    dc.alignment = _LEFT

    _set_row_height(ws, row, 36)


def _table_header(ws, row: int, columns: list[tuple]):
    """Write a standard table header row."""
    for col_idx, (header, width) in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[_col(col_idx)].width = width
    _set_row_height(ws, row, 30)


def _data_cell(ws, row: int, col: int, value, font=None, fill=None,
               alignment=None, border=True):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    else:
        cell.alignment = _LEFT_NOWRAP
    if border:
        cell.border = THIN_BORDER
    return cell


def _write_data_row(ws, row: int, values: list, row_height: float = 18):
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.border = THIN_BORDER
        cell.alignment = _LEFT_NOWRAP
    _set_row_height(ws, row, row_height)


# ── Sheet 1: Title & Overview ─────────────────────────────────────────

def write_title_sheet(wb: Workbook, records: list[dict], scored_data: list[tuple]):
    ws = wb.active
    ws.title = "Executive Overview"

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 40

    # Row 1-2: Merged title
    ws.merge_cells("A1:C2")
    title_cell = ws.cell(row=1, column=1, value="GSS NAM Intelligence Pipeline")
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    _set_row_height(ws, 1, 50)
    _set_row_height(ws, 2, 20)

    # Row 3: Subtitle
    ws.merge_cells("A3:C3")
    sub = ws.cell(row=3, column=1,
                  value="Executive Briefing — Manufacturing ERP Market Intelligence")
    sub.font = Font(bold=True, size=14, color="1F4E79")
    sub.alignment = _LEFT_NOWRAP
    _set_row_height(ws, 3, 24)

    # Row 4: Generated date
    ws.merge_cells("A4:C4")
    gen = ws.cell(row=4, column=1,
                  value=f"Generated: {date.today().strftime('%B %d, %Y')}")
    gen.font = Font(italic=True, size=10, color="888888")
    gen.alignment = _LEFT_NOWRAP
    _set_row_height(ws, 4, 18)

    # Row 5: blank
    _set_row_height(ws, 5, 10)

    # Row 6: separator
    _write_separator(ws, 6, 3)

    # ── KPI computations ──
    total_companies = len(records)
    assoc_config = load_associations_config()
    covered_assocs = set()
    for rec in records:
        for a in get_associations_list(rec):
            covered_assocs.add(a.upper())
    # count assocs that have extracted records
    extracted_assocs = len([k for k in assoc_config if k.upper() in covered_assocs])
    total_assocs = len(assoc_config)

    enriched_count = sum(
        1 for r in records
        if r.get("email_provider") or r.get("domain") or r.get("tech_stack")
    )
    enriched_pct = enriched_count / max(total_companies, 1) * 100

    high_quality = sum(
        1 for r in records
        if (r.get("quality_grade") or "F") in ("A", "B")
    )
    hq_pct = high_quality / max(total_companies, 1) * 100

    contact_count = sum(1 for r in records if get_contacts(r))
    decision_maker_count = sum(len(get_contacts(r)) for r in records)

    avg_qs = sum(r.get("quality_score", 0) or 0 for r in records) / max(total_companies, 1)

    # Row 7: KPI header
    _set_row_height(ws, 7, 8)
    kpi_header = ws.cell(row=7, column=1, value="Pipeline KPIs")
    kpi_header.font = Font(bold=True, size=12, color="1F4E79")
    kpi_header.alignment = _LEFT_NOWRAP

    # Rows 8-13: KPI grid
    kpis = [
        ("Total Companies",
         f"{total_companies:,}",
         "Unique manufacturing companies in pipeline"),
        (f"Associations Covered",
         f"{extracted_assocs} of {total_assocs}",
         "5 of 14 NAM-affiliated associations extracted"),
        ("Enriched Records",
         f"{enriched_count:,} ({enriched_pct:.0f}%)",
         "Companies with email provider, domain, or tech stack data"),
        ("High-Quality (B+ Grade)",
         f"{high_quality:,} ({hq_pct:.0f}%)",
         "Records with quality score ≥ 70 (A or B grade)"),
        ("Decision-Maker Contacts",
         f"{decision_maker_count:,}",
         f"Named contacts across {contact_count:,} companies"),
        ("Avg Quality Score",
         f"{avg_qs:.1f} / 100",
         "Pipeline average (B-grade threshold: 70)"),
    ]

    for i, (label, value, desc) in enumerate(kpis, 8):
        _write_kpi_row(ws, i, label, value, desc)

    # Row 14: separator
    _write_separator(ws, 14, 3)

    # Row 15: Pipeline Status header
    status_cell = ws.cell(row=15, column=1, value="Pipeline Status")
    status_cell.font = Font(bold=True, size=13, color="1F4E79")
    status_cell.alignment = _LEFT_NOWRAP
    _set_row_height(ws, 15, 22)

    # Rows 16-20: bullets
    tier_counts = Counter(t for _, _, t in scored_data)
    bullets = [
        f"  • Extracted {total_companies:,} manufacturing companies across 5 associations "
        f"(PMA, NEMA, AGMA, AIA, SOCMA) — 9 associations remain untapped",
        f"  • {high_quality:,} records ({hq_pct:.0f}%) meet B+ quality threshold; "
        f"pipeline average quality score {avg_qs:.1f}/100",
        f"  • ICP scoring identifies {tier_counts.get('Tier 1', 0)} Tier 1 strategic accounts, "
        f"{tier_counts.get('Tier 2', 0)} Tier 2 growth accounts ready for ABM campaigns",
        f"  • 91% tech stack coverage, 89% MX/email provider detection — "
        f"Microsoft 365 dominates at 662 companies (strong Dynamics displacement signal)",
        f"  • Free OSINT enrichment complete (Wappalyzer, DNS, SPF); "
        f"Clearbit/Apollo API keys will unlock full firmographic depth",
    ]

    for i, bullet in enumerate(bullets, 16):
        ws.merge_cells(f"A{i}:C{i}")
        cell = ws.cell(row=i, column=1, value=bullet)
        cell.font = _BULLET_FONT
        cell.alignment = _LEFT
        _set_row_height(ws, i, 28)

    # Page setup
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "landscape"
    ws.print_area = "A1:C20"

    print(f"  Sheet 'Executive Overview': {len(kpis)} KPIs, {len(bullets)} status bullets")


# ── Sheet 2: Market Opportunity ───────────────────────────────────────

def write_market_opportunity_sheet(wb: Workbook, records: list[dict]):
    ws = wb.create_sheet("Market Opportunity")

    assoc_config = load_associations_config()

    # Count extracted per association
    extracted_counts: dict[str, int] = defaultdict(int)
    for rec in records:
        for a in get_associations_list(rec):
            extracted_counts[a.upper()] += 1

    # ── Section A: Total Addressable Market ──
    row = 1
    write_section_header(ws, row, "A. Total Addressable Market — Association Coverage", 6)
    row += 1

    tam_cols = [
        ("Association", 35),
        ("Expected Members", 16),
        ("Extracted", 12),
        ("Penetration %", 14),
        ("Priority", 10),
        ("Status", 12),
    ]
    _table_header(ws, row, tam_cols)
    row += 1

    total_expected = 0
    total_extracted = 0

    for assoc_key, cfg in assoc_config.items():
        expected = int(cfg.get("expected_members", 0))
        extracted = extracted_counts.get(assoc_key.upper(), 0)
        pct = extracted / max(expected, 1) * 100
        priority = cfg.get("priority", "low").capitalize()
        status = "Active" if extracted > 0 else "Pending"
        total_expected += expected
        total_extracted += extracted

        values = [
            cfg.get("name", assoc_key),
            expected,
            extracted,
            f"{pct:.1f}%",
            priority,
            status,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = _data_cell(ws, row, col_idx, val)
        # Highlight pending associations in yellow
        if extracted == 0:
            for c in range(1, 7):
                ws.cell(row=row, column=c).fill = _YELLOW_FILL
        # Status color
        status_cell = ws.cell(row=row, column=6)
        if status == "Active":
            status_cell.fill = _GREEN_FILL
            status_cell.font = _BOLD_GREEN_FONT
        else:
            status_cell.fill = _YELLOW_FILL
            status_cell.font = _BOLD_YELLOW_FONT
        _set_row_height(ws, row, 18)
        row += 1

    # Total row
    total_pct = total_extracted / max(total_expected, 1) * 100
    total_vals = ["TOTAL", total_expected, total_extracted, f"{total_pct:.1f}%", "", ""]
    for col_idx, val in enumerate(total_vals, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = Font(bold=True, size=11, color="1F4E79")
        cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        cell.border = THIN_BORDER
        cell.alignment = _LEFT_NOWRAP
    _set_row_height(ws, row, 20)
    row += 2

    # ── Section B: Market Size Estimate ──
    write_section_header(ws, row, "B. Market Size Estimate", 6)
    row += 1

    remaining_expected = total_expected - total_extracted
    remaining_assocs = [
        k for k, v in assoc_config.items()
        if extracted_counts.get(k.upper(), 0) == 0
    ]
    remaining_large = sorted(
        [(k, assoc_config[k].get("expected_members", 0)) for k in remaining_assocs],
        key=lambda x: -x[1]
    )

    market_rows = [
        ("Total addressable market (all 14 associations)", f"{total_expected:,} companies"),
        ("Currently extracted", f"{total_extracted:,} companies ({total_pct:.0f}% penetration)"),
        ("Remaining opportunity", f"{remaining_expected:,} companies across {len(remaining_assocs)} associations"),
        (
            "Largest untapped association",
            f"{remaining_large[0][0] if remaining_large else 'N/A'} "
            f"(~{remaining_large[0][1]:,} expected members)" if remaining_large else "N/A",
        ),
        (
            "Growth path",
            f"Extracting remaining {len(remaining_assocs)} associations → estimated +{remaining_expected:,} companies",
        ),
    ]

    for label, value in market_rows:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(bold=True, size=10)
        lc.border = THIN_BORDER
        lc.alignment = _LEFT_NOWRAP
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = Font(size=10)
        vc.border = THIN_BORDER
        vc.alignment = _LEFT_NOWRAP
        # merge B–F for the value cell
        ws.merge_cells(f"B{row}:F{row}")
        _set_row_height(ws, row, 18)
        row += 1

    row += 1

    # ── Section C: Association Priority Matrix ──
    write_section_header(ws, row, "C. Association Priority Matrix", 6)
    row += 1

    priority_cols = [
        ("Association", 35),
        ("Priority", 10),
        ("Expected", 12),
        ("Extracted", 12),
        ("Gap", 10),
        ("Status", 12),
    ]
    _table_header(ws, row, priority_cols)
    row += 1

    # Sort by priority then expected members
    priority_order = {"high": 0, "medium": 1, "low": 2}
    sorted_assocs = sorted(
        assoc_config.items(),
        key=lambda x: (priority_order.get(x[1].get("priority", "low"), 3),
                       -x[1].get("expected_members", 0))
    )

    for assoc_key, cfg in sorted_assocs:
        expected = int(cfg.get("expected_members", 0))
        extracted = extracted_counts.get(assoc_key.upper(), 0)
        gap = max(expected - extracted, 0)
        priority = cfg.get("priority", "low").capitalize()
        status = "Active" if extracted > 0 else "Pending"

        values = [cfg.get("name", assoc_key), priority, expected, extracted, gap, status]
        for col_idx, val in enumerate(values, 1):
            _data_cell(ws, row, col_idx, val)

        # Color the entire row by status
        fill = _GREEN_FILL if status == "Active" else _YELLOW_FILL
        for c in range(1, 7):
            ws.cell(row=row, column=c).fill = fill

        status_font = _BOLD_GREEN_FONT if status == "Active" else _BOLD_YELLOW_FONT
        ws.cell(row=row, column=6).font = status_font
        _set_row_height(ws, row, 18)
        row += 1

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "landscape"
    print(f"  Sheet 'Market Opportunity': {len(assoc_config)} associations, "
          f"TAM {total_expected:,} expected members")


# ── Sheet 3: Data Quality Scorecard ──────────────────────────────────

def write_quality_scorecard_sheet(wb: Workbook, records: list[dict]):
    ws = wb.create_sheet("Data Quality Scorecard")
    total = len(records)

    # ── Section A: Grade Distribution ──
    row = 1
    write_section_header(ws, row, "A. Quality Grade Distribution", 5)
    row += 1

    grade_cols = [
        ("Grade", 10),
        ("Count", 12),
        ("Percentage", 12),
        ("Avg ICP Score", 14),
        ("Meaning", 30),
    ]
    _table_header(ws, row, grade_cols)
    row += 1

    grade_meanings = {
        "A": "Excellent — fully enriched, high-confidence data",
        "B": "Good — mostly complete, ready for outreach",
        "C": "Fair — partial enrichment, needs improvement",
        "D": "Poor — minimal data, enrichment required",
        "F": "Failing — raw extraction only, not outreach-ready",
    }

    # Pre-compute ICP scores for grade correlation
    icp_by_grade: dict[str, list[int]] = defaultdict(list)
    for rec in records:
        grade = (rec.get("quality_grade") or "F").upper()
        icp = compute_icp_score(rec)["icp_score"]
        icp_by_grade[grade].append(icp)

    grade_counts: Counter = Counter(
        (rec.get("quality_grade") or "F").upper() for rec in records
    )

    for grade in ["A", "B", "C", "D", "F"]:
        count = grade_counts.get(grade, 0)
        pct = count / max(total, 1) * 100
        icp_list = icp_by_grade.get(grade, [0])
        avg_icp = sum(icp_list) / max(len(icp_list), 1)
        meaning = grade_meanings.get(grade, "")

        values = [grade, count, f"{pct:.1f}%", f"{avg_icp:.1f}", meaning]
        for col_idx, val in enumerate(values, 1):
            cell = _data_cell(ws, row, col_idx, val)
        # Grade fill on first column
        grade_cell = ws.cell(row=row, column=1)
        if grade in GRADE_FILLS:
            grade_cell.fill = GRADE_FILLS[grade]
            if grade == "A":
                grade_cell.font = WHITE_FONT
            else:
                grade_cell.font = Font(bold=True, size=11)
        _set_row_height(ws, row, 18)
        row += 1

    row += 1

    # ── Section B: Enrichment Coverage ──
    write_section_header(ws, row, "B. Enrichment Coverage by Field", 5)
    row += 1

    enrich_cols = [
        ("Field Name", 25),
        ("Has Data", 12),
        ("Total", 10),
        ("Coverage %", 12),
        ("Status", 15),
    ]
    _table_header(ws, row, enrich_cols)
    row += 1

    enrichment_fields = [
        ("Website", lambda r: bool((r.get("website") or r.get("domain") or "").strip())),
        ("Email Provider", lambda r: bool((r.get("email_provider") or "").strip())),
        ("Tech Stack", lambda r: bool(get_tech_stack(r))),
        ("Contacts", lambda r: bool(get_contacts(r))),
        ("Phone", lambda r: bool((r.get("phone") or "").strip())),
        ("Employee Count", lambda r: bool(get_employee_count(r)[0] or get_employee_count(r)[1])),
        ("Street Address", lambda r: bool((r.get("street") or "").strip())),
        ("CMS Platform", lambda r: bool((r.get("cms") or "").strip())),
        ("SPF Services", lambda r: bool(get_spf_list(r))),
    ]

    for field_name, checker in enrichment_fields:
        has_data = sum(1 for r in records if checker(r))
        pct = has_data / max(total, 1) * 100
        if pct >= 80:
            status = "Good"
            status_fill = _GREEN_FILL
            status_font = _BOLD_GREEN_FONT
        elif pct >= 60:
            status = "Fair"
            status_fill = _YELLOW_FILL
            status_font = _BOLD_YELLOW_FONT
        else:
            status = "Needs Work"
            status_fill = _RED_FILL
            status_font = _BOLD_RED_FONT

        values = [field_name, has_data, total, f"{pct:.1f}%", status]
        for col_idx, val in enumerate(values, 1):
            _data_cell(ws, row, col_idx, val)
        ws.cell(row=row, column=5).fill = status_fill
        ws.cell(row=row, column=5).font = status_font
        _set_row_height(ws, row, 18)
        row += 1

    row += 1

    # ── Section C: Quality Trend & Summary ──
    write_section_header(ws, row, "C. Quality Summary & Improvement Targets", 5)
    row += 1

    high_quality = sum(1 for r in records if (r.get("quality_grade") or "F") in ("A", "B"))
    hq_pct = high_quality / max(total, 1) * 100
    avg_qs = sum(r.get("quality_score", 0) or 0 for r in records) / max(total, 1)
    missing_web = sum(1 for r in records if not has_website(r))
    missing_ep = sum(1 for r in records if not (r.get("email_provider") or "").strip())
    needs_enrichment = sum(1 for r in records if not has_website(r) or not (r.get("email_provider") or "").strip())

    summary_rows = [
        ("Current average quality score", f"{avg_qs:.1f} / 100"),
        ("B+ grade records (outreach-ready)", f"{high_quality:,} ({hq_pct:.0f}%)"),
        ("Records missing website", f"{missing_web:,} ({missing_web/max(total,1)*100:.0f}%)"),
        ("Records missing email provider", f"{missing_ep:,} ({missing_ep/max(total,1)*100:.0f}%)"),
        ("Records needing enrichment (website or email)", f"{needs_enrichment:,}"),
        ("Improvement target", "Procure Clearbit/Apollo API keys → lift avg to 85+"),
    ]

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 18

    for label, value in summary_rows:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(bold=True, size=10)
        lc.border = THIN_BORDER
        lc.alignment = _LEFT_NOWRAP
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = Font(size=10)
        vc.border = THIN_BORDER
        vc.alignment = _LEFT_NOWRAP
        ws.merge_cells(f"B{row}:E{row}")
        _set_row_height(ws, row, 18)
        row += 1

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "landscape"
    print(f"  Sheet 'Data Quality Scorecard': {total} companies, avg quality {avg_qs:.1f}")


# ── Sheet 4: Top 25 Target Accounts ──────────────────────────────────

def write_top_accounts_sheet(wb: Workbook, scored_data: list[tuple]):
    ws = wb.create_sheet("Top 25 Target Accounts")

    top25_cols = [
        ("Rank", 5),
        ("Company Name", 35),
        ("Website", 30),
        ("City", 18),
        ("State", 8),
        ("Association", 15),
        ("ICP Score", 12),
        ("ICP Tier", 12),
        ("Email Provider", 18),
        ("Employee Count", 15),
        ("Primary Contact", 25),
        ("Contact Email", 30),
        ("Quality Score", 12),
        ("Recommended Campaign", 30),
    ]
    _table_header(ws, 1, top25_cols)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:N1"

    top25 = scored_data[:25]

    for row_idx, (rec, icp, tier) in enumerate(top25, 2):
        contact_name, contact_email = get_primary_contact(rec)
        emp = _emp_display(rec)
        assocs = "; ".join(get_associations_list(rec))
        campaign = _recommend_campaign(rec, icp)

        values = [
            row_idx - 1,
            rec.get("company_name", ""),
            (rec.get("website") or "").strip(),
            (rec.get("city") or "").strip(),
            (rec.get("state") or "").strip(),
            assocs,
            icp["icp_score"],
            tier,
            get_email_provider(rec),
            emp,
            contact_name,
            contact_email,
            rec.get("quality_score", 0),
            campaign,
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = _LEFT_NOWRAP

        # Color-code Tier column (8)
        tier_cell = ws.cell(row=row_idx, column=8)
        if tier in TIER_FILLS:
            tier_cell.fill = TIER_FILLS[tier]
            tier_cell.font = TIER_FONTS[tier]
            tier_cell.alignment = _CENTER

        # Color-code ICP Score (7) — shade by value
        icp_score = icp["icp_score"]
        icp_cell = ws.cell(row=row_idx, column=7)
        if icp_score >= 75:
            icp_cell.fill = _GREEN_FILL
            icp_cell.font = _BOLD_GREEN_FONT
        elif icp_score >= 55:
            icp_cell.fill = _YELLOW_FILL
            icp_cell.font = _BOLD_YELLOW_FONT

        _set_row_height(ws, row_idx, 18)

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "landscape"
    ws.print_area = f"A1:N{len(top25) + 1}"

    print(f"  Sheet 'Top 25 Target Accounts': {len(top25)} accounts")


# ── Sheet 5: Competitive Threats ──────────────────────────────────────

def write_competitive_threats_sheet(wb: Workbook, records: list[dict]):
    ws = wb.create_sheet("Competitive Threats")

    competitors = load_competitors()
    events = load_events()

    # Detect competitor presence in pipeline records
    competitor_detection: Counter = Counter()
    for rec in records:
        comp = detect_competitor(rec)
        if comp:
            competitor_detection[comp] += 1

    greenfield = sum(1 for r in records if not detect_competitor(r)
                     and not (r.get("erp_system") or "").strip())

    # ── Section A: Competitor Overview ──
    row = 1
    write_section_header(ws, row, "A. Competitor Overview — ERP Competitive Landscape", 6)
    row += 1

    comp_cols = [
        ("Competitor", 22),
        ("Threat Level", 13),
        ("Show / Association Presence", 40),
        ("Strategy Notes", 45),
        ("GSS Response", 18),
        ("Companies Detected", 18),
    ]
    _table_header(ws, row, comp_cols)
    row += 1

    # Sort by threat level priority
    threat_order = {"HIGH": 0, "EMERGING": 1, "MEDIUM": 2, "LOW": 3}
    sorted_comps = sorted(
        competitors,
        key=lambda c: threat_order.get(c.get("threat_level", "LOW").upper(), 4)
    )

    for comp in sorted_comps:
        threat = comp.get("threat_level", "LOW").upper()
        comp_name = comp.get("competitor", "")
        detected = competitor_detection.get(comp_name.title(), 0)

        # Try case-insensitive match
        if detected == 0:
            for detected_name, cnt in competitor_detection.items():
                if detected_name.lower() in comp_name.lower() or comp_name.lower() in detected_name.lower():
                    detected = cnt
                    break

        gss_response = THREAT_RESPONSE.get(threat, "Track Only")

        values = [
            comp_name,
            threat,
            comp.get("presence", ""),
            comp.get("strategy_notes", ""),
            gss_response,
            detected,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = _LEFT_NOWRAP

        # Color-code threat level
        threat_cell = ws.cell(row=row, column=2)
        if threat in THREAT_FILLS:
            threat_cell.fill = THREAT_FILLS[threat]
            threat_cell.font = WHITE_FONT
            threat_cell.alignment = _CENTER

        # Detected companies styling
        det_cell = ws.cell(row=row, column=6)
        if detected > 0:
            det_cell.font = _BOLD_RED_FONT
        _set_row_height(ws, row, 20)
        row += 1

    row += 1

    # ── Section B: Key Competitive Insights ──
    write_section_header(ws, row, "B. Key Competitive Insights", 6)
    row += 1

    # Most detected competitor
    top_comp = competitor_detection.most_common(1)
    top_comp_str = f"{top_comp[0][0]} ({top_comp[0][1]} companies)" if top_comp else "None detected"

    # Most contested association — find which assoc has most competitor-using companies
    assoc_comp_count: Counter = Counter()
    for rec in records:
        if detect_competitor(rec):
            for a in get_associations_list(rec):
                assoc_comp_count[a.upper()] += 1
    most_contested_assoc = assoc_comp_count.most_common(1)
    most_contested_str = (
        f"{most_contested_assoc[0][0]} ({most_contested_assoc[0][1]} companies with competitor ERP)"
        if most_contested_assoc else "N/A"
    )

    # Events with high-threat competitors
    high_threat_event_comps = {
        c.get("competitor", "") for c in competitors
        if c.get("threat_level", "").upper() in ("HIGH", "EMERGING")
    }
    event_threat_map: dict[str, list[str]] = defaultdict(list)
    for event in events:
        event_name = event.get("event_name", "")
        if event.get("priority", "").upper() == "HIGH":
            for comp in competitors:
                comp_name = comp.get("competitor", "")
                presence = comp.get("presence", "")
                if comp_name in high_threat_event_comps:
                    for keyword in event_name.split():
                        if len(keyword) > 3 and keyword.upper() in presence.upper():
                            event_threat_map[event_name].append(comp_name)
                            break
    top_event_names = sorted(event_threat_map.keys(),
                              key=lambda e: -len(event_threat_map[e]))[:3]
    event_str = ", ".join(top_event_names) if top_event_names else "IMTS, FABTECH"

    total_with_competitor = sum(competitor_detection.values())

    insights = [
        (f"Greenfield opportunity", f"{greenfield:,} companies with no ERP detected — highest-priority "
         f"target segment for first-ERP positioning"),
        ("Top competitor by pipeline detection", top_comp_str),
        ("Most contested association", most_contested_str),
        ("Events with highest competitor presence", event_str),
        ("Microsoft 365 displacement", "662 companies using M365 — strong co-sell angle with "
         "Microsoft partner program and Dynamics transition story"),
        ("Total companies with competitor ERP", f"{total_with_competitor:,} — displacement "
         f"campaigns should prioritize these accounts"),
    ]

    for label, value in insights:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(bold=True, size=10, color="1F4E79")
        lc.border = THIN_BORDER
        lc.alignment = _LEFT_NOWRAP
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = Font(size=10)
        vc.border = THIN_BORDER
        vc.alignment = _LEFT
        ws.merge_cells(f"B{row}:F{row}")
        _set_row_height(ws, row, 24)
        row += 1

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "landscape"
    print(f"  Sheet 'Competitive Threats': {len(sorted_comps)} competitors, "
          f"{greenfield:,} greenfield accounts")


# ── Sheet 6: Investment Recommendations ──────────────────────────────

def write_recommendations_sheet(wb: Workbook):
    ws = wb.create_sheet("Investment Recommendations")

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 50

    rec_cols = [
        ("Action", 45),
        ("Impact", 10),
        ("Effort", 10),
        ("Expected Outcome", 50),
    ]

    # ── Section A: Immediate (30 days) ──
    row = 1
    write_section_header(ws, row, "A. Immediate Actions — Next 30 Days", 4)
    row += 1
    _table_header(ws, row, rec_cols)
    row += 1

    immediate_actions = [
        ("Procure Clearbit/Apollo API keys",
         "High", "Low",
         "Enable deep firmographic enrichment — employee count, revenue, SIC codes for all 2,083 companies"),
        ("Extract NTMA association (~1,400 expected members)",
         "High", "Med",
         "Largest single untapped association — single extraction run adds ~60% more companies"),
        ("Deploy Microsoft co-sell partnership campaign",
         "High", "Med",
         "662 M365 companies identified — Microsoft partner incentives + Dynamics displacement angle"),
        ("Run competitive displacement campaign (Epicor/Plex)",
         "High", "Low",
         "Target companies with HIGH-threat competitor ERP detected; lead with migration ROI story"),
    ]

    for action, impact, effort, outcome in immediate_actions:
        values = [action, impact, effort, outcome]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = _LEFT if col_idx == 4 else _LEFT_NOWRAP

        # Color Impact (col 2)
        ic = ws.cell(row=row, column=2)
        ic.fill = _IMPACT_FILLS.get(impact, _GRAY_FILL)
        ic.font = _IMPACT_FONTS.get(impact, _BOLD_GRAY_FONT)
        ic.alignment = _CENTER

        # Color Effort (col 3)
        ec = ws.cell(row=row, column=3)
        ec.fill = _EFFORT_FILLS.get(effort, _GRAY_FILL)
        ec.font = _EFFORT_FONTS.get(effort, _BOLD_GRAY_FONT)
        ec.alignment = _CENTER

        _set_row_height(ws, row, 30)
        row += 1

    row += 1

    # ── Section B: Medium-Term (30–90 days) ──
    write_section_header(ws, row, "B. Medium-Term Actions — 30 to 90 Days", 4)
    row += 1
    _table_header(ws, row, rec_cols)
    row += 1

    medium_actions = [
        ("Extract remaining 8 associations (PMPA, FIA, NADCA, AFS, FMA, AMT, PMMI, PLASTICS)",
         "High", "High",
         "~4,300 additional expected members; brings pipeline to 6,000+ total companies"),
        ("Build Salesforce integration for GSS CRM",
         "Med", "Med",
         "Automate CRM data sync; 712-row SFDC-ready import file already generated"),
        ("Implement incremental extraction (weekly refresh)",
         "Med", "Med",
         "Keep pipeline current as associations add/remove members; prevents data decay"),
        ("PMA website enrichment (968 records missing domains)",
         "Med", "Med",
         "Scrape PMA profile pages for company websites; unlocks DNS/tech stack enrichment for largest association"),
    ]

    for action, impact, effort, outcome in medium_actions:
        values = [action, impact, effort, outcome]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = _LEFT if col_idx in (1, 4) else _LEFT_NOWRAP

        ic = ws.cell(row=row, column=2)
        ic.fill = _IMPACT_FILLS.get(impact, _GRAY_FILL)
        ic.font = _IMPACT_FONTS.get(impact, _BOLD_GRAY_FONT)
        ic.alignment = _CENTER

        ec = ws.cell(row=row, column=3)
        ec.fill = _EFFORT_FILLS.get(effort, _GRAY_FILL)
        ec.font = _EFFORT_FONTS.get(effort, _BOLD_GRAY_FONT)
        ec.alignment = _CENTER

        _set_row_height(ws, row, 30)
        row += 1

    row += 1

    # ── Section C: Long-Term (90+ days) ──
    write_section_header(ws, row, "C. Long-Term Investments — 90+ Days", 4)
    row += 1
    _table_header(ws, row, rec_cols)
    row += 1

    long_actions = [
        ("Build self-service admin dashboard",
         "Low", "High",
         "Allow marketing team to run queries, export lists, and view pipeline stats without engineering support"),
        ("Implement predictive lead scoring (ML-based purchase intent)",
         "High", "High",
         "Train model on historical win/loss data + enrichment signals; "
         "prioritize accounts most likely to convert in next 90 days"),
    ]

    for action, impact, effort, outcome in long_actions:
        values = [action, impact, effort, outcome]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = _LEFT if col_idx in (1, 4) else _LEFT_NOWRAP

        ic = ws.cell(row=row, column=2)
        ic.fill = _IMPACT_FILLS.get(impact, _GRAY_FILL)
        ic.font = _IMPACT_FONTS.get(impact, _BOLD_GRAY_FONT)
        ic.alignment = _CENTER

        ec = ws.cell(row=row, column=3)
        ec.fill = _EFFORT_FILLS.get(effort, _GRAY_FILL)
        ec.font = _EFFORT_FONTS.get(effort, _BOLD_GRAY_FONT)
        ec.alignment = _CENTER

        _set_row_height(ws, row, 30)
        row += 1

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "landscape"
    ws.print_area = f"A1:D{row}"
    print(f"  Sheet 'Investment Recommendations': "
          f"{len(immediate_actions)} immediate, {len(medium_actions)} medium, "
          f"{len(long_actions)} long-term actions")


# ── Main ──────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("NAM Intelligence Pipeline — Executive Briefing")
    print("=" * 60)

    # Ensure output directory exists
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

    # Load data
    print("\nLoading and merging data...")
    records = load_and_merge_data()
    print(f"  Loaded {len(records)} companies")

    # Score all records for ICP
    print("Computing ICP scores...")
    scored_data: list[tuple] = []
    for rec in records:
        icp = compute_icp_score(rec)
        contacts = get_contacts(rec)
        qs = rec.get("quality_score", 0)
        if not isinstance(qs, (int, float)):
            try:
                qs = int(qs)
            except (ValueError, TypeError):
                qs = 0
        tier = assign_tier(icp["icp_score"], bool(contacts), qs)
        scored_data.append((rec, icp, tier))

    # Sort by ICP score descending
    scored_data.sort(key=lambda x: -x[1]["icp_score"])

    tier_counts = Counter(t for _, _, t in scored_data)
    print(f"  Tier 1: {tier_counts.get('Tier 1', 0)}, "
          f"Tier 2: {tier_counts.get('Tier 2', 0)}, "
          f"Tier 3: {tier_counts.get('Tier 3', 0)}")

    # Generate workbook
    print("\nGenerating workbook...")
    wb = Workbook()

    write_title_sheet(wb, records, scored_data)
    write_market_opportunity_sheet(wb, records)
    write_quality_scorecard_sheet(wb, records)
    write_top_accounts_sheet(wb, scored_data)
    write_competitive_threats_sheet(wb, records)
    write_recommendations_sheet(wb)

    wb.save(OUTPUT_PATH)
    print(f"\nSaved to: {OUTPUT_PATH}")
    print("Done!")


if __name__ == "__main__":
    main()
