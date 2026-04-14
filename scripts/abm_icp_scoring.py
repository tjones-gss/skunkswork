#!/usr/bin/env python3
"""
ICP Scoring Script
NAM Intelligence Pipeline

Generates: data/exports/GSS_ICP_Scored_Accounts.xlsx
  Sheet 1: ICP Scored Accounts (all companies with scores + tiers)
  Sheet 2: Tier 1 Strategic (Top 50)
  Sheet 3: ICP Score Distribution (summary tables)
  Sheet 4: ICP Model Weights (methodology documentation)
"""

import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Add parent to path for imports
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from scripts.abm_shared import (
    BASE_DIR, EXPORTS_DIR,
    HEADER_FILL, HEADER_FONT, HEADER_ALIGNMENT, THIN_BORDER,
    GRADE_FILLS, TIER_FILLS, TIER_FONTS,
    SECTION_FILL, SECTION_FONT, SUBTITLE_FONT, METRIC_FONT, METRIC_LARGE_FONT,
    load_and_merge_data, compute_icp_score, assign_tier,
    get_contacts, get_primary_contact, get_email_provider,
    get_associations_list, get_tech_stack, get_spf_list,
    get_employee_count, has_website, detect_competitor,
    style_header_row, write_section_header,
)

OUTPUT_PATH = EXPORTS_DIR / "GSS_ICP_Scored_Accounts.xlsx"

# Columns for the main scored accounts sheet
SCORED_COLUMNS = [
    ("Company Name", 35),
    ("Website", 30),
    ("Domain", 20),
    ("City", 18),
    ("State", 8),
    ("Association(s)", 15),
    ("ICP Score", 12),
    ("ICP Tier", 12),
    ("Tech Maturity", 14),
    ("Size Fit", 12),
    ("Geo Fit", 12),
    ("Assoc Engagement", 16),
    ("Tech Gap", 12),
    ("Data Quality", 14),
    ("Email Provider", 18),
    ("Employee Count", 15),
    ("Quality Score", 12),
    ("Quality Grade", 12),
    ("Primary Contact", 25),
    ("Contact Email", 30),
    ("Detected Competitor", 20),
]

# Columns for Tier 1 strategic sheet
TIER1_COLUMNS = SCORED_COLUMNS + [
    ("Recommended Approach", 40),
]


def _emp_display(rec: dict) -> str:
    ec_min, ec_max = get_employee_count(rec)
    if ec_min and ec_max and ec_min != ec_max:
        return f"{ec_min:,}-{ec_max:,}"
    if ec_max:
        return f"{ec_max:,}"
    if ec_min:
        return f"{ec_min:,}+"
    return ""


def _recommend_approach(rec: dict, icp: dict) -> str:
    """Generate a recommended engagement approach based on ICP dimension strengths."""
    # Find the strongest non-trivial dimension
    dims = [
        ("tech_maturity", icp["tech_maturity"], 25),
        ("size_fit", icp["size_fit"], 20),
        ("geo_fit", icp["geo_fit"], 15),
        ("tech_gap", icp["tech_gap"], 15),
        ("assoc_engagement", icp["assoc_engagement"], 15),
    ]
    # Sort by percentage of max achieved
    dims.sort(key=lambda d: d[1] / d[2], reverse=True)
    top = dims[0][0]

    ep = get_email_provider(rec).lower()
    competitor = detect_competitor(rec)

    if competitor:
        return f"Competitive displacement — lead with migration from {competitor.title()}"
    if top == "tech_gap" and icp["tech_gap"] >= 12:
        return "Greenfield ERP opportunity — lead with first-ERP value story"
    if top == "tech_maturity" and "microsoft" in ep:
        return "ABM 1:1 — Microsoft ecosystem modernization, Dynamics displacement angle"
    if top == "tech_maturity" and ep in ("self-hosted", "self-hosted (on-premise)"):
        return "Digital transformation — legacy IT modernization campaign"
    if top == "size_fit" and icp["size_fit"] >= 18:
        return "ABM 1:1 — perfect mid-market fit, personalized executive outreach"
    if top == "assoc_engagement" and icp["assoc_engagement"] >= 12:
        return "Association-centric — leverage peer trust and event presence"
    if top == "geo_fit" and icp["geo_fit"] >= 12:
        return "Regional campaign — local manufacturing cluster outreach"
    return "ABM 1:few — multi-touch nurture sequence"


def build_scored_row(rec: dict, icp: dict, tier: str) -> list:
    """Build a data row for the scored accounts sheet."""
    contact_name, contact_email = get_primary_contact(rec)
    return [
        rec.get("company_name", ""),
        (rec.get("website") or "").strip(),
        (rec.get("domain") or "").strip(),
        (rec.get("city") or "").strip(),
        (rec.get("state") or "").strip(),
        "; ".join(get_associations_list(rec)),
        icp["icp_score"],
        tier,
        icp["tech_maturity"],
        icp["size_fit"],
        icp["geo_fit"],
        icp["assoc_engagement"],
        icp["tech_gap"],
        icp["data_quality"],
        get_email_provider(rec),
        _emp_display(rec),
        rec.get("quality_score", 0),
        rec.get("quality_grade", ""),
        contact_name,
        contact_email,
        detect_competitor(rec),
    ]


def build_tier1_row(rec: dict, icp: dict, tier: str) -> list:
    """Build a row for Tier 1 sheet (includes recommendation)."""
    base = build_scored_row(rec, icp, tier)
    base.append(_recommend_approach(rec, icp))
    return base


def write_scored_sheet(wb: Workbook, scored_data: list[tuple]):
    """Sheet 1: All ICP Scored Accounts."""
    ws = wb.active
    ws.title = "ICP Scored Accounts"
    style_header_row(ws, SCORED_COLUMNS)

    for row_idx, (rec, icp, tier) in enumerate(scored_data, 2):
        row = build_scored_row(rec, icp, tier)
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)

        # Color-code tier column (col 8)
        tier_cell = ws.cell(row=row_idx, column=8)
        if tier in TIER_FILLS:
            tier_cell.fill = TIER_FILLS[tier]
            tier_cell.font = TIER_FONTS[tier]

        # Color-code grade column (col 18)
        grade = rec.get("quality_grade", "")
        grade_cell = ws.cell(row=row_idx, column=18)
        if grade in GRADE_FILLS:
            grade_cell.fill = GRADE_FILLS[grade]

    ws.auto_filter.ref = ws.dimensions
    print(f"  Sheet 'ICP Scored Accounts': {len(scored_data)} rows")


def write_tier1_sheet(wb: Workbook, scored_data: list[tuple]):
    """Sheet 2: Tier 1 Strategic Accounts."""
    ws = wb.create_sheet("Tier 1 Strategic (Top 50)")
    style_header_row(ws, TIER1_COLUMNS)

    tier1 = [(rec, icp, tier) for rec, icp, tier in scored_data if tier == "Tier 1"][:50]

    for row_idx, (rec, icp, tier) in enumerate(tier1, 2):
        row = build_tier1_row(rec, icp, tier)
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == len(row)))

        # Green tier fill
        tier_cell = ws.cell(row=row_idx, column=8)
        tier_cell.fill = TIER_FILLS["Tier 1"]
        tier_cell.font = TIER_FONTS["Tier 1"]

    ws.auto_filter.ref = ws.dimensions
    print(f"  Sheet 'Tier 1 Strategic': {len(tier1)} rows")


def write_distribution_sheet(wb: Workbook, scored_data: list[tuple]):
    """Sheet 3: ICP Score Distribution & Summary Tables."""
    ws = wb.create_sheet("ICP Score Distribution")

    # ── Table 1: Score Range Distribution ──
    row = 1
    write_section_header(ws, row, "ICP Score Distribution", 5)
    row += 1
    headers = [("Score Range", 15), ("Count", 12), ("Percentage", 12), ("Avg Quality", 14), ("Tier Mix", 25)]
    for col_idx, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    row += 1

    total = len(scored_data)
    ranges = [(90, 100), (80, 89), (70, 79), (60, 69), (50, 59), (40, 49), (30, 39), (0, 29)]
    for lo, hi in ranges:
        in_range = [(r, i, t) for r, i, t in scored_data if lo <= i["icp_score"] <= hi]
        count = len(in_range)
        pct = f"{count / total * 100:.1f}%" if total else "0%"
        avg_q = sum(r.get("quality_score", 0) for r, _, _ in in_range) / max(count, 1)
        tiers = Counter(t for _, _, t in in_range)
        tier_str = ", ".join(f"{k}: {v}" for k, v in sorted(tiers.items()))

        for col_idx, val in enumerate([f"{lo}-{hi}", count, pct, f"{avg_q:.1f}", tier_str], 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
        row += 1

    # ── Table 2: Average ICP by Association ──
    row += 2
    write_section_header(ws, row, "Average ICP Score by Association", 5)
    row += 1
    headers2 = [("Association", 20), ("Count", 12), ("Avg ICP", 12), ("Avg Quality", 14), ("Tier 1 Count", 14)]
    for col_idx, (h, w) in enumerate(headers2, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    row += 1

    assoc_data = defaultdict(list)
    for rec, icp, tier in scored_data:
        for a in get_associations_list(rec):
            assoc_data[a.upper()].append((icp["icp_score"], rec.get("quality_score", 0), tier))

    for assoc in sorted(assoc_data.keys(), key=lambda a: -len(assoc_data[a])):
        items = assoc_data[assoc]
        count = len(items)
        avg_icp = sum(i for i, _, _ in items) / count
        avg_q = sum(q for _, q, _ in items) / count
        t1_count = sum(1 for _, _, t in items if t == "Tier 1")
        for col_idx, val in enumerate([assoc, count, f"{avg_icp:.1f}", f"{avg_q:.1f}", t1_count], 1):
            ws.cell(row=row, column=col_idx, value=val).border = THIN_BORDER
        row += 1

    # ── Table 3: Average ICP by State (Top 15) ──
    row += 2
    write_section_header(ws, row, "Average ICP Score by State (Top 15)", 5)
    row += 1
    headers3 = [("State", 10), ("Count", 12), ("Avg ICP", 12), ("Avg Quality", 14), ("Tier 1 Count", 14)]
    for col_idx, (h, w) in enumerate(headers3, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    row += 1

    state_data = defaultdict(list)
    for rec, icp, tier in scored_data:
        st = (rec.get("state") or "").strip().upper()
        if st:
            state_data[st].append((icp["icp_score"], rec.get("quality_score", 0), tier))

    sorted_states = sorted(state_data.keys(), key=lambda s: -len(state_data[s]))[:15]
    for st in sorted_states:
        items = state_data[st]
        count = len(items)
        avg_icp = sum(i for i, _, _ in items) / count
        avg_q = sum(q for _, q, _ in items) / count
        t1_count = sum(1 for _, _, t in items if t == "Tier 1")
        for col_idx, val in enumerate([st, count, f"{avg_icp:.1f}", f"{avg_q:.1f}", t1_count], 1):
            ws.cell(row=row, column=col_idx, value=val).border = THIN_BORDER
        row += 1

    # ── Tier Summary ──
    row += 2
    write_section_header(ws, row, "Tier Distribution Summary", 5)
    row += 1
    for col_idx, (h, _) in enumerate([("Tier", 15), ("Count", 12), ("Percentage", 12), ("Avg ICP", 12), ("With Contacts %", 16)], 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    row += 1

    tier_groups = defaultdict(list)
    for rec, icp, tier in scored_data:
        tier_groups[tier].append((rec, icp))

    for tier_name in ["Tier 1", "Tier 2", "Tier 3", "Unqualified"]:
        items = tier_groups.get(tier_name, [])
        count = len(items)
        pct = f"{count / total * 100:.1f}%" if total else "0%"
        avg_icp = sum(i["icp_score"] for _, i in items) / max(count, 1)
        with_contacts = sum(1 for r, _ in items if get_contacts(r)) / max(count, 1) * 100
        for col_idx, val in enumerate([tier_name, count, pct, f"{avg_icp:.1f}", f"{with_contacts:.1f}%"], 1):
            ws.cell(row=row, column=col_idx, value=val).border = THIN_BORDER
        row += 1

    print(f"  Sheet 'ICP Score Distribution': summary tables written")


def write_model_sheet(wb: Workbook):
    """Sheet 4: ICP Model Weights — self-documenting methodology."""
    ws = wb.create_sheet("ICP Model Weights")
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 60

    row = 1
    write_section_header(ws, row, "ICP Scoring Model — GSS ERP Sales Fitness", 3)
    row += 2

    ws.cell(row=row, column=1, value="Purpose:").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Score each company 0-100 on likelihood to be a good GSS ERP customer").alignment = Alignment(wrap_text=True)
    row += 1
    ws.cell(row=row, column=1, value="Distinct from Quality Score:").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Quality measures data completeness. ICP measures sales fitness.").alignment = Alignment(wrap_text=True)
    row += 2

    # Dimension details
    dimensions = [
        ("Tech Maturity Signal", "0-25", [
            "Microsoft 365 email = 10 pts (Dynamics displacement opportunity)",
            "Proofpoint/Mimecast = 8 pts (enterprise security = budget available)",
            "Google Workspace = 5 pts (cloud but Google ecosystem)",
            "Self-hosted email = 12 pts (legacy = modernization opportunity)",
            "Salesforce/Pardot in SPF = +5 pts (CRM investment signals ERP budget)",
            "Marketing automation in SPF = +3 pts (tech-savvy)",
            "CDN (Cloudflare/Akamai) = +2 pts (web investment)",
        ]),
        ("Company Size Fit", "0-20", [
            "50-500 employees = 20 pts (GSS sweet spot)",
            "25-49 = 15 pts (slightly small)",
            "501-1000 = 12 pts (getting large)",
            "10-24 = 8 pts (small shop)",
            "1000+ = 5 pts (enterprise, outside typical GSS deal)",
            "Unknown = 10 pts (benefit of doubt)",
        ]),
        ("Geographic Fit", "0-15", [
            "Top mfg states (TX/OH/MI/IN/IL/PA/WI/CA/NC/TN) = 15 pts",
            "Mid-tier states (AL/SC/GA/MN/MO etc.) = 12 pts",
            "Other US states = 10 pts",
            "Unknown = 7 pts",
            "International = 5 pts",
        ]),
        ("Association Engagement", "0-15", [
            "3+ associations = 15 pts (highly engaged)",
            "2 associations = 12 pts",
            "1 high-priority association (PMA/NEMA/AGMA/AIA/SOCMA) = 10 pts",
            "1 medium-priority = 7 pts",
        ]),
        ("Tech Stack Gap", "0-15", [
            "No ERP detected = 15 pts (greenfield opportunity)",
            "Competitor ERP detected = 5 pts (displacement opportunity)",
            "GSS already detected = 0 pts (existing customer)",
        ]),
        ("Data Quality Bonus", "0-10", [
            "Has named contacts = +4 pts",
            "Has phone number = +2 pts",
            "Enrichment complete = +2 pts",
            "Quality score >= 80 = +2 pts",
        ]),
    ]

    for dim_name, dim_range, rules in dimensions:
        write_section_header(ws, row, f"{dim_name} ({dim_range})", 3)
        row += 1
        for rule in rules:
            ws.cell(row=row, column=1, value="").border = THIN_BORDER
            ws.cell(row=row, column=2, value="").border = THIN_BORDER
            ws.cell(row=row, column=3, value=rule).border = THIN_BORDER
            ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
            row += 1
        row += 1

    # Tiering rules
    row += 1
    write_section_header(ws, row, "Tier Assignment Rules", 3)
    row += 1
    tiers = [
        ("Tier 1 — Strategic", "ICP >= 75 AND has contacts AND quality >= 70", "Top 50 max. ABM 1:1 personalized outreach."),
        ("Tier 2 — Growth", "ICP >= 55 (or ICP >= 75 without contacts)", "Next 200. ABM 1:few multi-touch campaigns."),
        ("Tier 3 — Nurture", "ICP >= 35", "Remaining qualified. Content marketing + events."),
        ("Unqualified", "ICP < 35", "Excluded from active ABM programs."),
    ]
    for tier, criteria, desc in tiers:
        ws.cell(row=row, column=1, value=tier).font = Font(bold=True)
        ws.cell(row=row, column=2, value=criteria)
        ws.cell(row=row, column=3, value=desc).alignment = Alignment(wrap_text=True)
        row += 1

    print(f"  Sheet 'ICP Model Weights': methodology documented")


def main():
    print("=" * 60)
    print("NAM Intelligence Pipeline - ICP Scoring")
    print("=" * 60)

    # Load data
    print("\nLoading and merging data...")
    records = load_and_merge_data()
    print(f"  Loaded {len(records)} companies")

    # Score all records
    print("\nComputing ICP scores...")
    scored_data = []
    for rec in records:
        icp = compute_icp_score(rec)
        contacts = get_contacts(rec)
        qs = rec.get("quality_score", 0)
        if not isinstance(qs, (int, float)):
            try:
                qs = int(qs)
            except (ValueError, TypeError):
                qs = 0
        tier = assign_tier(icp["icp_score"], bool(contacts), qs)
        scored_data.append((rec, icp, tier))

    # Sort by ICP score descending
    scored_data.sort(key=lambda x: -x[1]["icp_score"])

    # Stats
    scores = [icp["icp_score"] for _, icp, _ in scored_data]
    tiers = Counter(t for _, _, t in scored_data)
    print(f"  Average ICP: {sum(scores) / len(scores):.1f}")
    print(f"  Tier 1: {tiers.get('Tier 1', 0)}, Tier 2: {tiers.get('Tier 2', 0)}, "
          f"Tier 3: {tiers.get('Tier 3', 0)}, Unqualified: {tiers.get('Unqualified', 0)}")

    # Generate workbook
    print("\nGenerating workbook...")
    wb = Workbook()

    write_scored_sheet(wb, scored_data)
    write_tier1_sheet(wb, scored_data)
    write_distribution_sheet(wb, scored_data)
    write_model_sheet(wb)

    wb.save(OUTPUT_PATH)
    print(f"\nSaved to: {OUTPUT_PATH}")
    print("Done!")


if __name__ == "__main__":
    main()
