#!/usr/bin/env python3
"""
ABM Market Intelligence Brief
NAM Intelligence Pipeline

Generates GSS_Market_Intelligence_Brief.xlsx — a 6-sheet market analysis workbook:
  Sheet 1: Geographic Density
  Sheet 2: Association Overlap Matrix
  Sheet 3: Technology Maturity Index
  Sheet 4: Email Infrastructure Analysis
  Sheet 5: Competitive Pressure Map
  Sheet 6: Market Opportunity Summary
"""

import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from scripts.abm_shared import (
    BASE_DIR,
    EXPORTS_DIR,
    HEADER_FILL,
    HEADER_FONT,
    HEADER_ALIGNMENT,
    THIN_BORDER,
    SECTION_FILL,
    SECTION_FONT,
    SUBTITLE_FONT,
    METRIC_FONT,
    THREAT_FILLS,
    THREAT_RESPONSE,
    WHITE_FONT,
    COMPETITOR_ALIASES,
    load_and_merge_data,
    load_events,
    load_competitors,
    load_associations_config,
    get_email_provider,
    get_spf_list,
    get_tech_stack,
    get_associations_list,
    get_primary_contact,
    get_contacts,
    get_employee_count,
    has_website,
    compute_icp_score,
    assign_tier,
    detect_competitor,
    style_header_row,
    write_section_header,
)

OUTPUT_PATH = EXPORTS_DIR / "GSS_Market_Intelligence_Brief.xlsx"

# ── Color fills for market classification ─────────────────────────────
HOT_FILL = PatternFill(start_color="006100", end_color="006100", fill_type="solid")
OPP_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
EMRG_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
DEV_FILL = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
HOT_FONT = Font(bold=True, color="FFFFFF")
OPP_FONT = Font(bold=True, color="7D5700")
EMRG_FONT = Font(bold=True, color="1F4E79")
DEV_FONT = Font(color="555555")

LIGHT_BLUE_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
DIAG_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
DIAG_FONT = Font(bold=True, color="FFFFFF")

MODERN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
TRANS_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
LEGACY_FILL = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
UNKN_FILL = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")

DASHBOARD_METRIC_FONT = Font(bold=True, size=20, color="1F4E79")
DASHBOARD_LABEL_FONT = Font(bold=True, size=11, color="555555")
HIGHLIGHT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# ── Tech Maturity Categories for scoring ──────────────────────────────
TECH_CATEGORIES = {
    "Analytics": [
        "Google Analytics", "Google Tag Manager", "Adobe DTM", "Adobe Launch",
        "Hotjar", "Matomo", "Heap", "Mixpanel", "Segment", "Amplitude",
    ],
    "CMS": [
        "WordPress", "Drupal", "HubSpot CMS", "Squarespace", "Wix", "TYPO3 CMS",
        "Joomla", "Kentico", "Sitecore", "Contentful", "Sanity", "Craft CMS",
    ],
    "CDN": [
        "Cloudflare", "Akamai", "Amazon CloudFront", "Fastly", "StackPath",
        "cdnjs", "jsDelivr CDN", "unpkg CDN", "Google CDN",
    ],
    "Framework": [
        "React", "Angular", "Vue.js", "Next.js", "Nuxt.js", "jQuery",
        "Bootstrap", "Tailwind CSS", "ASP.NET", "Express.js", "Laravel",
    ],
    "Email/Marketing": [
        "Microsoft 365", "Salesforce", "Mailchimp", "SendGrid", "HubSpot",
        "Pardot", "Marketo", "Mandrill", "Constant Contact", "Amazon SES",
        "Zendesk", "LiveChat",
    ],
    "Security": [
        "Cloudflare", "reCAPTCHA", "Sucuri", "Imperva", "Akamai",
    ],
    "Hosting": [
        "Nginx", "Apache", "Amazon S3", "Azure", "Google Cloud",
        "Vercel", "Netlify", "WP Engine",
    ],
}

# SPF service category mapping
SPF_CATEGORIES = {
    "Microsoft 365": "Email Infrastructure",
    "Salesforce": "CRM",
    "Pardot": "CRM",
    "Google Workspace": "Email Infrastructure",
    "Amazon SES": "Transactional",
    "Mailchimp": "Marketing",
    "SendGrid": "Transactional",
    "Constant Contact": "Marketing",
    "HubSpot": "Marketing",
    "Marketo": "Marketing",
    "ActiveCampaign": "Marketing",
    "Proofpoint": "Security",
    "Mimecast": "Security",
    "Barracuda": "Security",
    "Postmark": "Transactional",
    "Mandrill": "Transactional",
    "SparkPost": "Transactional",
}


def _col_letter(col_idx: int) -> str:
    from openpyxl.utils import get_column_letter
    return get_column_letter(col_idx)


def classify_market(count: int, avg_icp: float) -> str:
    if count >= 50 and avg_icp >= 60:
        return "Hot Market"
    if count >= 20 and avg_icp >= 50:
        return "Opportunity Market"
    if count >= 5:
        return "Emerging"
    return "Developing"


def compute_tech_maturity_score(rec: dict) -> int:
    """Compute a 0-100 tech maturity score for a single company."""
    score = 0
    ts = [t.lower() for t in get_tech_stack(rec)]
    spf = [s.lower() for s in get_spf_list(rec)]
    cms = (rec.get("cms") or "").lower()

    # CMS component (max 20 pts — scale CMS score 0-100 to 0-20)
    cms_raw = 0
    if "hubspot" in cms:
        cms_raw = 70
    elif "drupal" in cms:
        cms_raw = 50
    elif "wordpress" in cms:
        cms_raw = 40
    elif any(x in cms for x in ("squarespace", "wix")):
        cms_raw = 30
    elif cms:
        cms_raw = 20
    score += int(cms_raw / 100 * 20)

    # Analytics component (max 20 pts)
    analytics_raw = 0
    if any(t in ("adobe dtm", "adobe launch") for t in ts):
        analytics_raw = 50
    elif any(t in ("hotjar", "heap") for t in ts):
        analytics_raw = 40
    elif "google analytics" in ts:
        analytics_raw = 30
    if "google tag manager" in ts and analytics_raw > 0:
        analytics_raw = min(analytics_raw + 10, 60)
    score += int(analytics_raw / 100 * 20)

    # CDN component (max 20 pts)
    cdn_raw = 0
    if any(t in ("cloudflare", "akamai", "amazon cloudfront") for t in ts):
        cdn_raw = 20
    score += int(cdn_raw / 20 * 20)

    # Framework component (max 20 pts)
    fw_raw = 0
    if any(t in ("react", "angular", "vue.js", "next.js", "nuxt.js") for t in ts):
        fw_raw = 30
    elif "jquery" in ts:
        fw_raw = 10
    score += int(fw_raw / 30 * 20)

    # Marketing automation in SPF (max 20 pts)
    ma_raw = 0
    if any(s in ("hubspot", "marketo", "pardot") for s in spf):
        ma_raw = 20
    elif any(s in ("mailchimp", "sendgrid") for s in spf):
        ma_raw = 10
    score += int(ma_raw / 20 * 20)

    return min(score, 100)


def get_tech_maturity_tier(score: int, has_tech: bool) -> str:
    if not has_tech:
        return "Unknown"
    if score >= 70:
        return "Modern"
    if score >= 40:
        return "Transitional"
    return "Legacy"


# ──────────────────────────────────────────────────────────────────────
# Sheet 1: Geographic Density
# ──────────────────────────────────────────────────────────────────────

def build_geographic_density(ws, companies: list[dict]) -> None:
    ws.title = "Geographic Density"

    COLUMNS = [
        ("State", 10),
        ("Companies", 12),
        ("Avg Quality Score", 15),
        ("Avg ICP Score", 15),
        ("% With Website", 12),
        ("% With Email Provider", 15),
        ("% With Contacts", 12),
        ("% With Tech Stack", 12),
        ("Top Association", 20),
        ("Market Classification", 20),
    ]
    style_header_row(ws, COLUMNS, row=1)

    # Aggregate per state
    state_data: dict[str, dict] = defaultdict(lambda: {
        "count": 0,
        "quality_scores": [],
        "icp_scores": [],
        "with_website": 0,
        "with_email": 0,
        "with_contacts": 0,
        "with_tech": 0,
        "assoc_counter": Counter(),
    })

    for rec in companies:
        state = (rec.get("state") or "").strip().upper()
        if not state:
            state = "Unknown"
        d = state_data[state]
        d["count"] += 1
        d["quality_scores"].append(rec.get("quality_score", 0) or 0)
        icp = compute_icp_score(rec)["icp_score"]
        d["icp_scores"].append(icp)
        if has_website(rec):
            d["with_website"] += 1
        if get_email_provider(rec):
            d["with_email"] += 1
        if get_contacts(rec):
            d["with_contacts"] += 1
        if get_tech_stack(rec):
            d["with_tech"] += 1
        for assoc in get_associations_list(rec):
            d["assoc_counter"][assoc] += 1

    # Sort by count descending
    sorted_states = sorted(state_data.items(), key=lambda x: x[1]["count"], reverse=True)

    market_colors = {
        "Hot Market": (HOT_FILL, HOT_FONT),
        "Opportunity Market": (OPP_FILL, OPP_FONT),
        "Emerging": (EMRG_FILL, EMRG_FONT),
        "Developing": (DEV_FILL, DEV_FONT),
    }

    for row_idx, (state, d) in enumerate(sorted_states, 2):
        count = d["count"]
        avg_quality = sum(d["quality_scores"]) / count if count else 0
        avg_icp = sum(d["icp_scores"]) / count if count else 0
        pct_website = d["with_website"] / count * 100 if count else 0
        pct_email = d["with_email"] / count * 100 if count else 0
        pct_contacts = d["with_contacts"] / count * 100 if count else 0
        pct_tech = d["with_tech"] / count * 100 if count else 0
        top_assoc = d["assoc_counter"].most_common(1)[0][0] if d["assoc_counter"] else ""
        classification = classify_market(count, avg_icp)

        row_values = [
            state,
            count,
            round(avg_quality, 1),
            round(avg_icp, 1),
            f"{pct_website:.0f}%",
            f"{pct_email:.0f}%",
            f"{pct_contacts:.0f}%",
            f"{pct_tech:.0f}%",
            top_assoc,
            classification,
        ]
        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Color-code market classification cell
        fill, font = market_colors.get(classification, (None, None))
        class_cell = ws.cell(row=row_idx, column=10)
        if fill:
            class_cell.fill = fill
        if font:
            class_cell.font = font

    ws.row_dimensions[1].height = 30
    print(f"  Sheet 1: Geographic Density — {len(sorted_states)} states/regions")


# ──────────────────────────────────────────────────────────────────────
# Sheet 2: Association Overlap Matrix
# ──────────────────────────────────────────────────────────────────────

def build_association_overlap(ws, companies: list[dict]) -> None:
    ws.title = "Association Overlap Matrix"

    # Collect all associations
    all_assocs: Counter = Counter()
    company_assocs: list[list[str]] = []
    for rec in companies:
        assocs = get_associations_list(rec)
        if assocs:
            company_assocs.append(assocs)
            for a in assocs:
                all_assocs[a] += 1

    # Sort associations: known high-priority first, then alphabetically
    PRIORITY_ORDER = ["PMA", "AGMA", "NEMA", "AIA", "SOCMA"]
    assoc_list = []
    for a in PRIORITY_ORDER:
        if a in all_assocs:
            assoc_list.append(a)
    for a in sorted(all_assocs.keys()):
        if a not in assoc_list:
            assoc_list.append(a)

    n = len(assoc_list)
    assoc_idx = {a: i for i, a in enumerate(assoc_list)}

    # Build overlap matrix
    matrix = [[0] * n for _ in range(n)]
    for assocs in company_assocs:
        unique = list(set(assocs))
        for i, a in enumerate(unique):
            if a not in assoc_idx:
                continue
            ai = assoc_idx[a]
            matrix[ai][ai] += 1  # diagonal = total members
            for b in unique[i + 1:]:
                if b not in assoc_idx:
                    continue
                bi = assoc_idx[b]
                matrix[ai][bi] += 1
                matrix[bi][ai] += 1

    # Write header row (column headers = association codes)
    ws.cell(row=1, column=1, value="Association").fill = HEADER_FILL
    ws.cell(row=1, column=1).font = HEADER_FONT
    ws.cell(row=1, column=1).alignment = HEADER_ALIGNMENT
    ws.cell(row=1, column=1).border = THIN_BORDER
    ws.column_dimensions["A"].width = 14

    for col_idx, assoc in enumerate(assoc_list, 2):
        cell = ws.cell(row=1, column=col_idx, value=assoc)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[_col_letter(col_idx)].width = 10

    # "Total Members" column header
    total_col = n + 2
    cell = ws.cell(row=1, column=total_col, value="Total Members")
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = HEADER_ALIGNMENT
    cell.border = THIN_BORDER
    ws.column_dimensions[_col_letter(total_col)].width = 14

    # Write data rows
    for row_idx, assoc_row in enumerate(assoc_list, 2):
        ai = assoc_idx[assoc_row]

        # Row header
        cell = ws.cell(row=row_idx, column=1, value=assoc_row)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

        for col_idx, assoc_col in enumerate(assoc_list, 2):
            ci = assoc_idx[assoc_col]
            val = matrix[ai][ci]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

            if ai == ci:
                # Diagonal: bold, dark blue
                cell.fill = DIAG_FILL
                cell.font = DIAG_FONT
            elif val > 10:
                cell.fill = LIGHT_BLUE_FILL
                cell.font = Font(bold=True, color="1F4E79")

        # Total members column = diagonal value
        total_val = matrix[ai][ai]
        tcell = ws.cell(row=row_idx, column=total_col, value=total_val)
        tcell.alignment = Alignment(horizontal="center", vertical="center")
        tcell.border = THIN_BORDER
        tcell.font = Font(bold=True)

    # "Total Members" summary row at bottom
    total_row = len(assoc_list) + 2
    cell = ws.cell(row=total_row, column=1, value="Total Members")
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT
    cell.border = THIN_BORDER

    for col_idx, assoc_col in enumerate(assoc_list, 2):
        ci = assoc_idx[assoc_col]
        val = matrix[ci][ci]
        cell = ws.cell(row=total_row, column=col_idx, value=val)
        cell.fill = SECTION_FILL
        cell.font = Font(bold=True, color="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Grand total
    grand_total = sum(all_assocs.values())
    gtcell = ws.cell(row=total_row, column=total_col, value=grand_total)
    gtcell.fill = SECTION_FILL
    gtcell.font = Font(bold=True, color="1F4E79")
    gtcell.alignment = Alignment(horizontal="center", vertical="center")
    gtcell.border = THIN_BORDER

    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 30
    print(f"  Sheet 2: Association Overlap Matrix — {n} associations, {len(companies)} companies")


# ──────────────────────────────────────────────────────────────────────
# Sheet 3: Technology Maturity Index
# ──────────────────────────────────────────────────────────────────────

def build_tech_maturity(ws, companies: list[dict]) -> None:
    ws.title = "Technology Maturity Index"

    enriched = [c for c in companies if get_tech_stack(c) or c.get("cms") or c.get("email_provider")]
    total_enriched = len(enriched)

    # Compute maturity scores
    maturity_buckets: dict[str, list] = {
        "Modern": [],
        "Transitional": [],
        "Legacy": [],
        "Unknown": [],
    }

    for rec in companies:
        has_tech = bool(get_tech_stack(rec) or rec.get("cms"))
        score = compute_tech_maturity_score(rec) if has_tech else 0
        tier = get_tech_maturity_tier(score, has_tech)
        icp = compute_icp_score(rec)["icp_score"]
        maturity_buckets[tier].append({"rec": rec, "score": score, "icp": icp})

    # ── Section A: Summary Table ──────────────────────────────────────
    row = 1
    write_section_header(ws, row, "A. Technology Maturity Summary", col_span=5)
    row += 1

    SUMMARY_COLS = [
        ("Maturity Tier", 16),
        ("Count", 10),
        ("% of All", 12),
        ("Avg ICP Score", 14),
        ("Top Association", 20),
    ]
    style_header_row(ws, SUMMARY_COLS, row=row)
    row += 1

    tier_fills = {
        "Modern": MODERN_FILL,
        "Transitional": TRANS_FILL,
        "Legacy": LEGACY_FILL,
        "Unknown": UNKN_FILL,
    }
    tier_fonts = {
        "Modern": Font(bold=True, color="006100"),
        "Transitional": Font(bold=True, color="7D5700"),
        "Legacy": Font(bold=True, color="843C00"),
        "Unknown": Font(color="555555"),
    }

    total_all = len(companies)
    for tier in ["Modern", "Transitional", "Legacy", "Unknown"]:
        items = maturity_buckets[tier]
        count = len(items)
        pct = count / total_all * 100 if total_all else 0
        avg_icp = sum(x["icp"] for x in items) / count if count else 0
        # Top association
        assoc_counter: Counter = Counter()
        for x in items:
            for a in get_associations_list(x["rec"]):
                assoc_counter[a] += 1
        top_assoc = assoc_counter.most_common(1)[0][0] if assoc_counter else ""

        vals = [tier, count, f"{pct:.1f}%", round(avg_icp, 1), top_assoc]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=1).fill = tier_fills[tier]
        ws.cell(row=row, column=1).font = tier_fonts[tier]
        row += 1

    row += 1  # blank separator

    # ── Section B: Technology Adoption Table ─────────────────────────
    write_section_header(ws, row, "B. Technology Adoption by Category", col_span=4)
    row += 1

    TECH_COLS = [
        ("Technology", 30),
        ("Adoption Count", 14),
        ("% of Enriched", 13),
        ("Category", 18),
    ]
    style_header_row(ws, TECH_COLS, row=row)
    row += 1

    # Build tech counters from enriched records
    tech_counter: Counter = Counter()
    for rec in companies:
        for t in get_tech_stack(rec):
            tech_counter[t] += 1

    # Build category lookup (technology -> category)
    tech_to_cat: dict[str, str] = {}
    for cat, tech_list in TECH_CATEGORIES.items():
        for t in tech_list:
            tech_to_cat[t] = cat

    # Group by category, then by count descending
    rows_by_cat: dict[str, list] = defaultdict(list)
    for tech, count in tech_counter.items():
        cat = tech_to_cat.get(tech, "Other")
        rows_by_cat[cat].append((tech, count))

    CAT_ORDER = ["Analytics", "CMS", "CDN", "Framework", "Email/Marketing", "Security", "Hosting", "Other"]
    for cat in CAT_ORDER:
        if cat not in rows_by_cat:
            continue
        items_sorted = sorted(rows_by_cat[cat], key=lambda x: x[1], reverse=True)
        for tech, count in items_sorted:
            pct = count / total_enriched * 100 if total_enriched else 0
            vals = [tech, count, f"{pct:.1f}%", cat]
            for col_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "center",
                                           vertical="center")
            row += 1

    row += 1  # blank separator

    # ── Section C: CMS Distribution ──────────────────────────────────
    write_section_header(ws, row, "C. CMS Distribution", col_span=4)
    row += 1

    CMS_COLS = [
        ("CMS Platform", 25),
        ("Count", 12),
        ("% of Enriched", 13),
        ("Adoption Tier", 15),
    ]
    style_header_row(ws, CMS_COLS, row=row)
    row += 1

    cms_counter: Counter = Counter()
    for rec in companies:
        cms = (rec.get("cms") or "").strip()
        if cms:
            cms_counter[cms] += 1

    for cms, count in cms_counter.most_common():
        pct = count / total_enriched * 100 if total_enriched else 0
        if count >= 100:
            tier_label = "Dominant"
        elif count >= 30:
            tier_label = "Common"
        elif count >= 10:
            tier_label = "Niche"
        else:
            tier_label = "Rare"
        vals = [cms, count, f"{pct:.1f}%", tier_label]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

    ws.freeze_panes = "A3"
    print(f"  Sheet 3: Technology Maturity Index — {total_enriched} enriched records, "
          f"{len(tech_counter)} unique technologies")


# ──────────────────────────────────────────────────────────────────────
# Sheet 4: Email Infrastructure Analysis
# ──────────────────────────────────────────────────────────────────────

def build_email_infrastructure(ws, companies: list[dict]) -> None:
    ws.title = "Email Infrastructure Analysis"

    total = len(companies)

    # Pre-compute per-company data
    email_provider_data: dict[str, list[dict]] = defaultdict(list)
    spf_service_data: dict[str, list[dict]] = defaultdict(list)
    m365_records: list[dict] = []

    for rec in companies:
        ep = get_email_provider(rec) or "Unknown"
        email_provider_data[ep].append(rec)

        for svc in get_spf_list(rec):
            spf_service_data[svc].append(rec)

        if "microsoft 365" in ep.lower():
            m365_records.append(rec)

    # ── Section A: Email Provider Distribution ────────────────────────
    row = 1
    write_section_header(ws, row, "A. Email Provider Distribution", col_span=5)
    row += 1

    PROV_COLS = [
        ("Provider", 22),
        ("Count", 10),
        ("%", 8),
        ("Avg Employee Count", 18),
        ("Top Association", 22),
    ]
    style_header_row(ws, PROV_COLS, row=row)
    row += 1

    # Sort by count descending, "Unknown" last
    provider_sorted = sorted(
        email_provider_data.items(),
        key=lambda x: (x[0] == "Unknown", -len(x[1]))
    )

    for provider, recs in provider_sorted:
        count = len(recs)
        pct = count / total * 100 if total else 0

        emp_vals = []
        for r in recs:
            ec_min, ec_max = get_employee_count(r)
            mid = (ec_min + ec_max) / 2 if ec_max else ec_min
            if mid > 0:
                emp_vals.append(mid)
        avg_emp = int(sum(emp_vals) / len(emp_vals)) if emp_vals else 0
        avg_emp_str = str(avg_emp) if avg_emp else "N/A"

        assoc_c: Counter = Counter()
        for r in recs:
            for a in get_associations_list(r):
                assoc_c[a] += 1
        top_assoc = assoc_c.most_common(1)[0][0] if assoc_c else ""

        vals = [provider, count, f"{pct:.1f}%", avg_emp_str, top_assoc]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

    row += 1  # blank separator

    # ── Section B: SPF Services Distribution ─────────────────────────
    write_section_header(ws, row, "B. SPF Services Distribution", col_span=4)
    row += 1

    SPF_COLS = [
        ("Service", 22),
        ("Count", 10),
        ("%", 8),
        ("Category", 20),
    ]
    style_header_row(ws, SPF_COLS, row=row)
    row += 1

    spf_sorted = sorted(spf_service_data.items(), key=lambda x: -len(x[1]))
    enriched_count = sum(1 for r in companies if get_spf_list(r))

    for svc, recs in spf_sorted:
        count = len(recs)
        pct = count / total * 100 if total else 0
        cat = SPF_CATEGORIES.get(svc, "Other")
        vals = [svc, count, f"{pct:.1f}%", cat]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

    row += 1  # blank separator

    # ── Section C: Microsoft Ecosystem Analysis ───────────────────────
    write_section_header(ws, row, "C. Microsoft Ecosystem Analysis (Co-Sell Partnership Potential)", col_span=5)
    row += 1

    M365_COLS = [
        ("Segment", 35),
        ("Count", 10),
        ("% of M365 Users", 18),
        ("Notes", 35),
    ]
    style_header_row(ws, M365_COLS, row=row)
    row += 1

    m365_total = len(m365_records)

    # M365 + Salesforce
    m365_sfdc = [r for r in m365_records if any("salesforce" in s.lower() for s in get_spf_list(r))]
    # M365 + Marketing Automation
    ma_terms = ("hubspot", "marketo", "pardot", "mailchimp", "activecampaign", "constant contact")
    m365_ma = [r for r in m365_records if any(s.lower() in ma_terms for s in get_spf_list(r))]

    # M365 by size segment
    m365_small = [r for r in m365_records if (get_employee_count(r)[1] or get_employee_count(r)[0]) < 100]
    m365_mid = [r for r in m365_records
                if 100 <= (get_employee_count(r)[1] or get_employee_count(r)[0] or 0) <= 500]
    m365_large = [r for r in m365_records
                  if (get_employee_count(r)[1] or get_employee_count(r)[0] or 0) > 500]
    m365_unknown_size = [r for r in m365_records
                         if (get_employee_count(r)[1] or get_employee_count(r)[0]) == 0]

    m365_segments = [
        ("Total Microsoft 365 Users",
         m365_total, "100.0%",
         "Prime candidates for Microsoft Dynamics co-sell"),
        ("M365 + Salesforce CRM",
         len(m365_sfdc), f"{len(m365_sfdc)/m365_total*100:.1f}%" if m365_total else "0%",
         "Salesforce displacement or Dynamics + SFDC co-exist"),
        ("M365 + Marketing Automation",
         len(m365_ma), f"{len(m365_ma)/m365_total*100:.1f}%" if m365_total else "0%",
         "Tech-savvy buyers with existing cloud budget"),
        ("M365 Small (<100 employees)",
         len(m365_small), f"{len(m365_small)/m365_total*100:.1f}%" if m365_total else "0%",
         "Dynamics 365 Business Central target segment"),
        ("M365 Mid (100-500 employees)",
         len(m365_mid), f"{len(m365_mid)/m365_total*100:.1f}%" if m365_total else "0%",
         "Dynamics 365 F&O / Business Central sweet spot"),
        ("M365 Large (>500 employees)",
         len(m365_large), f"{len(m365_large)/m365_total*100:.1f}%" if m365_total else "0%",
         "Dynamics 365 F&O enterprise — longer sales cycle"),
        ("M365 Unknown Size",
         len(m365_unknown_size), f"{len(m365_unknown_size)/m365_total*100:.1f}%" if m365_total else "0%",
         "Needs enrichment before segmentation"),
    ]

    for seg_name, count, pct, notes in m365_segments:
        vals = [seg_name, count, pct, notes]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="left" if col_idx in (1, 4) else "center",
                vertical="center"
            )
        row += 1

    ws.freeze_panes = "A3"
    print(f"  Sheet 4: Email Infrastructure Analysis — {m365_total} M365 users identified")


# ──────────────────────────────────────────────────────────────────────
# Sheet 5: Competitive Pressure Map
# ──────────────────────────────────────────────────────────────────────

def build_competitive_pressure(ws, companies: list[dict]) -> None:
    ws.title = "Competitive Pressure Map"

    # Associations to analyze (rows)
    ASSOC_ROWS = ["PMA", "AGMA", "NEMA", "AIA", "SOCMA"]

    # Competitor columns
    COMP_COLS = [
        "Epicor", "SYSPRO", "Plex", "Acumatica",
        "Infor", "SAP", "Microsoft Dynamics",
        "Oracle", "Netsuite", "Sage", "QAD",
        "Other ERP",
    ]
    GREENFIELD_COL = "No ERP Detected"

    # Build lookup: for each association, count competitor presence
    assoc_comp_counts: dict[str, Counter] = {a: Counter() for a in ASSOC_ROWS}
    assoc_greenfield: dict[str, int] = {a: 0 for a in ASSOC_ROWS}

    # Normalize competitor column names to COMPETITOR_ALIASES keys
    comp_col_to_key = {
        "Epicor": "epicor",
        "SYSPRO": "syspro",
        "Plex": "plex",
        "Acumatica": "acumatica",
        "Infor": "infor",
        "SAP": "sap",
        "Microsoft Dynamics": "microsoft dynamics",
        "Oracle": "oracle",
        "Netsuite": "netsuite",
        "Sage": "sage",
        "QAD": "qad",
    }

    for rec in companies:
        rec_assocs = get_associations_list(rec)
        matched_assocs = [a for a in rec_assocs if a in ASSOC_ROWS]
        if not matched_assocs:
            continue

        detected = detect_competitor(rec)
        if not detected:
            for a in matched_assocs:
                assoc_greenfield[a] += 1
        else:
            # Match detected name to column
            det_lower = detected.lower()
            matched_col = "Other ERP"
            for col, key in comp_col_to_key.items():
                if key in det_lower or det_lower in key:
                    matched_col = col
                    break
            for a in matched_assocs:
                assoc_comp_counts[a][matched_col] += 1

    # ── Main Matrix ───────────────────────────────────────────────────
    row = 1
    write_section_header(ws, row, "Competitive Presence by Association (company count per competitor)", col_span=len(COMP_COLS) + 3)
    row += 1

    # Header row
    all_headers = ["Association"] + COMP_COLS + [GREENFIELD_COL, "Total Companies"]
    for col_idx, h in enumerate(all_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[_col_letter(col_idx)].width = 14 if col_idx == 1 else 12
    ws.freeze_panes = "B4"
    row += 1

    # Data rows
    assoc_totals = Counter()
    for rec in companies:
        for a in get_associations_list(rec):
            if a in ASSOC_ROWS:
                assoc_totals[a] += 1

    for assoc in ASSOC_ROWS:
        total_in_assoc = assoc_totals[assoc]
        row_vals = [assoc]
        for comp_col in COMP_COLS:
            row_vals.append(assoc_comp_counts[assoc].get(comp_col, 0))
        row_vals.append(assoc_greenfield[assoc])
        row_vals.append(total_in_assoc)

        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_idx == 1:
                cell.font = Font(bold=True)
            elif col_idx == len(COMP_COLS) + 2:  # Greenfield column
                if isinstance(val, int) and val > 0:
                    cell.fill = MODERN_FILL
                    cell.font = Font(bold=True, color="006100")
            elif isinstance(val, int) and val > 20:
                cell.fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
                cell.font = Font(bold=True, color="843C00")
        row += 1

    row += 2  # blank separator

    # ── Key Takeaways ─────────────────────────────────────────────────
    write_section_header(ws, row, "Key Takeaways by Association", col_span=len(COMP_COLS) + 3)
    row += 1

    TKWY_COLS = [
        ("Association", 14),
        ("Strongest Competitor", 20),
        ("Competitor Count", 16),
        ("Greenfield Count", 16),
        ("Greenfield %", 14),
        ("Recommended Strategy", 40),
    ]
    for col_idx, (h, w) in enumerate(TKWY_COLS, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[_col_letter(col_idx)].width = w
    row += 1

    for assoc in ASSOC_ROWS:
        total_in_assoc = assoc_totals[assoc]
        comp_counts = assoc_comp_counts[assoc]
        greenfield = assoc_greenfield[assoc]
        greenfield_pct = greenfield / total_in_assoc * 100 if total_in_assoc else 0

        if comp_counts:
            top_comp, top_comp_count = comp_counts.most_common(1)[0]
        else:
            top_comp, top_comp_count = "None detected", 0

        if greenfield_pct >= 70:
            strategy = f"Strong greenfield play — {greenfield_pct:.0f}% have no ERP. Lead with value + education."
        elif greenfield_pct >= 40:
            strategy = f"Mixed market — {top_comp} is strongest threat. Differentiate on service/price."
        else:
            strategy = f"Competitive market — counter {top_comp} with displacement campaign."

        vals = [assoc, top_comp, top_comp_count, greenfield, f"{greenfield_pct:.0f}%", strategy]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="left" if col_idx in (2, 6) else "center",
                vertical="center",
                wrap_text=col_idx == 6,
            )
        row += 1

    ws.row_dimensions[1].height = 20
    print(f"  Sheet 5: Competitive Pressure Map — {len(ASSOC_ROWS)} associations analyzed")


# ──────────────────────────────────────────────────────────────────────
# Sheet 6: Market Opportunity Summary
# ──────────────────────────────────────────────────────────────────────

def build_market_opportunity(ws, companies: list[dict], assoc_config: dict) -> None:
    ws.title = "Market Opportunity Summary"

    # Pre-compute aggregate metrics
    total_companies = len(companies)
    icp_scores = [compute_icp_score(r)["icp_score"] for r in companies]
    avg_icp = sum(icp_scores) / total_companies if total_companies else 0

    # TAM computation
    total_tam = sum(
        cfg.get("expected_members", 0)
        for cfg in assoc_config.values()
    )
    penetration_pct = total_companies / total_tam * 100 if total_tam else 0

    # Greenfield count
    greenfield_count = sum(1 for r in companies if not detect_competitor(r))

    # M365 count
    m365_count = sum(1 for r in companies if "microsoft 365" in get_email_provider(r).lower())

    # State aggregates for top opportunities
    state_data: dict[str, dict] = defaultdict(lambda: {"count": 0, "icp_scores": []})
    for rec, score in zip(companies, icp_scores):
        state = (rec.get("state") or "").strip().upper()
        if not state or len(state) != 2:
            continue
        state_data[state]["count"] += 1
        state_data[state]["icp_scores"].append(score)

    state_opportunity = []
    for state, d in state_data.items():
        c = d["count"]
        avg = sum(d["icp_scores"]) / c if c else 0
        state_opportunity.append((state, c, round(avg, 1)))
    state_opportunity.sort(key=lambda x: (-x[1], -x[2]))

    row = 1

    # ── Dashboard Title ───────────────────────────────────────────────
    title_cell = ws.cell(row=row, column=1, value="GSS Market Opportunity Dashboard")
    title_cell.font = Font(bold=True, size=16, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 28
    row += 1

    subtitle_cell = ws.cell(row=row, column=1, value=f"NAM Intelligence Pipeline — 14 Associations — {total_companies:,} Companies")
    subtitle_cell.font = SUBTITLE_FONT
    row += 2

    # ── Section A: Total Addressable Market ───────────────────────────
    write_section_header(ws, row, "A. Total Addressable Market (TAM)", col_span=4)
    row += 1

    TAM_COLS = [("Metric", 30), ("Value", 20), ("Notes", 45)]
    for col_idx, (h, w) in enumerate(TAM_COLS, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[_col_letter(col_idx)].width = w
    row += 1

    tam_rows = [
        ("Total TAM (14 associations)", f"{total_tam:,}", "Sum of expected_members from associations.yaml"),
        ("Current Coverage", f"{total_companies:,}", "Unique companies in pipeline"),
        ("Penetration %", f"{penetration_pct:.1f}%", "Coverage / TAM"),
        ("Remaining Opportunity", f"{max(0, total_tam - total_companies):,}", "Companies not yet captured"),
        ("Avg ICP Score", f"{avg_icp:.1f}/100", "Across all current companies"),
        ("High-Quality (ICP >= 60)", f"{sum(1 for s in icp_scores if s >= 60):,}", "Priority sales targets"),
    ]
    for metric, value, notes in tam_rows:
        ws.cell(row=row, column=1, value=metric).border = THIN_BORDER
        val_cell = ws.cell(row=row, column=2, value=value)
        val_cell.border = THIN_BORDER
        val_cell.font = METRIC_FONT
        val_cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=notes).border = THIN_BORDER
        row += 1

    row += 1

    # ── Section B: Top 5 States by Opportunity ────────────────────────
    write_section_header(ws, row, "B. Top 5 States by Opportunity", col_span=4)
    row += 1

    STATE_COLS = [("State", 10), ("Companies", 12), ("Avg ICP Score", 15), ("Market Classification", 20)]
    for col_idx, (h, w) in enumerate(STATE_COLS, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[_col_letter(col_idx)].width = w
    row += 1

    for state, count, avg in state_opportunity[:5]:
        classification = classify_market(count, avg)
        vals = [state, count, avg, classification]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

    row += 1

    # ── Section C: Top Competitive Gaps (Greenfield) ──────────────────
    write_section_header(ws, row, "C. Top Competitive Gaps (Greenfield ERP Opportunities)", col_span=4)
    row += 1

    GAP_COLS = [("Association", 14), ("Total Companies", 16), ("Greenfield Count", 16), ("Greenfield %", 14)]
    for col_idx, (h, w) in enumerate(GAP_COLS, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[_col_letter(col_idx)].width = w
    row += 1

    ASSOC_ROWS = ["PMA", "AGMA", "NEMA", "AIA", "SOCMA"]
    assoc_totals: Counter = Counter()
    assoc_greenfield_counts: Counter = Counter()
    for rec in companies:
        for a in get_associations_list(rec):
            if a in ASSOC_ROWS:
                assoc_totals[a] += 1
                if not detect_competitor(rec):
                    assoc_greenfield_counts[a] += 1

    gap_rows = []
    for assoc in ASSOC_ROWS:
        total_a = assoc_totals[assoc]
        gf = assoc_greenfield_counts[assoc]
        gf_pct = gf / total_a * 100 if total_a else 0
        gap_rows.append((assoc, total_a, gf, gf_pct))
    gap_rows.sort(key=lambda x: -x[3])

    for assoc, total_a, gf, gf_pct in gap_rows:
        vals = [assoc, total_a, gf, f"{gf_pct:.0f}%"]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
        if gf_pct >= 70:
            ws.cell(row=row, column=4).fill = MODERN_FILL
            ws.cell(row=row, column=4).font = Font(bold=True, color="006100")
        row += 1

    row += 1

    # ── Section D: Investment Priorities ─────────────────────────────
    write_section_header(ws, row, "D. Investment Priorities (Recommended Next Actions)", col_span=4)
    row += 1

    PRI_COLS = [("Priority", 10), ("Action", 50), ("Impact", 15), ("Effort", 12)]
    for col_idx, (h, w) in enumerate(PRI_COLS, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[_col_letter(col_idx)].width = w
    row += 1

    priorities = [
        (1, "Extract remaining 9 associations (~5,000 new companies)", "High", "Medium"),
        (2, "Enrich PMA records with website/domain data (964 records need enrichment)", "High", "Medium"),
        (3, "Procure API keys: BuiltWith Free + Hunter.io for deeper firmographic enrichment", "Medium", "Low"),
        (4, f"Deploy Microsoft co-sell partnership ({m365_count:,} M365 users identified)", "High", "Low"),
        (5, f"Target greenfield ERP opportunities ({greenfield_count:,} companies with no ERP detected)", "High", "Low"),
    ]

    impact_fills = {
        "High": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "Medium": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "Low": PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid"),
    }

    for pri, action, impact, effort in priorities:
        vals = [pri, action, impact, effort]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center" if col_idx in (1, 3, 4) else "left",
                vertical="center",
                wrap_text=col_idx == 2,
            )
        ws.cell(row=row, column=3).fill = impact_fills.get(impact, UNKN_FILL)
        ws.row_dimensions[row].height = 24
        row += 1

    ws.freeze_panes = "A4"
    print(f"  Sheet 6: Market Opportunity Summary — TAM: {total_tam:,}, "
          f"Coverage: {penetration_pct:.1f}%, Greenfield: {greenfield_count:,}")


# ──────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────

def main() -> None:
    import os
    os.makedirs(EXPORTS_DIR, exist_ok=True)

    print("Loading data...")
    companies = load_and_merge_data()
    assoc_config = load_associations_config()
    print(f"  Loaded {len(companies)} companies")
    print(f"  Loaded {len(assoc_config)} association configs")

    print("\nBuilding GSS_Market_Intelligence_Brief.xlsx...")
    wb = Workbook()
    # Remove the default empty sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    ws1 = wb.create_sheet("Geographic Density")
    build_geographic_density(ws1, companies)

    ws2 = wb.create_sheet("Association Overlap Matrix")
    build_association_overlap(ws2, companies)

    ws3 = wb.create_sheet("Technology Maturity Index")
    build_tech_maturity(ws3, companies)

    ws4 = wb.create_sheet("Email Infrastructure Analysis")
    build_email_infrastructure(ws4, companies)

    ws5 = wb.create_sheet("Competitive Pressure Map")
    build_competitive_pressure(ws5, companies)

    ws6 = wb.create_sheet("Market Opportunity Summary")
    build_market_opportunity(ws6, companies, assoc_config)

    wb.save(str(OUTPUT_PATH))
    size_kb = OUTPUT_PATH.stat().st_size / 1024
    print(f"\nSaved: {OUTPUT_PATH} ({size_kb:.0f} KB)")
    print("Done.")


if __name__ == "__main__":
    main()
