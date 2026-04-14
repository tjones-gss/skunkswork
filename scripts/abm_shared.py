"""
ABM Shared Module
NAM Intelligence Pipeline

Shared data loading, ICP scoring, and Excel styling used by all ABM scripts.
"""

import csv
import json
import re
from pathlib import Path

import yaml
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── Paths ─────────────────────────────────────────────────────────────

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
EXPORTS_DIR = DATA_DIR / "exports"
ENRICHED_PATH = DATA_DIR / "processed" / "enriched_all.jsonl"
CSV_PATH = EXPORTS_DIR / "companies_all.csv"
EVENTS_PATH = EXPORTS_DIR / "events_2026.csv"
COMPETITORS_PATH = EXPORTS_DIR / "competitor_analysis.csv"
ASSOCIATIONS_PATH = BASE_DIR / "config" / "associations.yaml"

# ── Excel Styling Constants ───────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

GRADE_FILLS = {
    "A": PatternFill(start_color="006100", end_color="006100", fill_type="solid"),
    "B": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "C": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "D": PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid"),
    "F": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}

TIER_FILLS = {
    "Tier 1": PatternFill(start_color="006100", end_color="006100", fill_type="solid"),
    "Tier 2": PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid"),
    "Tier 3": PatternFill(start_color="C65911", end_color="C65911", fill_type="solid"),
}
TIER_FONTS = {
    "Tier 1": Font(bold=True, color="FFFFFF"),
    "Tier 2": Font(bold=True, color="FFFFFF"),
    "Tier 3": Font(bold=True, color="FFFFFF"),
}

THREAT_FILLS = {
    "HIGH": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
    "MEDIUM": PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid"),
    "LOW": PatternFill(start_color="44AA44", end_color="44AA44", fill_type="solid"),
    "EMERGING": PatternFill(start_color="4488CC", end_color="4488CC", fill_type="solid"),
}
WHITE_FONT = Font(bold=True, color="FFFFFF")

SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(bold=True, size=12, color="1F4E79")

SUBTITLE_FONT = Font(italic=True, size=10, color="555555")
METRIC_FONT = Font(bold=True, size=14, color="1F4E79")
METRIC_LARGE_FONT = Font(bold=True, size=24, color="1F4E79")

# ── Competitor Data ───────────────────────────────────────────────────

COMPETITOR_ALIASES = {
    "epicor": ["epicor", "epicor erp", "epicor prophet 21", "epicor kinetic"],
    "odoo": ["odoo", "odoo erp", "odoo manufacturing"],
    "infor": ["infor", "infor cloudsuite", "infor syteline", "infor m3", "infor ln"],
    "iqms": ["iqms", "iqms erp", "delmiaworks"],
    "jobboss": ["jobboss", "job boss", "jobboss erp"],
    "syspro": ["syspro", "syspro erp"],
    "acumatica": ["acumatica", "acumatica erp", "acumatica cloud erp"],
    "microsoft dynamics": ["dynamics 365", "microsoft dynamics", "d365", "navision"],
    "aptean": ["aptean", "aptean industrial", "aptean erp"],
    "fulcrum": ["fulcrum", "fulcrum erp"],
    "sap": ["sap", "sap erp", "sap business one", "sap s/4hana", "sap b1"],
    "oracle": ["oracle", "oracle erp", "jd edwards", "jde"],
    "netsuite": ["netsuite", "oracle netsuite"],
    "plex": ["plex", "plex systems", "plex manufacturing cloud"],
    "sage": ["sage", "sage 100", "sage 300", "sage x3", "sage intacct"],
    "qad": ["qad", "qad erp"],
}

THREAT_RESPONSE = {
    "HIGH": "Must Counter",
    "MEDIUM": "Monitor Closely",
    "LOW": "Track Only",
    "EMERGING": "Watch & Prepare",
}

# ── Manufacturing States (ranked by manufacturing GDP) ────────────────

HIGH_MFG_STATES = {"TX", "OH", "MI", "IN", "IL", "PA", "WI", "CA", "NC", "TN"}
MED_MFG_STATES = {"AL", "SC", "GA", "MN", "MO", "IA", "KY", "MS", "NY", "VA", "CT", "OR", "WA", "NJ"}

# ── Data Loading ──────────────────────────────────────────────────────


def _normalize(name: str) -> str:
    """Normalize company name for merge key."""
    return re.sub(r"[^a-z0-9]", "", name.lower()) if name else ""


def _parse_csv_contacts(raw: str) -> list[dict]:
    """Parse CSV contact string like 'Name <email>; Name2 <email2>'."""
    contacts = []
    if not raw:
        return contacts
    for part in raw.split(";"):
        part = part.strip()
        if not part:
            continue
        m = re.match(r"^(.+?)\s*<([^>]+)>$", part)
        if m:
            contacts.append({"name": m.group(1).strip(), "email": m.group(2).strip()})
        else:
            contacts.append({"name": part, "email": ""})
    return contacts


def load_enriched(path: Path = ENRICHED_PATH) -> list[dict]:
    """Load enriched JSONL records."""
    records = []
    if not path.exists():
        print(f"Warning: enriched file not found: {path}")
        return records
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    records.append(json.loads(line))
                except json.JSONDecodeError:
                    continue
    return records


def load_csv_records(path: Path = CSV_PATH) -> list[dict]:
    """Load companies CSV records with type conversion."""
    records = []
    if not path.exists():
        print(f"Warning: CSV file not found: {path}")
        return records
    with open(path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                row["quality_score"] = int(row.get("quality_score", 0))
            except (ValueError, TypeError):
                row["quality_score"] = 0

            contacts_raw = row.get("contacts", "")
            if contacts_raw and isinstance(contacts_raw, str):
                row["contacts"] = _parse_csv_contacts(contacts_raw)
            elif not contacts_raw:
                row["contacts"] = []

            ts = row.get("tech_stack", "")
            if isinstance(ts, str) and ts:
                row["tech_stack"] = [t.strip() for t in ts.replace(",", ";").split(";") if t.strip()]
            else:
                row["tech_stack"] = []

            spf = row.get("spf_services", "")
            if isinstance(spf, str) and spf:
                row["spf_services"] = [s.strip() for s in spf.replace(",", ";").split(";") if s.strip()]
            else:
                row["spf_services"] = []

            # Parse employee counts
            for field in ("employee_count_min", "employee_count_max"):
                try:
                    row[field] = int(row.get(field, 0) or 0)
                except (ValueError, TypeError):
                    row[field] = 0

            records.append(row)
    return records


def merge_records(enriched: list[dict], csv_records: list[dict]) -> list[dict]:
    """Merge enriched JSONL and companies CSV by normalized company_name."""
    enriched_map = {}
    for rec in enriched:
        key = _normalize(rec.get("company_name", ""))
        if key:
            enriched_map[key] = rec

    merged = []
    seen_keys = set()

    for csv_rec in csv_records:
        key = _normalize(csv_rec.get("company_name", ""))
        if not key:
            continue
        enriched_rec = enriched_map.get(key)
        if enriched_rec:
            combined = dict(enriched_rec)
            combined["quality_score"] = csv_rec.get("quality_score", 0)
            combined["quality_grade"] = csv_rec.get("quality_grade", "F")
            if csv_rec.get("associations"):
                combined["associations"] = csv_rec["associations"]
            if not combined.get("contacts") and csv_rec.get("contacts"):
                combined["contacts"] = csv_rec["contacts"]
            for field in ["city", "state", "country", "phone", "industry", "source_url",
                          "member_type", "notes", "street", "zip_code",
                          "employee_count_min", "employee_count_max"]:
                if not combined.get(field) and csv_rec.get(field):
                    combined[field] = csv_rec[field]
        else:
            combined = dict(csv_rec)
        seen_keys.add(key)
        merged.append(combined)

    for rec in enriched:
        key = _normalize(rec.get("company_name", ""))
        if key and key not in seen_keys:
            if "quality_score" not in rec:
                rec["quality_score"] = 50
            if "quality_grade" not in rec:
                rec["quality_grade"] = "D"
            merged.append(rec)
            seen_keys.add(key)

    return merged


def load_and_merge_data() -> list[dict]:
    """Load and merge enriched + CSV data. Returns merged list of company dicts."""
    enriched = load_enriched()
    csv_records = load_csv_records()
    return merge_records(enriched, csv_records)


def load_events() -> list[dict]:
    """Load events CSV."""
    events = []
    if not EVENTS_PATH.exists():
        return events
    with open(EVENTS_PATH, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            events.append(row)
    return events


def load_competitors() -> list[dict]:
    """Load competitor analysis CSV."""
    comps = []
    if not COMPETITORS_PATH.exists():
        return comps
    with open(COMPETITORS_PATH, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            comps.append(row)
    return comps


def load_associations_config() -> dict:
    """Load associations.yaml config."""
    if not ASSOCIATIONS_PATH.exists():
        return {}
    with open(ASSOCIATIONS_PATH, encoding="utf-8") as f:
        data = yaml.safe_load(f)
    return data.get("associations", {})


# ── Helper Functions ──────────────────────────────────────────────────


def get_email_provider(rec: dict) -> str:
    """Get normalized email provider."""
    return (rec.get("email_provider") or "").strip()


def get_spf_list(rec: dict) -> list[str]:
    """Get spf_services as a list."""
    spf = rec.get("spf_services", [])
    if isinstance(spf, list):
        return [s.strip() for s in spf if s and str(s).strip()]
    if isinstance(spf, str) and spf.strip():
        return [s.strip() for s in spf.split(";") if s.strip()]
    return []


def get_tech_stack(rec: dict) -> list[str]:
    """Get tech_stack as a list."""
    ts = rec.get("tech_stack", [])
    if isinstance(ts, list):
        return ts
    if isinstance(ts, str) and ts.strip():
        return [t.strip() for t in ts.replace(",", ";").split(";") if t.strip()]
    return []


def get_contacts(rec: dict) -> list[dict]:
    """Get contacts as a list of dicts."""
    contacts = rec.get("contacts", [])
    if isinstance(contacts, list):
        return contacts
    return []


def get_primary_contact(rec: dict) -> tuple[str, str]:
    """Return (name, email) of first contact."""
    contacts = get_contacts(rec)
    if contacts:
        c = contacts[0]
        if isinstance(c, dict):
            return c.get("name", ""), c.get("email", "")
    return "", ""


def get_associations_list(rec: dict) -> list[str]:
    """Parse associations field into a list."""
    assoc = rec.get("associations", "") or rec.get("association", "")
    if isinstance(assoc, list):
        return [a.strip() for a in assoc if a and str(a).strip()]
    if isinstance(assoc, str) and assoc.strip():
        return [a.strip() for a in assoc.replace(",", ";").split(";") if a.strip()]
    return []


def has_website(rec: dict) -> bool:
    return bool((rec.get("website") or rec.get("domain") or "").strip())


def get_employee_count(rec: dict) -> tuple[int, int]:
    """Return (min, max) employee counts."""
    try:
        ec_min = int(rec.get("employee_count_min", 0) or 0)
    except (ValueError, TypeError):
        ec_min = 0
    try:
        ec_max = int(rec.get("employee_count_max", 0) or 0)
    except (ValueError, TypeError):
        ec_max = 0
    return ec_min, ec_max


# ── ICP Scoring Model ─────────────────────────────────────────────────


def compute_icp_score(rec: dict) -> dict:
    """
    Compute ICP (Ideal Customer Profile) score for a company record.

    Returns dict with total score and breakdown per dimension:
    {
        "icp_score": int (0-100),
        "tech_maturity": int (0-25),
        "size_fit": int (0-20),
        "geo_fit": int (0-15),
        "assoc_engagement": int (0-15),
        "tech_gap": int (0-15),
        "data_quality": int (0-10),
    }
    """

    # 1. Tech Maturity Signal (0-25)
    tech_maturity = 0
    ep = get_email_provider(rec).lower()
    spf = [s.lower() for s in get_spf_list(rec)]
    ts = [t.lower() for t in get_tech_stack(rec)]

    # Email provider signals
    if "microsoft 365" in ep:
        tech_maturity += 10  # M365 ecosystem = Dynamics displacement opportunity
    elif "proofpoint" in ep or "mimecast" in ep:
        tech_maturity += 8   # Enterprise security = budget available
    elif "google workspace" in ep:
        tech_maturity += 5   # Cloud, but Google ecosystem
    elif ep in ("self-hosted", "self-hosted (on-premise)", "other"):
        tech_maturity += 12  # Legacy = modernization opportunity
    elif not ep:
        tech_maturity += 3   # Unknown

    # SPF service signals
    if any(s in ("salesforce", "pardot") for s in spf):
        tech_maturity += 5   # CRM investment = budget for ERP
    elif any(s in ("hubspot", "marketo", "activecampaign") for s in spf):
        tech_maturity += 3   # Marketing automation = tech-savvy
    if any(s in ("mailchimp", "constant contact", "sendgrid", "amazon ses") for s in spf):
        tech_maturity += 2   # Email marketing

    # CDN/security signals
    if any(t in ("cloudflare", "akamai", "amazon cloudfront") for t in ts):
        tech_maturity += 2   # CDN = web investment

    tech_maturity = min(tech_maturity, 25)

    # 2. Company Size Fit (0-20) — GSS sweet spot: 50-500 employees
    size_fit = 0
    ec_min, ec_max = get_employee_count(rec)
    emp = ec_max or ec_min

    if 50 <= emp <= 500:
        size_fit = 20    # Sweet spot
    elif 25 <= emp < 50:
        size_fit = 15    # Slightly small
    elif 500 < emp <= 1000:
        size_fit = 12    # Getting large
    elif 10 <= emp < 25:
        size_fit = 8     # Small shop
    elif emp > 1000:
        size_fit = 5     # Enterprise (too large for GSS typical deal)
    elif emp == 0:
        size_fit = 10    # Unknown — benefit of doubt

    # 3. Geographic Fit (0-15) — US manufacturing states weighted
    geo_fit = 0
    state = (rec.get("state") or "").strip().upper()

    if state in HIGH_MFG_STATES:
        geo_fit = 15
    elif state in MED_MFG_STATES:
        geo_fit = 12
    elif state and len(state) == 2:
        geo_fit = 10     # Other US state
    elif state:
        geo_fit = 5      # International (Canada, Mexico)
    else:
        geo_fit = 7      # Unknown

    # 4. Association Engagement (0-15) — multi-membership = more engaged
    assoc_engagement = 0
    assocs = get_associations_list(rec)
    high_priority = {"PMA", "NEMA", "AGMA", "AIA", "SOCMA"}

    if len(assocs) >= 3:
        assoc_engagement = 15
    elif len(assocs) == 2:
        assoc_engagement = 12
    elif len(assocs) == 1:
        if assocs[0].upper() in high_priority:
            assoc_engagement = 10
        else:
            assoc_engagement = 7
    # else: 0

    # 5. Tech Stack Gap (0-15) — no ERP = greenfield opportunity
    tech_gap = 0
    erp = (rec.get("erp_system") or "").strip().lower()
    all_tech_text = " ".join(ts + spf + [erp])

    # Check if any known competitor ERP is detected
    competitor_detected = False
    for comp_key, aliases in COMPETITOR_ALIASES.items():
        if comp_key == "global shop solutions":
            continue  # Our product — skip
        for alias in aliases:
            if alias in all_tech_text:
                competitor_detected = True
                break
        if competitor_detected:
            break

    if not erp and not competitor_detected:
        tech_gap = 15    # Greenfield — no ERP detected
    elif competitor_detected:
        tech_gap = 5     # Displacement opportunity
    elif "global shop" in all_tech_text:
        tech_gap = 0     # Already our customer

    # 6. Data Quality Bonus (0-10)
    data_quality = 0
    contacts = get_contacts(rec)
    if contacts:
        data_quality += 4
    if rec.get("phone"):
        data_quality += 2
    if rec.get("enrichment_status") == "complete":
        data_quality += 2
    qs = rec.get("quality_score", 0)
    if isinstance(qs, (int, float)) and qs >= 80:
        data_quality += 2
    data_quality = min(data_quality, 10)

    total = tech_maturity + size_fit + geo_fit + assoc_engagement + tech_gap + data_quality

    return {
        "icp_score": min(total, 100),
        "tech_maturity": tech_maturity,
        "size_fit": size_fit,
        "geo_fit": geo_fit,
        "assoc_engagement": assoc_engagement,
        "tech_gap": tech_gap,
        "data_quality": data_quality,
    }


def assign_tier(icp_score: int, has_contacts: bool, quality_score: int) -> str:
    """Assign ABM tier based on ICP score and data readiness."""
    if icp_score >= 75 and has_contacts and quality_score >= 70:
        return "Tier 1"
    if icp_score >= 55:
        return "Tier 2"
    if icp_score >= 35:
        return "Tier 3"
    return "Unqualified"


def detect_competitor(rec: dict) -> str:
    """Detect if a known competitor product is present in tech stack/erp/spf."""
    erp = (rec.get("erp_system") or "").strip().lower()
    ts = [t.lower() for t in get_tech_stack(rec)]
    spf = [s.lower() for s in get_spf_list(rec)]
    all_text = " ".join(ts + spf + [erp])

    for comp_key, aliases in COMPETITOR_ALIASES.items():
        if comp_key == "global shop solutions":
            continue
        for alias in aliases:
            if alias in all_text:
                return comp_key.title()
    return ""


# ── Excel Helper Functions ────────────────────────────────────────────


def style_header_row(ws, columns: list[tuple], row: int = 1):
    """Style a header row with standard ABM formatting.

    columns: list of (header_text, width) tuples
    """
    for col_idx, (header, width) in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[_col_letter(col_idx)].width = width
    ws.freeze_panes = f"A{row + 1}"
    ws.auto_filter.ref = ws.dimensions


def write_section_header(ws, row: int, text: str, col_span: int = 10):
    """Write a section header row (blue background, large font)."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(2, col_span + 1):
        ws.cell(row=row, column=col).fill = SECTION_FILL


def _col_letter(col_idx: int) -> str:
    """Convert 1-based column index to letter."""
    from openpyxl.utils import get_column_letter
    return get_column_letter(col_idx)
