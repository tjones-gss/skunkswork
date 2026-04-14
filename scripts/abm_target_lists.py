#!/usr/bin/env python3
"""
ABM Target Lists
NAM Intelligence Pipeline

Generates: data/exports/GSS_ABM_Target_Lists.xlsx
  Sheet 1: Tier 1 — Strategic (Top 50) — full detail + campaign assignment
  Sheet 2: Tier 2 — Growth (200) — full detail + enrichment gaps
  Sheet 3: Tier 3 — Nurture — trimmed columns + enrichment gaps
  Sheet 4: Account Assignment Matrix — for sales managers
  Sheet 5: Tier Summary Statistics — distribution + gap analysis tables
"""

import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from scripts.abm_shared import (
    BASE_DIR, EXPORTS_DIR,
    HEADER_FILL, HEADER_FONT, HEADER_ALIGNMENT, THIN_BORDER,
    GRADE_FILLS, TIER_FILLS, TIER_FONTS,
    SECTION_FILL, SECTION_FONT, SUBTITLE_FONT,
    load_and_merge_data, load_events, load_competitors,
    get_email_provider, get_spf_list, get_tech_stack, get_associations_list,
    get_primary_contact, get_contacts, get_employee_count, has_website,
    compute_icp_score, assign_tier, detect_competitor,
    style_header_row, write_section_header,
)

OUTPUT_PATH = EXPORTS_DIR / "GSS_ABM_Target_Lists.xlsx"

# ── Column definitions ─────────────────────────────────────────────────

# Tier 1 / Tier 2 shared columns (header text, width)
FULL_COLUMNS = [
    ("Rank", 5),
    ("Company Name", 35),
    ("Website", 30),
    ("Domain", 20),
    ("City", 18),
    ("State", 8),
    ("Association(s)", 15),
    ("ICP Score", 12),
    ("ICP Tier", 12),
    ("Email Provider", 18),
    ("Employee Count", 15),
    ("ERP System", 15),
    ("Tech Stack", 40),
    ("SPF Services", 30),
    ("Primary Contact Name", 25),
    ("Contact Title", 25),
    ("Contact Email", 30),
    ("Contact Phone", 18),
    ("Quality Score", 12),
    ("Recommended Campaign", 25),
    ("Event Overlap", 30),
    ("Competitor Presence", 20),
]

TIER2_EXTRA_COLUMNS = FULL_COLUMNS + [("Enrichment Gaps", 30)]

NURTURE_COLUMNS = [
    ("Rank", 5),
    ("Company Name", 35),
    ("Website", 30),
    ("City", 18),
    ("State", 8),
    ("Association", 15),
    ("ICP Score", 12),
    ("Tier", 12),
    ("Email Provider", 18),
    ("Quality Score", 12),
    ("Enrichment Gaps", 30),
]

ASSIGNMENT_COLUMNS = [
    ("Company Name", 35),
    ("ICP Tier", 12),
    ("ICP Score", 12),
    ("State", 8),
    ("Association", 12),
    ("Assigned Rep", 20),
    ("Campaign Stage", 20),
    ("Next Action Date", 18),
    ("Notes", 40),
]

# Yellow fill for "to be filled" columns in Assignment Matrix
FILL_PENDING = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# ── Association → event mapping ────────────────────────────────────────

# Maps association code prefixes to relevant event name fragments
ASSOC_EVENT_MAP = {
    "PMA": ["FABTECH", "PMA Forming", "METALCON", "IMTS"],
    "NEMA": ["NEMA Annual", "NEMA Premium", "IPC APEX", "Automate"],
    "AGMA": ["AGMA Annual", "IMTS", "FABTECH"],
    "AIA": ["MD&M West", "RAPID+TCT", "Automate"],
    "SOCMA": ["SOCMA Show", "PACK EXPO"],
    "NTMA": ["IMTS", "FABTECH", "PMPA"],
    "PMPA": ["PMPA National", "IMTS"],
    "AMT": ["AMT MFG", "IMTS"],
    "FIA": ["FABTECH", "METALCON"],
    "NADCA": ["METALCON", "Ceramics Expo"],
    "AFS": ["Ceramics Expo", "IMTS"],
    "FMA": ["FABTECH", "METALCON", "IMTS"],
    "PMMI": ["PACK EXPO", "IMTS"],
    "PLASTICS": ["NPE", "IMTS"],
}

# ── Helper functions ───────────────────────────────────────────────────


def _col_letter(col_idx: int) -> str:
    return get_column_letter(col_idx)


def _emp_display(rec: dict) -> str:
    ec_min, ec_max = get_employee_count(rec)
    if ec_min and ec_max and ec_min != ec_max:
        return f"{ec_min:,}-{ec_max:,}"
    if ec_max:
        return f"{ec_max:,}"
    if ec_min:
        return f"{ec_min:,}+"
    return ""


def _recommend_campaign(rec: dict, icp: dict) -> str:
    """Auto-assign recommended campaign based on strongest ICP dimension."""
    ep = get_email_provider(rec).lower()
    competitor = detect_competitor(rec)

    # Competitor takes priority
    if competitor:
        return "Competitive Displacement"

    # Evaluate dimension percentages (score / max_score)
    dims = {
        "tech_maturity": icp["tech_maturity"] / 25,
        "tech_gap":      icp["tech_gap"] / 15,
        "assoc_engagement": icp["assoc_engagement"] / 15,
        "size_fit":      icp["size_fit"] / 20,
        "geo_fit":       icp["geo_fit"] / 15,
    }
    top = max(dims, key=dims.get)

    if top == "tech_maturity":
        if "microsoft" in ep or "m365" in ep:
            return "Microsoft Ecosystem"
        if ep in ("self-hosted", "self-hosted (on-premise)", "other"):
            return "Legacy Transformation"
        return "Tech Modernization"
    if top == "tech_gap" and icp["tech_gap"] >= 12:
        return "Greenfield ERP"
    if top == "assoc_engagement" and icp["assoc_engagement"] >= 10:
        return "Association Blitz"
    return "Multi-touch Nurture"


def _event_overlap(rec: dict, events: list[dict]) -> str:
    """Return semicolon-joined list of relevant events for this account."""
    assocs = get_associations_list(rec)
    if not assocs or not events:
        return ""

    relevant_fragments: list[str] = []
    for assoc in assocs:
        upper = assoc.strip().upper()
        # Match by exact prefix or partial key
        for key, fragments in ASSOC_EVENT_MAP.items():
            if upper == key or upper.startswith(key):
                relevant_fragments.extend(fragments)

    if not relevant_fragments:
        # Default: all HIGH-priority general manufacturing events
        relevant_fragments = ["IMTS", "FABTECH"]

    matched: list[str] = []
    seen: set[str] = set()
    for evt in events:
        name = evt.get("event_name", "")
        for frag in relevant_fragments:
            if frag.lower() in name.lower() and name not in seen:
                matched.append(name)
                seen.add(name)
                break

    return "; ".join(matched) if matched else ""


def _enrichment_gaps(rec: dict) -> str:
    """Return human-readable list of missing enrichment fields."""
    gaps = []
    contacts = get_contacts(rec)
    if not contacts:
        gaps.append("contacts")
    ec_min, ec_max = get_employee_count(rec)
    if not ec_min and not ec_max:
        gaps.append("employee count")
    ts = get_tech_stack(rec)
    if not ts:
        gaps.append("tech stack")
    ep = get_email_provider(rec)
    if not ep:
        gaps.append("email provider")
    if not (rec.get("phone") or ""):
        gaps.append("phone")
    if not has_website(rec):
        gaps.append("website")
    return "needs: " + ", ".join(gaps) if gaps else ""


def _write_data_row(
    ws,
    row_idx: int,
    values: list,
    tier: str | None = None,
    quality_score: int = 0,
    quality_grade: str = "",
    tier_col_idx: int | None = None,
    grade_col_idx: int | None = None,
):
    """Write a data row with standard formatting + optional tier/grade coloring."""
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=False)

        # Hyperlink for website-looking values
        if isinstance(val, str) and val.startswith("http"):
            cell.hyperlink = val
            cell.font = Font(color="0563C1", underline="single")

    # Tier column coloring
    if tier_col_idx and tier and tier in TIER_FILLS:
        tc = ws.cell(row=row_idx, column=tier_col_idx)
        tc.fill = TIER_FILLS[tier]
        tc.font = TIER_FONTS[tier]

    # Grade column coloring (if grade present)
    if grade_col_idx and quality_grade and quality_grade in GRADE_FILLS:
        gc = ws.cell(row=row_idx, column=grade_col_idx)
        gc.fill = GRADE_FILLS[quality_grade]


# ── Sheet writers ──────────────────────────────────────────────────────


def write_tier1_sheet(wb: Workbook, scored_data: list[tuple], events: list[dict]) -> list[tuple]:
    """Sheet 1: Tier 1 — Strategic (Top 50). Returns the tier1 subset."""
    ws = wb.create_sheet("Tier 1 — Strategic (Top 50)")
    style_header_row(ws, FULL_COLUMNS)

    # Filter strict Tier 1 (ICP >= 75 + contacts + quality >= 70), cap at 50
    tier1 = [
        (rec, icp, tier)
        for rec, icp, tier in scored_data
        if tier == "Tier 1"
    ][:50]

    # ICP tier column = col 9, quality score = col 19
    TIER_COL = 9
    QS_COL = 19

    for rank, (rec, icp, tier) in enumerate(tier1, 1):
        contact_name, contact_email = get_primary_contact(rec)
        contacts = get_contacts(rec)
        contact_title = ""
        contact_phone = ""
        if contacts and isinstance(contacts[0], dict):
            contact_title = contacts[0].get("title", "")
            contact_phone = contacts[0].get("phone", "")

        ts_str = "; ".join(get_tech_stack(rec)[:8])
        spf_str = "; ".join(get_spf_list(rec)[:5])

        row_vals = [
            rank,
            rec.get("company_name", ""),
            (rec.get("website") or "").strip(),
            (rec.get("domain") or "").strip(),
            (rec.get("city") or "").strip(),
            (rec.get("state") or "").strip(),
            "; ".join(get_associations_list(rec)),
            icp["icp_score"],
            tier,
            get_email_provider(rec),
            _emp_display(rec),
            (rec.get("erp_system") or "").strip(),
            ts_str,
            spf_str,
            contact_name,
            contact_title,
            contact_email,
            contact_phone,
            rec.get("quality_score", 0),
            _recommend_campaign(rec, icp),
            _event_overlap(rec, events),
            detect_competitor(rec),
        ]

        _write_data_row(
            ws, rank + 1, row_vals,
            tier=tier,
            quality_score=rec.get("quality_score", 0),
            quality_grade=rec.get("quality_grade", ""),
            tier_col_idx=TIER_COL,
        )

    print(f"  Sheet 'Tier 1 — Strategic': {len(tier1)} rows")
    return tier1


def write_tier2_sheet(wb: Workbook, scored_data: list[tuple], events: list[dict]) -> list[tuple]:
    """Sheet 2: Tier 2 — Growth (200). Returns the tier2 subset."""
    ws = wb.create_sheet("Tier 2 — Growth (200)")
    style_header_row(ws, TIER2_EXTRA_COLUMNS)

    # Tier 2 = ICP >= 55 (or ICP >= 75 without contacts). Sort by ICP desc.
    tier2 = [
        (rec, icp, tier)
        for rec, icp, tier in scored_data
        if tier == "Tier 2"
    ][:200]

    TIER_COL = 9

    for rank, (rec, icp, tier) in enumerate(tier2, 1):
        contact_name, contact_email = get_primary_contact(rec)
        contacts = get_contacts(rec)
        contact_title = ""
        contact_phone = ""
        if contacts and isinstance(contacts[0], dict):
            contact_title = contacts[0].get("title", "")
            contact_phone = contacts[0].get("phone", "")

        ts_str = "; ".join(get_tech_stack(rec)[:8])
        spf_str = "; ".join(get_spf_list(rec)[:5])

        row_vals = [
            rank,
            rec.get("company_name", ""),
            (rec.get("website") or "").strip(),
            (rec.get("domain") or "").strip(),
            (rec.get("city") or "").strip(),
            (rec.get("state") or "").strip(),
            "; ".join(get_associations_list(rec)),
            icp["icp_score"],
            tier,
            get_email_provider(rec),
            _emp_display(rec),
            (rec.get("erp_system") or "").strip(),
            ts_str,
            spf_str,
            contact_name,
            contact_title,
            contact_email,
            contact_phone,
            rec.get("quality_score", 0),
            _recommend_campaign(rec, icp),
            _event_overlap(rec, events),
            detect_competitor(rec),
            _enrichment_gaps(rec),
        ]

        _write_data_row(
            ws, rank + 1, row_vals,
            tier=tier,
            quality_score=rec.get("quality_score", 0),
            quality_grade=rec.get("quality_grade", ""),
            tier_col_idx=TIER_COL,
        )

    print(f"  Sheet 'Tier 2 — Growth': {len(tier2)} rows")
    return tier2


def write_tier3_sheet(wb: Workbook, scored_data: list[tuple]) -> list[tuple]:
    """Sheet 3: Tier 3 — Nurture (trimmed columns)."""
    ws = wb.create_sheet("Tier 3 — Nurture")
    style_header_row(ws, NURTURE_COLUMNS)

    tier3 = [
        (rec, icp, tier)
        for rec, icp, tier in scored_data
        if tier == "Tier 3"
    ]

    TIER_COL = 8  # "Tier" column

    for rank, (rec, icp, tier) in enumerate(tier3, 1):
        row_vals = [
            rank,
            rec.get("company_name", ""),
            (rec.get("website") or "").strip(),
            (rec.get("city") or "").strip(),
            (rec.get("state") or "").strip(),
            "; ".join(get_associations_list(rec)),
            icp["icp_score"],
            tier,
            get_email_provider(rec),
            rec.get("quality_score", 0),
            _enrichment_gaps(rec),
        ]

        _write_data_row(
            ws, rank + 1, row_vals,
            tier=tier,
            quality_score=rec.get("quality_score", 0),
            quality_grade=rec.get("quality_grade", ""),
            tier_col_idx=TIER_COL,
        )

    print(f"  Sheet 'Tier 3 — Nurture': {len(tier3)} rows")
    return tier3


def write_assignment_sheet(wb: Workbook, tier1: list[tuple], tier2: list[tuple]):
    """Sheet 4: Account Assignment Matrix (sales manager working document)."""
    ws = wb.create_sheet("Account Assignment Matrix")
    style_header_row(ws, ASSIGNMENT_COLUMNS)

    # Empty-column fill indices (1-based): Assigned Rep=6, Campaign Stage=7, Next Action Date=8, Notes=9
    EMPTY_COLS = {6, 7, 8, 9}

    all_accounts = list(tier1) + list(tier2)

    for row_idx, (rec, icp, tier) in enumerate(all_accounts, 2):
        assocs = "; ".join(get_associations_list(rec))

        row_vals = [
            rec.get("company_name", ""),
            tier,
            icp["icp_score"],
            (rec.get("state") or "").strip(),
            assocs,
            "",   # Assigned Rep — to be filled
            "",   # Campaign Stage — to be filled
            "",   # Next Action Date — to be filled
            "",   # Notes — to be filled
        ]

        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if col_idx in EMPTY_COLS:
                cell.fill = FILL_PENDING

        # Color-code tier column (col 2)
        tc = ws.cell(row=row_idx, column=2)
        if tier in TIER_FILLS:
            tc.fill = TIER_FILLS[tier]
            tc.font = TIER_FONTS[tier]

    print(f"  Sheet 'Account Assignment Matrix': {len(all_accounts)} rows (Tier 1 + 2)")


def write_summary_sheet(
    wb: Workbook,
    scored_data: list[tuple],
    tier1: list[tuple],
    tier2: list[tuple],
    tier3: list[tuple],
):
    """Sheet 5: Tier Summary Statistics."""
    ws = wb.create_sheet("Tier Summary Statistics")

    # Fix column widths manually (no auto-filter for summary sheets)
    col_widths = [20, 10, 8, 10, 12, 16, 14, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[_col_letter(i)].width = w

    row = 1
    total = len(scored_data)

    # ── Table 1: Tier Distribution ──────────────────────────────────────
    write_section_header(ws, row, "Table 1: Tier Distribution", 8)
    row += 1

    t1_headers = [
        ("Tier", 20), ("Count", 10), ("%", 8),
        ("Avg ICP", 10), ("Avg Quality", 12),
        ("With Contacts %", 16), ("With Website %", 14), ("With Tech Stack %", 16),
    ]
    for col_idx, (h, w) in enumerate(t1_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        ws.column_dimensions[_col_letter(col_idx)].width = w
    row += 1

    tier_groups: dict[str, list[tuple]] = defaultdict(list)
    for rec, icp, tier in scored_data:
        tier_groups[tier].append((rec, icp))

    for tier_name in ["Tier 1", "Tier 2", "Tier 3", "Unqualified"]:
        items = tier_groups.get(tier_name, [])
        count = len(items)
        pct = f"{count / total * 100:.1f}%" if total else "0%"
        avg_icp = sum(i["icp_score"] for _, i in items) / max(count, 1)
        avg_q = sum(r.get("quality_score", 0) for r, _ in items) / max(count, 1)
        with_contacts = sum(1 for r, _ in items if get_contacts(r)) / max(count, 1) * 100
        with_web = sum(1 for r, _ in items if has_website(r)) / max(count, 1) * 100
        with_ts = sum(1 for r, _ in items if get_tech_stack(r)) / max(count, 1) * 100

        row_vals = [
            tier_name, count, pct,
            f"{avg_icp:.1f}", f"{avg_q:.1f}",
            f"{with_contacts:.1f}%", f"{with_web:.1f}%", f"{with_ts:.1f}%",
        ]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
        # Color-code tier name cell
        tc = ws.cell(row=row, column=1)
        if tier_name in TIER_FILLS:
            tc.fill = TIER_FILLS[tier_name]
            tc.font = TIER_FONTS[tier_name]
        row += 1

    # ── Table 2: Tier 1 by State (top 10) ──────────────────────────────
    row += 2
    write_section_header(ws, row, "Table 2: Tier 1 Accounts by State (Top 10)", 3)
    row += 1

    t2_headers = [("State", 10), ("Count", 10), ("Avg ICP", 12)]
    for col_idx, (h, w) in enumerate(t2_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        ws.column_dimensions[_col_letter(col_idx)].width = w
    row += 1

    state_icp: dict[str, list[int]] = defaultdict(list)
    for rec, icp, tier in tier1:
        st = (rec.get("state") or "").strip().upper()
        if st:
            state_icp[st].append(icp["icp_score"])

    top_states = sorted(state_icp.keys(), key=lambda s: -len(state_icp[s]))[:10]
    for st in top_states:
        scores = state_icp[st]
        avg = sum(scores) / len(scores)
        for col_idx, val in enumerate([st, len(scores), f"{avg:.1f}"], 1):
            ws.cell(row=row, column=col_idx, value=val).border = THIN_BORDER
        row += 1

    # ── Table 3: Tier 1 by Association ─────────────────────────────────
    row += 2
    write_section_header(ws, row, "Table 3: Tier 1 Accounts by Association", 4)
    row += 1

    t3_headers = [("Association", 15), ("Count", 10), ("Avg ICP", 12), ("Avg Quality", 14)]
    for col_idx, (h, w) in enumerate(t3_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        ws.column_dimensions[_col_letter(col_idx)].width = w
    row += 1

    assoc_data: dict[str, list[tuple]] = defaultdict(list)
    for rec, icp, tier in tier1:
        for a in get_associations_list(rec):
            assoc_data[a.upper()].append((icp["icp_score"], rec.get("quality_score", 0)))

    for assoc in sorted(assoc_data.keys(), key=lambda a: -len(assoc_data[a])):
        items = assoc_data[assoc]
        count = len(items)
        avg_icp = sum(i for i, _ in items) / count
        avg_q = sum(q for _, q in items) / count
        for col_idx, val in enumerate([assoc, count, f"{avg_icp:.1f}", f"{avg_q:.1f}"], 1):
            ws.cell(row=row, column=col_idx, value=val).border = THIN_BORDER
        row += 1

    # ── Table 4: Enrichment Gap Analysis (Tier 2) ─────────────────────
    row += 2
    write_section_header(ws, row, "Table 4: Enrichment Gap Analysis — Tier 2 Accounts", 4)
    row += 1

    t4_headers = [("Field", 20), ("Missing Count", 14), ("Missing %", 12), ("Priority", 10)]
    for col_idx, (h, w) in enumerate(t4_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        ws.column_dimensions[_col_letter(col_idx)].width = w
    row += 1

    tier2_recs = [rec for rec, _, _ in tier2]
    t2_count = max(len(tier2_recs), 1)

    gap_fields = [
        ("contacts",       lambda r: not get_contacts(r),          "High"),
        ("employee_count", lambda r: not any(get_employee_count(r)), "High"),
        ("tech_stack",     lambda r: not get_tech_stack(r),         "High"),
        ("email_provider", lambda r: not get_email_provider(r),     "Medium"),
        ("phone",          lambda r: not (r.get("phone") or ""),    "Medium"),
        ("website",        lambda r: not has_website(r),            "Low"),
    ]

    # Priority fill colors
    priority_fills = {
        "High":   PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "Medium": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "Low":    PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    }

    for field_label, missing_fn, priority in gap_fields:
        missing_count = sum(1 for r in tier2_recs if missing_fn(r))
        missing_pct = f"{missing_count / t2_count * 100:.1f}%"
        row_vals = [field_label, missing_count, missing_pct, priority]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
        # Color priority cell
        pc = ws.cell(row=row, column=4)
        if priority in priority_fills:
            pc.fill = priority_fills[priority]
        row += 1

    print(f"  Sheet 'Tier Summary Statistics': 4 tables written")


# ── Main ───────────────────────────────────────────────────────────────


def main():
    print("=" * 60)
    print("NAM Intelligence Pipeline - ABM Target Lists")
    print("=" * 60)

    # Load data
    print("\nLoading and merging data...")
    records = load_and_merge_data()
    print(f"  Loaded {len(records)} companies")

    events = load_events()
    print(f"  Loaded {len(events)} events")

    load_competitors()  # available but used via detect_competitor() per-record

    # Score all records
    print("\nComputing ICP scores...")
    scored_data: list[tuple] = []
    for rec in records:
        icp = compute_icp_score(rec)
        contacts = get_contacts(rec)
        qs = rec.get("quality_score", 0)
        try:
            qs = int(qs)
        except (ValueError, TypeError):
            qs = 0
        tier = assign_tier(icp["icp_score"], bool(contacts), qs)
        scored_data.append((rec, icp, tier))

    # Sort by ICP score descending (primary), quality_score descending (secondary)
    scored_data.sort(key=lambda x: (-x[1]["icp_score"], -x[0].get("quality_score", 0)))

    # Stats
    scores = [icp["icp_score"] for _, icp, _ in scored_data]
    tiers = Counter(t for _, _, t in scored_data)
    print(f"  Average ICP: {sum(scores) / len(scores):.1f}")
    print(
        f"  Tier 1: {tiers.get('Tier 1', 0)}, "
        f"Tier 2: {tiers.get('Tier 2', 0)}, "
        f"Tier 3: {tiers.get('Tier 3', 0)}, "
        f"Unqualified: {tiers.get('Unqualified', 0)}"
    )

    # Generate workbook
    print("\nGenerating workbook...")
    wb = Workbook()
    # Remove auto-created blank sheet
    if wb.active and wb.active.title == "Sheet":
        wb.remove(wb.active)

    tier1 = write_tier1_sheet(wb, scored_data, events)
    tier2 = write_tier2_sheet(wb, scored_data, events)
    tier3 = write_tier3_sheet(wb, scored_data)
    write_assignment_sheet(wb, tier1, tier2)
    write_summary_sheet(wb, scored_data, tier1, tier2, tier3)

    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"\nSaved: {OUTPUT_PATH}")
    print("Done!")


if __name__ == "__main__":
    main()
