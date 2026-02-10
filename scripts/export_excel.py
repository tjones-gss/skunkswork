"""
Excel Export Script
NAM Intelligence Pipeline

Generates:
  1. data/exports/GSS_NAM_Intelligence_Master.xlsx  (7-sheet workbook)
  2. data/exports/GSS_Salesforce_Import.csv          (SFDC-ready CSV)
"""

import csv
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

BASE_DIR = Path(__file__).resolve().parent.parent
ENRICHED_PATH = BASE_DIR / "data" / "processed" / "enriched_all.jsonl"
CSV_PATH = BASE_DIR / "data" / "exports" / "companies_all.csv"
EXCEL_OUT = BASE_DIR / "data" / "exports" / "GSS_NAM_Intelligence_Master.xlsx"
SFDC_OUT = BASE_DIR / "data" / "exports" / "GSS_Salesforce_Import.csv"


def _join_list(val, sep="; "):
    """Join a list to string, or return empty string."""
    if isinstance(val, list):
        return sep.join(str(v) for v in val if v)
    if isinstance(val, str):
        return val
    return ""


def _normalize(name: str) -> str:
    """Normalize company name for merge key."""
    return re.sub(r"[^a-z0-9]", "", name.lower()) if name else ""


def _parse_csv_contacts(raw: str) -> list[dict]:
    """Parse CSV contact string like 'Mr. Tim Daily <tedaily@example.com>; Ms. Stacy Kelly <sk@example.com>'."""
    contacts = []
    if not raw:
        return contacts
    for part in raw.split(";"):
        part = part.strip()
        if not part:
            continue
        m = re.match(r"^(.+?)\s*<([^>]+)>$", part)
        if m:
            contacts.append({"name": m.group(1).strip(), "email": m.group(2).strip()})
        else:
            contacts.append({"name": part})
    return contacts


class ExcelExporter:
    """Export merged pipeline data to Excel workbook and SFDC CSV."""

    # Extended column definitions for All Companies sheet
    ALL_COLUMNS = [
        ("Company Name", "company_name", 35),
        ("Website", "website", 30),
        ("Domain", "domain", 20),
        ("City", "city", 18),
        ("State", "state", 8),
        ("Country", "country", 15),
        ("Association(s)", "associations", 15),
        ("Membership Tier", "membership_tier", 15),
        ("Member Since", "member_since", 12),
        ("Phone", "phone", 16),
        ("Fax", "fax", 16),
        ("Employee Count", "employee_count_display", 15),
        ("Revenue Range", "revenue_display", 18),
        ("Year Founded", "year_founded", 12),
        ("Industry", "industry", 25),
        ("NAICS Code", "naics_code", 12),
        ("ERP System", "erp_system", 18),
        ("CRM System", "crm_system", 15),
        ("Tech Stack", "tech_stack_display", 40),
        ("Email Provider", "email_provider", 18),
        ("CMS", "cms", 15),
        ("SPF Services", "spf_services_display", 25),
        ("Manufacturing Processes", "manufacturing_processes_display", 35),
        ("Facility Size", "facility_size_sqft", 15),
        ("Certifications", "certifications_display", 30),
        ("Markets Served", "markets_served_display", 30),
        ("Publicly Traded", "publicly_traded_display", 14),
        ("Contact Name", "primary_contact_name", 25),
        ("Contact Title", "primary_contact_title", 25),
        ("Contact Email", "primary_contact_email", 30),
        ("Contact Phone", "primary_contact_phone", 18),
        ("Quality Score", "quality_score", 12),
        ("Quality Grade", "quality_grade", 12),
        ("Source URL", "source_url", 50),
    ]

    # Colors
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    GRADE_COLORS = {
        "A": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "B": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "C": PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
        "D": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "F": PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"),
    }

    def __init__(self, records: list[dict]):
        self.records = records
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)

    def export(self, excel_path: str, sfdc_path: str):
        """Export all sheets and save workbook + SFDC CSV."""
        print(f"Exporting {len(self.records)} records...")

        self._prepare_records()

        # Sort all records by quality_score descending
        self.records.sort(key=lambda r: r.get("quality_score", 0), reverse=True)

        self._create_all_companies_sheet()
        self._create_high_quality_sheet()
        self._create_salesforce_sheet()
        self._create_by_association_sheet()
        self._create_by_state_sheet()
        self._create_contacts_sheet()
        self._create_quality_report_sheet()

        # Save workbook
        Path(excel_path).parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(excel_path)
        print(f"Excel saved: {excel_path}")

        # Save SFDC CSV
        self._write_sfdc_csv(sfdc_path)
        print(f"SFDC CSV saved: {sfdc_path}")

    # ------------------------------------------------------------------
    # Record preparation
    # ------------------------------------------------------------------

    def _prepare_records(self):
        """Prepare records with display fields."""
        for rec in self.records:
            # Employee count display
            emp_min = rec.get("employee_count_min")
            emp_max = rec.get("employee_count_max")
            if emp_min and emp_max:
                if emp_min == emp_max:
                    rec["employee_count_display"] = str(emp_min)
                else:
                    rec["employee_count_display"] = f"{emp_min}-{emp_max}"
            elif emp_min:
                rec["employee_count_display"] = f"{emp_min}+"

            # Revenue display
            rev_min = rec.get("revenue_min_usd")
            rev_max = rec.get("revenue_max_usd")
            if rev_min:
                rec["revenue_display"] = self._format_currency(rev_min, rev_max)

            # Tech stack display
            tech_stack = rec.get("tech_stack", [])
            if isinstance(tech_stack, list) and tech_stack:
                rec["tech_stack_display"] = "; ".join(tech_stack[:10])
            elif isinstance(tech_stack, str) and tech_stack:
                rec["tech_stack_display"] = tech_stack

            # List-to-display fields
            rec["spf_services_display"] = _join_list(rec.get("spf_services"))
            rec["manufacturing_processes_display"] = _join_list(rec.get("manufacturing_processes"))
            rec["certifications_display"] = _join_list(rec.get("certifications"))
            rec["markets_served_display"] = _join_list(rec.get("markets_served"))

            # Publicly traded display
            pt = rec.get("publicly_traded")
            if pt is True:
                rec["publicly_traded_display"] = "Yes"
            elif pt is False:
                rec["publicly_traded_display"] = "No"
            else:
                rec["publicly_traded_display"] = ""

            # Associations display (normalize to string)
            assoc = rec.get("associations") or rec.get("association") or ""
            if isinstance(assoc, list):
                assoc = "; ".join(a for a in assoc if a)
            rec["associations"] = assoc

            # Primary contact
            contacts = rec.get("contacts", [])
            if isinstance(contacts, list) and contacts:
                primary = contacts[0]
                rec["primary_contact_name"] = primary.get("name", "")
                rec["primary_contact_title"] = primary.get("title", "")
                rec["primary_contact_email"] = primary.get("email", "")
                rec["primary_contact_phone"] = primary.get("phone", "")

            # Domain from website if missing
            website = rec.get("website", "")
            if website and not rec.get("domain"):
                if "://" in website:
                    domain = website.split("://")[1].split("/")[0]
                    if domain.startswith("www."):
                        domain = domain[4:]
                    rec["domain"] = domain

    @staticmethod
    def _format_currency(min_val, max_val=None) -> str:
        def fmt(n):
            n = int(n)
            if n >= 1_000_000_000:
                return f"${n / 1_000_000_000:.1f}B"
            elif n >= 1_000_000:
                return f"${n / 1_000_000:.0f}M"
            elif n >= 1_000:
                return f"${n / 1_000:.0f}K"
            return f"${n:,}"
        if max_val and max_val != min_val:
            return f"{fmt(min_val)} - {fmt(max_val)}"
        return fmt(min_val)

    # ------------------------------------------------------------------
    # Shared helpers
    # ------------------------------------------------------------------

    def _write_header_row(self, ws, columns):
        """Write styled header row and set column widths."""
        for col_idx, (header, _, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = "A2"

    def _write_data_rows(self, ws, records, columns, start_row=2):
        """Write data rows with hyperlinks and grade coloring."""
        for row_idx, rec in enumerate(records, start_row):
            for col_idx, (_, field, _) in enumerate(columns, 1):
                value = rec.get(field, "")
                if isinstance(value, list):
                    value = "; ".join(str(v) for v in value)
                if value is None:
                    value = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if field in ("website", "source_url") and isinstance(value, str) and value.startswith("http"):
                    cell.hyperlink = value
                    cell.font = Font(color="0563C1", underline="single")
                if field == "quality_grade" and value in self.GRADE_COLORS:
                    cell.fill = self.GRADE_COLORS[value]

    # ------------------------------------------------------------------
    # Sheet 1: All Companies
    # ------------------------------------------------------------------

    def _create_all_companies_sheet(self):
        ws = self.wb.create_sheet("All Companies")
        self._write_header_row(ws, self.ALL_COLUMNS)
        self._write_data_rows(ws, self.records, self.ALL_COLUMNS)
        ws.auto_filter.ref = ws.dimensions
        print(f"  Sheet 1 - All Companies: {len(self.records)} rows")

    # ------------------------------------------------------------------
    # Sheet 2: High-Quality Targets
    # ------------------------------------------------------------------

    def _create_high_quality_sheet(self):
        hq = [r for r in self.records if r.get("quality_grade") == "B"]
        hq.sort(key=lambda r: r.get("quality_score", 0), reverse=True)
        ws = self.wb.create_sheet("High-Quality Targets")
        self._write_header_row(ws, self.ALL_COLUMNS)
        self._write_data_rows(ws, hq, self.ALL_COLUMNS)
        if hq:
            ws.auto_filter.ref = ws.dimensions
        print(f"  Sheet 2 - High-Quality Targets: {len(hq)} rows")

    # ------------------------------------------------------------------
    # Sheet 3: Salesforce Import
    # ------------------------------------------------------------------

    def _create_salesforce_sheet(self):
        hq = [r for r in self.records if r.get("quality_grade") == "B"]
        hq.sort(key=lambda r: r.get("quality_score", 0), reverse=True)

        sfdc_cols = [
            ("Account Name", "_sfdc_account_name", 35),
            ("Website", "website", 30),
            ("Phone", "phone", 16),
            ("BillingCity", "city", 18),
            ("BillingState", "state", 8),
            ("BillingCountry", "country", 15),
            ("BillingStreet", "street", 30),
            ("BillingPostalCode", "zip_code", 14),
            ("Industry", "_sfdc_industry", 18),
            ("NumberOfEmployees", "_sfdc_employees", 16),
            ("Description", "_sfdc_description", 50),
            ("Email_Provider__c", "email_provider", 18),
            ("Tech_Stack__c", "tech_stack_display", 40),
            ("CMS__c", "cms", 15),
            ("Quality_Score__c", "quality_score", 14),
            ("Data_Source__c", "_sfdc_data_source", 22),
            ("SPF_Services__c", "spf_services_display", 25),
            ("Primary_Contact__c", "_sfdc_primary_contact", 40),
        ]

        # Pre-compute SFDC fields
        for rec in hq:
            rec["_sfdc_account_name"] = rec.get("company_name", "")
            rec["_sfdc_industry"] = "Manufacturing"
            rec["_sfdc_employees"] = rec.get("employee_count_max") or rec.get("employee_count_min") or ""
            rec["_sfdc_data_source"] = "NAM Intelligence Pipeline"

            # Description
            parts = []
            assoc = rec.get("associations", "")
            if assoc:
                parts.append(f"Association: {assoc}")
            tier = rec.get("membership_tier", "")
            if tier:
                parts.append(f"Membership: {tier}")
            procs = rec.get("manufacturing_processes_display", "")
            if procs:
                parts.append(f"Processes: {procs}")
            status = rec.get("enrichment_status", "")
            if status:
                parts.append(f"Enrichment: {status}")
            rec["_sfdc_description"] = ". ".join(parts) if parts else ""

            # Primary contact
            cn = rec.get("primary_contact_name", "")
            ct = rec.get("primary_contact_title", "")
            ce = rec.get("primary_contact_email", "")
            contact_parts = [p for p in [cn, ct, ce] if p]
            rec["_sfdc_primary_contact"] = " ".join(contact_parts)

        ws = self.wb.create_sheet("Salesforce Import")
        self._write_header_row(ws, sfdc_cols)
        self._write_data_rows(ws, hq, sfdc_cols)
        if hq:
            ws.auto_filter.ref = ws.dimensions
        print(f"  Sheet 3 - Salesforce Import: {len(hq)} rows")

        # Store for CSV output
        self._sfdc_records = hq
        self._sfdc_cols = sfdc_cols

    # ------------------------------------------------------------------
    # Sheet 4: By Association
    # ------------------------------------------------------------------

    def _create_by_association_sheet(self):
        ws = self.wb.create_sheet("By Association")

        assoc_stats = {}
        for rec in self.records:
            raw = rec.get("associations", "")
            assocs = [a.strip() for a in raw.replace(",", ";").split(";") if a.strip()] if raw else ["Unknown"]
            for assoc in assocs:
                if assoc not in assoc_stats:
                    assoc_stats[assoc] = {
                        "count": 0, "with_website": 0, "with_email_provider": 0,
                        "with_tech_stack": 0, "with_contacts": 0, "total_score": 0,
                    }
                s = assoc_stats[assoc]
                s["count"] += 1
                if rec.get("website"):
                    s["with_website"] += 1
                if rec.get("email_provider"):
                    s["with_email_provider"] += 1
                ts = rec.get("tech_stack") or rec.get("tech_stack_display")
                if ts:
                    s["with_tech_stack"] += 1
                if rec.get("contacts") or rec.get("primary_contact_name"):
                    s["with_contacts"] += 1
                s["total_score"] += rec.get("quality_score", 0) or 0

        headers = [
            ("Association", 20), ("Companies", 12), ("With Website", 14),
            ("With Email Provider", 20), ("With Tech Stack", 16),
            ("With Contacts", 14), ("Avg Quality Score", 18),
        ]
        for col_idx, (h, w) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        for row_idx, (assoc, s) in enumerate(
            sorted(assoc_stats.items(), key=lambda x: -x[1]["count"]), 2
        ):
            n = s["count"]
            ws.cell(row=row_idx, column=1, value=assoc)
            ws.cell(row=row_idx, column=2, value=n)
            ws.cell(row=row_idx, column=3, value=f"{s['with_website'] / n * 100:.0f}%" if n else "0%")
            ws.cell(row=row_idx, column=4, value=f"{s['with_email_provider'] / n * 100:.0f}%" if n else "0%")
            ws.cell(row=row_idx, column=5, value=f"{s['with_tech_stack'] / n * 100:.0f}%" if n else "0%")
            ws.cell(row=row_idx, column=6, value=f"{s['with_contacts'] / n * 100:.0f}%" if n else "0%")
            ws.cell(row=row_idx, column=7, value=round(s["total_score"] / n, 1) if n else 0)

        ws.freeze_panes = "A2"
        print(f"  Sheet 4 - By Association: {len(assoc_stats)} associations")

    # ------------------------------------------------------------------
    # Sheet 5: By State
    # ------------------------------------------------------------------

    def _create_by_state_sheet(self):
        ws = self.wb.create_sheet("By State")

        state_stats = {}
        for rec in self.records:
            st = rec.get("state", "").strip()
            if not st:
                st = "Unknown"
            if st not in state_stats:
                state_stats[st] = {"count": 0, "total_score": 0, "companies": []}
            state_stats[st]["count"] += 1
            state_stats[st]["total_score"] += rec.get("quality_score", 0) or 0
            state_stats[st]["companies"].append(rec.get("company_name", ""))

        headers = [("State", 10), ("Companies", 12), ("Avg Quality Score", 18), ("Top Companies", 60)]
        for col_idx, (h, w) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        sorted_states = sorted(state_stats.items(), key=lambda x: -x[1]["count"])
        for row_idx, (st, s) in enumerate(sorted_states, 2):
            n = s["count"]
            ws.cell(row=row_idx, column=1, value=st)
            ws.cell(row=row_idx, column=2, value=n)
            ws.cell(row=row_idx, column=3, value=round(s["total_score"] / n, 1) if n else 0)
            top3 = [c for c in s["companies"][:3] if c]
            ws.cell(row=row_idx, column=4, value="; ".join(top3))

        ws.freeze_panes = "A2"
        print(f"  Sheet 5 - By State: {len(state_stats)} states")

    # ------------------------------------------------------------------
    # Sheet 6: Decision-Maker Contacts
    # ------------------------------------------------------------------

    def _create_contacts_sheet(self):
        ws = self.wb.create_sheet("Decision-Maker Contacts")

        contact_rows = []
        for rec in self.records:
            contacts = rec.get("contacts", [])
            if isinstance(contacts, list):
                for c in contacts:
                    if isinstance(c, dict):
                        contact_rows.append({
                            "company": rec.get("company_name", ""),
                            "name": c.get("name", ""),
                            "title": c.get("title", ""),
                            "email": c.get("email", ""),
                            "phone": c.get("phone", ""),
                            "association": rec.get("associations", ""),
                            "quality_score": rec.get("quality_score", ""),
                        })
            # Also try parsing from CSV-style contact string (already parsed at merge time)
            # Contacts from CSV have already been converted to list[dict] during merge.

        columns = [
            ("Company", "company", 35),
            ("Contact Name", "name", 25),
            ("Title", "title", 30),
            ("Email", "email", 35),
            ("Phone", "phone", 18),
            ("Association", "association", 15),
            ("Quality Score", "quality_score", 14),
        ]

        for col_idx, (h, _, w) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        for row_idx, cr in enumerate(contact_rows, 2):
            for col_idx, (_, field, _) in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=cr.get(field, ""))

        ws.freeze_panes = "A2"
        if contact_rows:
            ws.auto_filter.ref = ws.dimensions
        print(f"  Sheet 6 - Decision-Maker Contacts: {len(contact_rows)} contacts")

    # ------------------------------------------------------------------
    # Sheet 7: Data Quality Report
    # ------------------------------------------------------------------

    def _create_quality_report_sheet(self):
        ws = self.wb.create_sheet("Data Quality Report")
        total = len(self.records)

        # Grade distribution
        grade_dist = {"A": 0, "B": 0, "C": 0, "D": 0, "F": 0}
        total_score = 0
        for rec in self.records:
            g = rec.get("quality_grade", "F")
            grade_dist[g] = grade_dist.get(g, 0) + 1
            total_score += rec.get("quality_score", 0) or 0

        # Summary header
        ws["A1"] = "Data Quality Summary"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A3"] = "Total Records:"
        ws["B3"] = total
        ws["A4"] = "Average Quality Score:"
        ws["B4"] = round(total_score / total, 1) if total else 0

        # Grade distribution table
        ws["A6"] = "Quality Grade Distribution"
        ws["A6"].font = Font(bold=True, size=12)
        for col_idx, h in enumerate(["Grade", "Count", "Percentage"], 1):
            cell = ws.cell(row=7, column=col_idx, value=h)
            cell.font = Font(bold=True)
        for row_idx, (g, cnt) in enumerate(grade_dist.items(), 8):
            ws.cell(row=row_idx, column=1, value=g)
            ws.cell(row=row_idx, column=2, value=cnt)
            pct = cnt / total * 100 if total else 0
            ws.cell(row=row_idx, column=3, value=f"{pct:.1f}%")
            if g in self.GRADE_COLORS:
                ws.cell(row=row_idx, column=1).fill = self.GRADE_COLORS[g]

        # Field coverage table
        coverage_fields = [
            ("website", "Website"),
            ("phone", "Phone"),
            ("email_provider", "Email Provider"),
            ("tech_stack_display", "Tech Stack"),
            ("cms", "CMS"),
            ("spf_services_display", "SPF Services"),
            ("primary_contact_name", "Contacts"),
            ("city", "City"),
            ("state", "State"),
            ("employee_count_display", "Employee Count"),
        ]

        start_row = 15
        ws.cell(row=start_row, column=1, value="Field Coverage").font = Font(bold=True, size=12)
        for col_idx, h in enumerate(["Field", "Has Data", "Coverage %"], 1):
            cell = ws.cell(row=start_row + 1, column=col_idx, value=h)
            cell.font = Font(bold=True)

        for i, (field, label) in enumerate(coverage_fields):
            row = start_row + 2 + i
            has_count = sum(1 for r in self.records if r.get(field))
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=has_count)
            pct = has_count / total * 100 if total else 0
            ws.cell(row=row, column=3, value=f"{pct:.1f}%")

        # Enrichment status breakdown
        status_row = start_row + 2 + len(coverage_fields) + 2
        ws.cell(row=status_row, column=1, value="Enrichment Status").font = Font(bold=True, size=12)
        status_counts = {}
        for rec in self.records:
            es = rec.get("enrichment_status", "none")
            status_counts[es] = status_counts.get(es, 0) + 1
        for col_idx, h in enumerate(["Status", "Count", "Percentage"], 1):
            cell = ws.cell(row=status_row + 1, column=col_idx, value=h)
            cell.font = Font(bold=True)
        for i, (status, cnt) in enumerate(sorted(status_counts.items(), key=lambda x: -x[1])):
            row = status_row + 2 + i
            ws.cell(row=row, column=1, value=status)
            ws.cell(row=row, column=2, value=cnt)
            pct = cnt / total * 100 if total else 0
            ws.cell(row=row, column=3, value=f"{pct:.1f}%")

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        print("  Sheet 7 - Data Quality Report: complete")

    # ------------------------------------------------------------------
    # SFDC CSV
    # ------------------------------------------------------------------

    def _write_sfdc_csv(self, path: str):
        """Write Salesforce Import CSV (deduped by Account Name, highest quality_score wins)."""
        records = self._sfdc_records
        cols = self._sfdc_cols

        # Deduplicate by Account Name (keep highest quality_score)
        seen = {}
        for rec in records:
            name = rec.get("_sfdc_account_name", "")
            key = _normalize(name)
            if not key:
                continue
            existing = seen.get(key)
            if not existing or (rec.get("quality_score", 0) or 0) > (existing.get("quality_score", 0) or 0):
                seen[key] = rec

        deduped = list(seen.values())
        deduped.sort(key=lambda r: r.get("quality_score", 0) or 0, reverse=True)

        Path(path).parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            # Header row
            writer.writerow([h for h, _, _ in cols])
            for rec in deduped:
                row = []
                for _, field, _ in cols:
                    val = rec.get(field, "")
                    if val is None:
                        val = ""
                    if isinstance(val, list):
                        val = "; ".join(str(v) for v in val)
                    row.append(val)
                writer.writerow(row)

        print(f"  SFDC CSV: {len(deduped)} unique accounts (deduped from {len(records)})")


# ======================================================================
# Data loading and merge
# ======================================================================

def load_enriched(path: Path) -> list[dict]:
    """Load enriched JSONL records."""
    records = []
    if not path.exists():
        print(f"Warning: enriched file not found: {path}")
        return records
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    records.append(json.loads(line))
                except json.JSONDecodeError:
                    continue
    return records


def load_csv(path: Path) -> list[dict]:
    """Load companies CSV records."""
    records = []
    if not path.exists():
        print(f"Warning: CSV file not found: {path}")
        return records
    with open(path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Convert quality_score to int
            try:
                row["quality_score"] = int(row.get("quality_score", 0))
            except (ValueError, TypeError):
                row["quality_score"] = 0

            # Parse contacts string to list[dict]
            contacts_raw = row.get("contacts", "")
            if contacts_raw and isinstance(contacts_raw, str):
                row["contacts"] = _parse_csv_contacts(contacts_raw)
            elif not contacts_raw:
                row["contacts"] = []

            # Parse tech_stack string to consistent format
            ts = row.get("tech_stack", "")
            if isinstance(ts, str) and ts:
                row["tech_stack"] = [t.strip() for t in ts.replace(",", ";").split(";") if t.strip()]
            else:
                row["tech_stack"] = []

            # Parse spf_services string
            spf = row.get("spf_services", "")
            if isinstance(spf, str) and spf:
                row["spf_services"] = [s.strip() for s in spf.replace(",", ";").split(";") if s.strip()]
            else:
                row["spf_services"] = []

            records.append(row)
    return records


def merge_records(enriched: list[dict], csv_records: list[dict]) -> list[dict]:
    """Merge enriched JSONL and companies CSV by normalized company_name."""
    # Build enriched lookup by normalized name
    enriched_map = {}
    for rec in enriched:
        key = _normalize(rec.get("company_name", ""))
        if key:
            # If duplicate keys, keep the one that appears later (more recent)
            enriched_map[key] = rec

    merged = []
    seen_keys = set()

    # Process CSV records (these have quality_score/grade)
    for csv_rec in csv_records:
        key = _normalize(csv_rec.get("company_name", ""))
        if not key:
            continue

        enriched_rec = enriched_map.get(key)
        if enriched_rec:
            # Merge: start with enriched data, overlay CSV fields
            combined = dict(enriched_rec)
            # CSV provides quality_score, quality_grade, associations (may differ), contacts parsed
            combined["quality_score"] = csv_rec.get("quality_score", 0)
            combined["quality_grade"] = csv_rec.get("quality_grade", "F")
            # Use CSV associations (already merged/deduped in quality pipeline)
            if csv_rec.get("associations"):
                combined["associations"] = csv_rec["associations"]
            # Use CSV contacts if enriched has none
            if not combined.get("contacts") and csv_rec.get("contacts"):
                combined["contacts"] = csv_rec["contacts"]
            # Fill in missing fields from CSV
            for field in ["city", "state", "country", "phone", "industry", "source_url", "member_type", "notes"]:
                if not combined.get(field) and csv_rec.get(field):
                    combined[field] = csv_rec[field]
        else:
            # CSV-only record
            combined = dict(csv_rec)

        seen_keys.add(key)
        merged.append(combined)

    # Add enriched records not in CSV
    for rec in enriched:
        key = _normalize(rec.get("company_name", ""))
        if key and key not in seen_keys:
            # Give a default quality score
            if "quality_score" not in rec:
                rec["quality_score"] = 50
            if "quality_grade" not in rec:
                rec["quality_grade"] = "D"
            merged.append(rec)
            seen_keys.add(key)

    return merged


# ======================================================================
# Main
# ======================================================================

def main():
    print("=" * 60)
    print("NAM Intelligence Pipeline - Master Export")
    print("=" * 60)

    # Load data
    print(f"\nLoading enriched data from: {ENRICHED_PATH}")
    enriched = load_enriched(ENRICHED_PATH)
    print(f"  Loaded {len(enriched)} enriched records")

    print(f"Loading CSV data from: {CSV_PATH}")
    csv_records = load_csv(CSV_PATH)
    print(f"  Loaded {len(csv_records)} CSV records")

    # Merge
    print("\nMerging records...")
    merged = merge_records(enriched, csv_records)
    print(f"  Merged total: {len(merged)} unique companies")

    if not merged:
        print("Error: No records to export")
        sys.exit(1)

    # Export
    print()
    exporter = ExcelExporter(merged)
    exporter.export(str(EXCEL_OUT), str(SFDC_OUT))

    # Print file sizes
    print("\n" + "=" * 60)
    if EXCEL_OUT.exists():
        size_mb = EXCEL_OUT.stat().st_size / (1024 * 1024)
        print(f"Excel: {EXCEL_OUT} ({size_mb:.2f} MB)")
    if SFDC_OUT.exists():
        size_kb = SFDC_OUT.stat().st_size / 1024
        print(f"SFDC:  {SFDC_OUT} ({size_kb:.1f} KB)")
    print("=" * 60)
    print("Done!")


if __name__ == "__main__":
    main()
