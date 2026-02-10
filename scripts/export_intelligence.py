#!/usr/bin/env python3
"""Export tech intelligence and event strategy workbooks."""

import csv
import json
import os
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
EXPORTS_DIR = DATA_DIR / "exports"
ENRICHED_PATH = DATA_DIR / "processed" / "enriched_all.jsonl"
COMPANIES_PATH = EXPORTS_DIR / "companies_all.csv"
EVENTS_PATH = EXPORTS_DIR / "events_2026.csv"
COMPETITORS_PATH = EXPORTS_DIR / "competitor_analysis.csv"

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

RED_FILL = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")
GREEN_FILL = PatternFill(start_color="44AA44", end_color="44AA44", fill_type="solid")
BLUE_FILL = PatternFill(start_color="4488CC", end_color="4488CC", fill_type="solid")
WHITE_FONT = Font(bold=True, color="FFFFFF")

THREAT_FILLS = {
    "HIGH": RED_FILL,
    "MEDIUM": ORANGE_FILL,
    "LOW": GREEN_FILL,
    "EMERGING": BLUE_FILL,
}
THREAT_RESPONSE = {
    "HIGH": "Must Counter",
    "MEDIUM": "Monitor Closely",
    "LOW": "Track Only",
    "EMERGING": "Watch & Prepare",
}

TECH_CATEGORIES = {
    "Analytics": [
        "Google Analytics", "Google Tag Manager", "Adobe DTM", "Adobe Launch",
        "Hotjar", "Matomo", "Heap", "Mixpanel", "Segment", "Amplitude",
    ],
    "CMS": [
        "WordPress", "Drupal", "HubSpot CMS", "Squarespace", "Wix", "TYPO3 CMS",
        "Joomla", "Kentico", "Sitecore", "Contentful", "Sanity", "Craft CMS",
    ],
    "CDN": [
        "Cloudflare", "Akamai", "Amazon CloudFront", "Fastly", "StackPath",
        "cdnjs", "jsDelivr CDN", "unpkg CDN", "Google CDN",
    ],
    "Framework": [
        "React", "Angular", "Vue.js", "Next.js", "Nuxt.js", "jQuery",
        "Bootstrap", "Tailwind CSS", "ASP.NET", "Express.js", "Laravel",
    ],
    "Email/Marketing": [
        "Microsoft 365", "Salesforce", "Mailchimp", "SendGrid", "HubSpot",
        "Pardot", "Marketo", "Mandrill", "Constant Contact", "Amazon SES",
        "Zendesk", "LiveChat",
    ],
    "Security": [
        "Cloudflare", "reCAPTCHA", "Sucuri", "Imperva", "Akamai",
    ],
    "Hosting": [
        "Nginx", "Apache", "Amazon S3", "Azure", "Google Cloud",
        "Vercel", "Netlify", "WP Engine",
    ],
}


def style_header_row(ws, num_cols):
    """Apply dark-blue header styling, freeze top row, add auto-filter."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{ws.max_row}"


def auto_width(ws, min_width=10, max_width=50):
    """Set column widths based on content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            lengths.append(len(val))
        best = min(max(max(lengths) + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = best


def load_enriched():
    """Load enriched JSONL records."""
    records = []
    with open(ENRICHED_PATH, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                records.append(json.loads(line))
    return records


def load_csv(path):
    """Load CSV file as list of dicts."""
    rows = []
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    return rows


def write_summary_and_detail(ws, title_field, counter, detail_map, total, headers_summary, headers_detail):
    """Write a summary table then a detail table below it, separated by a blank row."""
    # Summary headers
    for ci, h in enumerate(headers_summary, 1):
        ws.cell(row=1, column=ci, value=h)

    sorted_items = counter.most_common()
    row = 2
    for name, count in sorted_items:
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=f"{count / total * 100:.1f}%")
        row += 1

    style_header_row(ws, len(headers_summary))

    # Blank separator row
    detail_start = row + 1

    # Detail headers
    for ci, h in enumerate(headers_detail, 1):
        ws.cell(row=detail_start, column=ci, value=h)
        cell = ws.cell(row=detail_start, column=ci)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    drow = detail_start + 1
    for name, _ in sorted_items:
        for company_name, domain, association in detail_map[name]:
            ws.cell(row=drow, column=1, value=name)
            ws.cell(row=drow, column=2, value=company_name)
            ws.cell(row=drow, column=3, value=domain)
            ws.cell(row=drow, column=4, value=association)
            drow += 1

    auto_width(ws)
    return len(sorted_items)


# ---------------------------------------------------------------------------
# Tech Intelligence Workbook
# ---------------------------------------------------------------------------

def build_tech_intelligence(records):
    """Build GSS_Tech_Intelligence.xlsx from enriched records."""
    wb = Workbook()
    total = len(records)

    # --- Sheet 1: Email Provider Summary ---
    ws1 = wb.active
    ws1.title = "Email Provider Summary"
    provider_counter = Counter()
    provider_detail = defaultdict(list)
    for r in records:
        prov = r.get("email_provider", "") or ""
        if not prov:
            prov = "Unknown"
        provider_counter[prov] += 1
        provider_detail[prov].append((r.get("company_name", ""), r.get("domain", ""), r.get("association", "")))

    n1 = write_summary_and_detail(
        ws1, "email_provider", provider_counter, provider_detail, total,
        ["Provider", "Count", "% of Total"],
        ["Provider", "Company Name", "Domain", "Association"],
    )
    print(f"  Sheet 1: Email Provider Summary - {n1} providers, {total} records")

    # --- Sheet 2: CMS Distribution ---
    ws2 = wb.create_sheet("CMS Distribution")
    cms_counter = Counter()
    cms_detail = defaultdict(list)
    for r in records:
        cms = r.get("cms", "") or ""
        if not cms:
            continue
        cms_counter[cms] += 1
        cms_detail[cms].append((r.get("company_name", ""), r.get("domain", ""), r.get("association", "")))

    cms_total = sum(cms_counter.values())
    n2 = write_summary_and_detail(
        ws2, "cms", cms_counter, cms_detail, cms_total,
        ["CMS", "Count", "% of CMS Users"],
        ["CMS", "Company Name", "Domain", "Association"],
    )
    print(f"  Sheet 2: CMS Distribution - {n2} CMS platforms, {cms_total} companies with CMS")

    # --- Sheet 3: Marketing Tools (SPF) ---
    ws3 = wb.create_sheet("Marketing Tools (SPF)")
    spf_counter = Counter()
    spf_detail = defaultdict(list)
    for r in records:
        services = r.get("spf_services", []) or []
        if isinstance(services, str):
            services = [s.strip() for s in services.split(";") if s.strip()]
        for svc in services:
            spf_counter[svc] += 1
            spf_detail[svc].append((r.get("company_name", ""), r.get("domain", ""), r.get("association", "")))

    n3 = write_summary_and_detail(
        ws3, "spf_services", spf_counter, spf_detail, total,
        ["Service", "Count", "% of Enriched Records"],
        ["Service", "Company Name", "Domain", "Association"],
    )
    print(f"  Sheet 3: Marketing Tools (SPF) - {n3} services detected")

    # --- Sheet 4: Top Technologies ---
    ws4 = wb.create_sheet("Top Technologies")
    tech_counter = Counter()
    for r in records:
        stack = r.get("tech_stack", []) or []
        for t in stack:
            tech_counter[t] += 1

    # Table 1: Top 30
    headers_t1 = ["Technology", "Count", "% Penetration"]
    for ci, h in enumerate(headers_t1, 1):
        ws4.cell(row=1, column=ci, value=h)

    top30 = tech_counter.most_common(30)
    row = 2
    for name, count in top30:
        ws4.cell(row=row, column=1, value=name)
        ws4.cell(row=row, column=2, value=count)
        ws4.cell(row=row, column=3, value=f"{count / total * 100:.1f}%")
        row += 1

    style_header_row(ws4, len(headers_t1))

    # Table 2: Category groupings
    cat_start = row + 2
    cat_headers = ["Category", "Technology", "Count"]
    for ci, h in enumerate(cat_headers, 1):
        cell = ws4.cell(row=cat_start, column=ci, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    crow = cat_start + 1
    for category, techs in TECH_CATEGORIES.items():
        for tech in techs:
            count = tech_counter.get(tech, 0)
            if count > 0:
                ws4.cell(row=crow, column=1, value=category)
                ws4.cell(row=crow, column=2, value=tech)
                ws4.cell(row=crow, column=3, value=count)
                crow += 1

    auto_width(ws4)
    print(f"  Sheet 4: Top Technologies - {len(tech_counter)} unique technologies, top 30 shown")

    out_path = EXPORTS_DIR / "GSS_Tech_Intelligence.xlsx"
    wb.save(str(out_path))
    return out_path


# ---------------------------------------------------------------------------
# Event Strategy Workbook
# ---------------------------------------------------------------------------

def count_companies_for_industry(industry, companies_by_assoc):
    """Rough match: count companies whose association overlaps with the event industry."""
    industry_lower = industry.lower() if industry else ""
    # Map industry keywords to association codes
    keyword_map = {
        "metalforming": ["PMA"],
        "metal fabrication": ["PMA"],
        "metal construction": ["PMA"],
        "electrical": ["NEMA"],
        "motor": ["NEMA"],
        "gear": ["AGMA"],
        "specialty chemical": ["SOCMA"],
        "chemical": ["SOCMA"],
        "aerospace": ["AIA"],
        "defense": ["AIA"],
        "manufacturing technology": ["AMT", "PMA", "NEMA", "AGMA"],
        "automation": ["NEMA", "AMT"],
        "robotics": ["NEMA", "AMT"],
        "packaging": ["PMMI"],
        "plastics": ["PLASTICS"],
        "precision machining": ["PMPA"],
        "spring": ["PMA"],
        "valve": ["PMA"],
        "marine": [],
        "medical device": [],
        "electronics": ["NEMA"],
        "additive": [],
        "advanced material": [],
        "ceramics": [],
    }
    matched_assocs = set()
    for keyword, assocs in keyword_map.items():
        if keyword in industry_lower:
            matched_assocs.update(assocs)

    if not matched_assocs:
        return 0

    total = 0
    for assoc in matched_assocs:
        total += companies_by_assoc.get(assoc, 0)
    return total


def build_event_strategy(events, competitors, companies_by_assoc):
    """Build GSS_Event_Strategy.xlsx."""
    wb = Workbook()

    # --- Sheet 1: Events Calendar ---
    ws1 = wb.active
    ws1.title = "Events Calendar"
    headers = [
        "Event", "Dates", "City", "Venue", "Attendance", "Industry",
        "Registration URL", "Priority", "Notes", "Member Companies",
    ]
    for ci, h in enumerate(headers, 1):
        ws1.cell(row=1, column=ci, value=h)

    for ri, ev in enumerate(events, 2):
        ws1.cell(row=ri, column=1, value=ev.get("event_name", ""))
        ws1.cell(row=ri, column=2, value=ev.get("dates", ""))
        ws1.cell(row=ri, column=3, value=ev.get("city", ""))
        ws1.cell(row=ri, column=4, value=ev.get("venue", ""))
        ws1.cell(row=ri, column=5, value=ev.get("attendance", ""))
        ws1.cell(row=ri, column=6, value=ev.get("industry", ""))
        url = ev.get("registration_url", "")
        if url and not url.startswith("http"):
            url = "https://" + url
        cell = ws1.cell(row=ri, column=7, value=url)
        if url:
            cell.hyperlink = url
            cell.font = Font(color="0563C1", underline="single")
        ws1.cell(row=ri, column=8, value=ev.get("priority", ""))
        ws1.cell(row=ri, column=9, value=ev.get("notes", ""))
        member_count = count_companies_for_industry(ev.get("industry", ""), companies_by_assoc)
        ws1.cell(row=ri, column=10, value=member_count)

    style_header_row(ws1, len(headers))
    auto_width(ws1)
    print(f"  Sheet 1: Events Calendar - {len(events)} events")

    # --- Sheet 2: Competitor Landscape ---
    ws2 = wb.create_sheet("Competitor Landscape")
    c_headers = ["Competitor", "Threat Level", "Show Presence", "Strategy Notes", "GSS Response"]
    for ci, h in enumerate(c_headers, 1):
        ws2.cell(row=1, column=ci, value=h)

    for ri, comp in enumerate(competitors, 2):
        ws2.cell(row=ri, column=1, value=comp.get("competitor", ""))
        threat = comp.get("threat_level", "")
        cell_threat = ws2.cell(row=ri, column=2, value=threat)
        if threat in THREAT_FILLS:
            cell_threat.fill = THREAT_FILLS[threat]
            cell_threat.font = WHITE_FONT
        ws2.cell(row=ri, column=3, value=comp.get("presence", ""))
        ws2.cell(row=ri, column=4, value=comp.get("strategy_notes", ""))
        ws2.cell(row=ri, column=5, value=THREAT_RESPONSE.get(threat, ""))

    style_header_row(ws2, len(c_headers))
    auto_width(ws2)
    print(f"  Sheet 2: Competitor Landscape - {len(competitors)} competitors")

    # --- Sheet 3: Recommended Actions ---
    ws3 = wb.create_sheet("Recommended Actions")
    a_headers = [
        "Event", "Priority", "Competitor Threat", "Recommendation",
        "Budget Tier", "Expected ROI", "Key Actions",
    ]
    for ci, h in enumerate(a_headers, 1):
        ws3.cell(row=1, column=ci, value=h)

    # Build competitor threat lookup from presence text
    # Check if any HIGH-threat competitor mentions the event name or show
    high_threat_names = [c["competitor"] for c in competitors if c.get("threat_level") == "HIGH"]
    high_threat_presence = " ".join(
        c.get("presence", "") for c in competitors if c.get("threat_level") == "HIGH"
    ).upper()
    medium_threat_presence = " ".join(
        c.get("presence", "") for c in competitors if c.get("threat_level") == "MEDIUM"
    ).upper()

    show_keywords = {
        "IMTS": "IMTS",
        "FABTECH": "FABTECH",
        "PACK EXPO": "PACK EXPO",
        "NPE": "NPE",
        "SOCMA": "SOCMA",
        "PMA": "PMA",
        "AUTOMATE": "AUTOMATE",
    }

    ri = 2
    for ev in events:
        event_name = ev.get("event_name", "")
        priority = ev.get("priority", "")
        notes = ev.get("notes", "")

        # Determine competitor threat level for this event
        event_upper = event_name.upper()
        has_high_threat = False
        has_medium_threat = False

        for keyword in show_keywords:
            if keyword in event_upper:
                if keyword in high_threat_presence or show_keywords[keyword] in high_threat_presence:
                    has_high_threat = True
                if keyword in medium_threat_presence or show_keywords[keyword] in medium_threat_presence:
                    has_medium_threat = True

        no_erp_competition = "NO ERP Competition" in notes

        # Decision logic
        if has_high_threat:
            threat_label = "HIGH"
        elif has_medium_threat:
            threat_label = "MEDIUM"
        else:
            threat_label = "LOW"

        if priority == "HIGH" and has_high_threat:
            recommendation = "SPONSOR"
            budget = "$$$"
            roi = "High"
            actions = f"Book premium booth; counter {', '.join(high_threat_names)}; host demo sessions"
        elif priority == "HIGH" and no_erp_competition:
            recommendation = "EXHIBIT"
            budget = "$$"
            roi = "High"
            actions = "Exhibit with standard booth; no competitor pressure allows organic engagement"
        elif priority == "HIGH":
            recommendation = "EXHIBIT"
            budget = "$$"
            roi = "High"
            actions = "Standard booth; product demos; lead capture"
        elif priority == "MEDIUM":
            recommendation = "ATTEND"
            budget = "$"
            roi = "Medium"
            actions = "Send sales team; attend sessions; network with members"
        else:
            recommendation = "SKIP"
            budget = "Free"
            roi = "Low"
            actions = "Monitor remotely; follow up on published attendee lists"

        ws3.cell(row=ri, column=1, value=event_name)
        ws3.cell(row=ri, column=2, value=priority)
        ws3.cell(row=ri, column=3, value=threat_label)
        ws3.cell(row=ri, column=4, value=recommendation)
        ws3.cell(row=ri, column=5, value=budget)
        ws3.cell(row=ri, column=6, value=roi)
        ws3.cell(row=ri, column=7, value=actions)
        ri += 1

    style_header_row(ws3, len(a_headers))
    auto_width(ws3)
    print(f"  Sheet 3: Recommended Actions - {len(events)} events analyzed")

    out_path = EXPORTS_DIR / "GSS_Event_Strategy.xlsx"
    wb.save(str(out_path))
    return out_path


def main():
    os.makedirs(EXPORTS_DIR, exist_ok=True)

    # Load enriched data
    print("Loading enriched data...")
    records = load_enriched()
    print(f"  Loaded {len(records)} enriched records from {ENRICHED_PATH.name}")

    # Load companies for association counts
    companies = load_csv(COMPANIES_PATH)
    companies_by_assoc = Counter()
    for c in companies:
        assocs = c.get("associations", "")
        for a in assocs.split(";"):
            a = a.strip()
            if a:
                companies_by_assoc[a] += 1
    print(f"  Loaded {len(companies)} companies ({len(companies_by_assoc)} associations)")

    # Load events and competitors
    events = load_csv(EVENTS_PATH)
    competitors = load_csv(COMPETITORS_PATH)
    print(f"  Loaded {len(events)} events, {len(competitors)} competitors")

    # Build Tech Intelligence workbook
    print("\nBuilding GSS_Tech_Intelligence.xlsx...")
    tech_path = build_tech_intelligence(records)
    size_kb = tech_path.stat().st_size / 1024
    print(f"  Saved: {tech_path} ({size_kb:.0f} KB)")

    # Build Event Strategy workbook
    print("\nBuilding GSS_Event_Strategy.xlsx...")
    event_path = build_event_strategy(events, competitors, companies_by_assoc)
    size_kb = event_path.stat().st_size / 1024
    print(f"  Saved: {event_path} ({size_kb:.0f} KB)")

    print("\nDone. Both workbooks exported.")


if __name__ == "__main__":
    main()
