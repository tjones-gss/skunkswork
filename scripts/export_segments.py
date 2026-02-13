#!/usr/bin/env python3
"""Export segmented target lists for marketing campaigns."""

import csv
import json
import os
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent.parent
ENRICHED_PATH = BASE_DIR / "data" / "processed" / "enriched_all.jsonl"
CSV_PATH = BASE_DIR / "data" / "exports" / "companies_all.csv"
OUTPUT_PATH = BASE_DIR / "data" / "exports" / "GSS_Segmented_Lists.xlsx"

COLUMNS = [
    ("Company Name", 35),
    ("Website", 30),
    ("Domain", 20),
    ("City", 18),
    ("State", 8),
    ("Street Address", 35),
    ("Zip Code", 12),
    ("Association", 12),
    ("Email Provider", 18),
    ("Phone", 18),
    ("Primary Contact", 25),
    ("Primary Email", 30),
    ("Tech Stack", 40),
    ("Quality Score", 12),
    ("Quality Grade", 12),
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBTITLE_FONT = Font(italic=True, size=10, color="555555")

GRADE_FILLS = {
    "B": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "C": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "D": PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid"),
    "F": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}

THIN_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)


def normalize_name(name: str) -> str:
    """Lowercase, strip whitespace for merge key."""
    return name.strip().lower()


def parse_spf_services(val) -> list[str]:
    """Parse spf_services from list or semicolon-delimited string."""
    if isinstance(val, list):
        return [s.strip() for s in val if s.strip()]
    if isinstance(val, str) and val.strip():
        return [s.strip() for s in val.split(";") if s.strip()]
    return []


def parse_contacts_csv(val: str) -> list[dict]:
    """Parse CSV contacts format: 'Name <email>; Name2 <email2>'."""
    contacts = []
    if not val or not val.strip():
        return contacts
    for entry in val.split(";"):
        entry = entry.strip()
        m = re.match(r"^(.+?)\s*<([^>]+)>$", entry)
        if m:
            contacts.append({"name": m.group(1).strip(), "email": m.group(2).strip()})
        elif entry:
            contacts.append({"name": entry, "email": ""})
    return contacts


def parse_tech_stack(val) -> list[str]:
    """Parse tech_stack from list or semicolon-delimited string."""
    if isinstance(val, list):
        return val
    if isinstance(val, str) and val.strip():
        return [s.strip() for s in val.split(";") if s.strip()]
    return []


def load_data() -> list[dict]:
    """Load and merge enriched JSONL + CSV data."""
    records = {}

    # Primary: enriched JSONL (richer data)
    if ENRICHED_PATH.exists():
        with open(ENRICHED_PATH, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                rec = json.loads(line)
                key = normalize_name(rec.get("company_name", ""))
                if key:
                    records[key] = rec

    print(f"Loaded {len(records)} records from enriched_all.jsonl")

    # Secondary: CSV for quality scores and additional records
    csv_count = 0
    merge_count = 0
    if CSV_PATH.exists():
        with open(CSV_PATH, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                key = normalize_name(row.get("company_name", ""))
                if not key:
                    continue
                csv_count += 1

                qs = row.get("quality_score", "")
                qg = row.get("quality_grade", "")
                qs_int = int(qs) if qs and qs.isdigit() else 0

                if key in records:
                    # Merge quality score/grade from CSV
                    records[key]["quality_score"] = qs_int
                    records[key]["quality_grade"] = qg
                    # Merge associations if CSV has more
                    csv_assoc = row.get("associations", "")
                    if csv_assoc and ";" in csv_assoc:
                        records[key]["associations_csv"] = csv_assoc
                    # Merge employee counts if missing in JSONL
                    if not records[key].get("employee_count_min"):
                        ec_min = row.get("employee_count_min", "")
                        if ec_min and ec_min.isdigit():
                            records[key]["employee_count_min"] = int(ec_min)
                    if not records[key].get("employee_count_max"):
                        ec_max = row.get("employee_count_max", "")
                        if ec_max and ec_max.isdigit():
                            records[key]["employee_count_max"] = int(ec_max)
                    # Merge contacts if JSONL has none
                    if not records[key].get("contacts"):
                        csv_contacts = parse_contacts_csv(row.get("contacts", ""))
                        if csv_contacts:
                            records[key]["contacts"] = csv_contacts
                    merge_count += 1
                else:
                    # New record from CSV only
                    rec = {
                        "company_name": row.get("company_name", ""),
                        "website": row.get("website", ""),
                        "domain": row.get("domain", ""),
                        "city": row.get("city", ""),
                        "state": row.get("state", ""),
                        "country": row.get("country", ""),
                        "association": row.get("associations", "").split(";")[0].strip() if row.get("associations") else "",
                        "associations_csv": row.get("associations", ""),
                        "phone": row.get("phone", ""),
                        "member_type": row.get("member_type", ""),
                        "email_provider": row.get("email_provider", ""),
                        "spf_services": parse_spf_services(row.get("spf_services", "")),
                        "tech_stack": parse_tech_stack(row.get("tech_stack", "")),
                        "cms": row.get("cms", ""),
                        "quality_score": qs_int,
                        "quality_grade": qg,
                        "employee_count_min": int(row["employee_count_min"]) if row.get("employee_count_min", "").isdigit() else 0,
                        "employee_count_max": int(row["employee_count_max"]) if row.get("employee_count_max", "").isdigit() else 0,
                        "contacts": parse_contacts_csv(row.get("contacts", "")),
                        "enrichment_status": row.get("enrichment_status", ""),
                        "membership_tier": "",
                    }
                    records[key] = rec

    print(f"Loaded {csv_count} records from companies_all.csv ({merge_count} merged, {csv_count - merge_count} new)")
    print(f"Total merged records: {len(records)}")
    return list(records.values())


def get_association(rec: dict) -> str:
    """Get association display string."""
    csv_assoc = rec.get("associations_csv", "")
    if csv_assoc:
        return csv_assoc.replace(";", "; ").strip()
    return rec.get("association", "")


def get_spf_list(rec: dict) -> list[str]:
    """Get normalized spf_services as a list."""
    return parse_spf_services(rec.get("spf_services", []))


def get_email_provider(rec: dict) -> str:
    """Get email provider string."""
    return (rec.get("email_provider") or "").strip()


def get_primary_contact(rec: dict) -> tuple[str, str]:
    """Return (name, email) of first contact."""
    contacts = rec.get("contacts")
    if isinstance(contacts, list) and contacts:
        c = contacts[0]
        if isinstance(c, dict):
            return c.get("name", ""), c.get("email", "")
    return "", ""


def get_tech_stack_str(rec: dict) -> str:
    """Get tech_stack as semicolon-delimited string."""
    ts = rec.get("tech_stack")
    if isinstance(ts, list):
        return "; ".join(ts)
    if isinstance(ts, str):
        return ts
    return ""


# ── Segment filter functions ───────────────────────────────────────

def filter_salesforce(rec: dict) -> bool:
    spf = get_spf_list(rec)
    return any(s.lower() in ("salesforce", "pardot") for s in spf)


def filter_legacy_email(rec: dict) -> bool:
    ep = get_email_provider(rec).lower()
    return ep in ("self-hosted", "self-hosted (on-premise)", "other", "") or not ep


def filter_microsoft_365(rec: dict) -> bool:
    return get_email_provider(rec) == "Microsoft 365"


MARTECH_KEYWORDS = {"hubspot", "marketo", "mailchimp", "pardot", "activecampaign", "constant contact"}

def filter_marketing_automation(rec: dict) -> bool:
    spf = get_spf_list(rec)
    return any(s.strip().lower() in MARTECH_KEYWORDS for s in spf)


def filter_small_mfg(rec: dict) -> bool:
    ec_max = rec.get("employee_count_max", 0) or 0
    website = (rec.get("website") or "").strip()
    return 0 < ec_max <= 100 and bool(website)


def filter_large_mfg(rec: dict) -> bool:
    ec_min = rec.get("employee_count_min", 0) or 0
    return ec_min >= 500


def filter_pma_premium(rec: dict) -> bool:
    tier = (rec.get("membership_tier") or "").upper()
    return tier in ("PLATINUM", "GOLD")


SEGMENTS = [
    ("Salesforce Users", "Filter: SPF services contain Salesforce or Pardot", filter_salesforce),
    ("Legacy Email (On-Prem)", "Filter: email_provider is Self-hosted, Other, or empty", filter_legacy_email),
    ("Microsoft 365 Stack", "Filter: email_provider is Microsoft 365", filter_microsoft_365),
    ("Marketing Automation Users", "Filter: SPF services contain HubSpot, Marketo, Mailchimp, Pardot, ActiveCampaign, or Constant Contact", filter_marketing_automation),
    ("Small Manufacturers", "Filter: employee_count_max <= 100 and has website", filter_small_mfg),
    ("Large Manufacturers", "Filter: employee_count_min >= 500", filter_large_mfg),
    ("PMA Premium Members", "Filter: membership_tier is PLATINUM or GOLD", filter_pma_premium),
]


def build_row(rec: dict) -> list:
    """Build a data row for the sheet."""
    contact_name, contact_email = get_primary_contact(rec)
    return [
        rec.get("company_name", ""),
        (rec.get("website") or "").strip(),
        (rec.get("domain") or "").strip(),
        (rec.get("city") or "").strip(),
        (rec.get("state") or "").strip(),
        (rec.get("street") or "").strip(),
        (rec.get("zip_code") or "").strip(),
        get_association(rec),
        get_email_provider(rec),
        (rec.get("phone") or "").strip(),
        contact_name,
        contact_email,
        get_tech_stack_str(rec),
        rec.get("quality_score", 0) or 0,
        rec.get("quality_grade", ""),
    ]


def write_segment_sheet(wb: Workbook, sheet_name: str, subtitle: str, rows: list[list]):
    """Create a formatted segment sheet."""
    ws = wb.create_sheet(title=sheet_name)

    # Row 1: subtitle
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    cell = ws.cell(row=1, column=1, value=f"{sheet_name} -- {subtitle}")
    cell.font = SUBTITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")

    # Row 2: blank spacer
    # Row 3: headers
    header_row = 3
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_width

    # Data rows starting at row 4
    for row_idx, row_data in enumerate(rows, start=header_row + 1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 11))  # Tech Stack wraps

        # Website hyperlink (column 2)
        website = row_data[1]
        if website:
            ws_cell = ws.cell(row=row_idx, column=2)
            url = website if website.startswith("http") else f"http://{website}"
            ws_cell.hyperlink = url
            ws_cell.font = Font(color="0563C1", underline="single")

        # Quality grade color (column 13)
        grade = row_data[12]
        if grade in GRADE_FILLS:
            ws.cell(row=row_idx, column=13).fill = GRADE_FILLS[grade]

    # Freeze top rows (freeze below header row 3)
    ws.freeze_panes = f"A{header_row + 1}"

    # Auto-filter on header row
    if rows:
        last_col = get_column_letter(len(COLUMNS))
        last_row = header_row + len(rows)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{last_row}"

    return len(rows)


def main():
    print("=" * 60)
    print("GSS Segmented Target Lists Export")
    print("=" * 60)

    all_records = load_data()

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("\n--- Segment Results ---")
    total_rows = 0
    for sheet_name, subtitle, filter_fn in SEGMENTS:
        # Filter, sort by quality_score descending
        matched = [r for r in all_records if filter_fn(r)]
        matched.sort(key=lambda r: r.get("quality_score", 0) or 0, reverse=True)
        rows = [build_row(r) for r in matched]
        count = write_segment_sheet(wb, sheet_name, subtitle, rows)
        total_rows += count
        print(f"  {sheet_name:30s}  {count:>5,} rows")

    # Save
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT_PATH))
    file_size = OUTPUT_PATH.stat().st_size
    print(f"\nTotal segment rows: {total_rows:,} (companies can appear in multiple segments)")
    print(f"Output: {OUTPUT_PATH}")
    print(f"Size: {file_size:,} bytes ({file_size / 1024:.1f} KB)")
    print("Done.")


if __name__ == "__main__":
    main()
