#!/usr/bin/env python3
"""Import seed data from the GSS ULTIMATE NAM Intelligence xlsx file.

Reads association member lists, trade shows, competitor analysis, and key contacts
from the xlsx and writes JSONL files to data/raw/ directories.

Usage:
    python -m scripts.import_xlsx_seed
    python scripts/import_xlsx_seed.py
"""

import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = PROJECT_ROOT / "resources" / "Ai-Marketing-Research-2026" / "GSS_ULTIMATE_NAM_Intelligence_COMPLETE_2026.xlsx"

TIMESTAMP = datetime.now(timezone.utc).isoformat()


def write_jsonl(records: list[dict], output_path: Path) -> int:
    """Write records to a JSONL file, creating directories as needed."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        for rec in records:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")
    return len(records)


def import_association_members(wb, sheet_name: str, assoc_code: str, extra_fields: list[str]) -> int:
    """Import an association member sheet to JSONL.

    Args:
        wb: openpyxl workbook
        sheet_name: Name of the sheet (e.g. 'SOCMA Members (Complete)')
        assoc_code: Association code (e.g. 'SOCMA')
        extra_fields: Column names after 'Company' and 'Member Type'
                      e.g. ['Notes'] or ['Sector', 'Notes'] or ['Industry', 'Notes']
    """
    ws = wb[sheet_name]
    records = []
    all_columns = ["Company", "Member Type"] + extra_fields

    for row in ws.iter_rows(min_row=2, values_only=True):
        company = row[0]
        if not company or str(company).strip() == "":
            continue

        rec = {
            "company_name": str(company).strip(),
            "association": assoc_code,
            "source_url": "xlsx_seed",
            "extracted_at": TIMESTAMP,
            "country": "United States",
            "member_type": str(row[1]).strip() if row[1] else "",
        }

        # Add extra fields dynamically
        for i, field_name in enumerate(extra_fields, start=2):
            key = field_name.lower().replace(" ", "_")
            val = row[i] if i < len(row) and row[i] else ""
            rec[key] = str(val).strip() if val else ""

        records.append(rec)

    output_path = PROJECT_ROOT / "data" / "raw" / assoc_code / "records.jsonl"
    return write_jsonl(records, output_path)


def import_trade_shows(wb) -> int:
    """Import 2026 Trade Shows sheet."""
    ws = wb["2026 Trade Shows"]
    records = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        event = row[0]
        if not event or str(event).strip() == "":
            continue

        rec = {
            "event_name": str(row[0]).strip() if row[0] else "",
            "dates": str(row[1]).strip() if row[1] else "",
            "city": str(row[2]).strip() if row[2] else "",
            "venue": str(row[3]).strip() if row[3] else "",
            "attendance": str(row[4]).strip() if row[4] else "",
            "industry": str(row[5]).strip() if row[5] else "",
            "registration_url": str(row[6]).strip() if row[6] else "",
            "notes": str(row[7]).strip() if row[7] else "",
            "priority": str(row[8]).strip() if row[8] else "",
            "source": "xlsx_seed",
        }
        records.append(rec)

    output_path = PROJECT_ROOT / "data" / "raw" / "events" / "trade_shows.jsonl"
    return write_jsonl(records, output_path)


def import_competitors(wb) -> int:
    """Import Competitor Analysis sheet."""
    ws = wb["Competitor Analysis"]
    records = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        competitor = row[0]
        if not competitor or str(competitor).strip() == "":
            continue

        rec = {
            "competitor": str(row[0]).strip(),
            "presence": str(row[1]).strip() if row[1] else "",
            "strategy_notes": str(row[2]).strip() if row[2] else "",
            "threat_level": str(row[3]).strip() if row[3] else "",
            "source": "xlsx_seed",
        }
        records.append(rec)

    output_path = PROJECT_ROOT / "data" / "raw" / "intelligence" / "competitors.jsonl"
    return write_jsonl(records, output_path)


def import_contacts(wb) -> int:
    """Import Key Contacts sheet."""
    ws = wb["Key Contacts"]
    records = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        org = row[0]
        if not org or str(org).strip() == "":
            continue

        rec = {
            "organization": str(row[0]).strip(),
            "name": str(row[1]).strip() if row[1] else "",
            "email": str(row[2]).strip() if row[2] else "",
            "phone": str(row[3]).strip() if row[3] else "",
            "notes": str(row[4]).strip() if row[4] else "",
            "source": "xlsx_seed",
        }
        records.append(rec)

    output_path = PROJECT_ROOT / "data" / "raw" / "contacts" / "association_contacts.jsonl"
    return write_jsonl(records, output_path)


def main():
    if not XLSX_PATH.exists():
        print(f"ERROR: xlsx file not found at {XLSX_PATH}")
        sys.exit(1)

    print(f"Loading {XLSX_PATH.name}...")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)

    summary = {}

    # Association member sheets
    summary["SOCMA"] = import_association_members(
        wb, "SOCMA Members (Complete)", "SOCMA", ["Notes"]
    )
    summary["AIA"] = import_association_members(
        wb, "AIA Members (Complete)", "AIA", ["Sector", "Notes"]
    )
    summary["AGMA"] = import_association_members(
        wb, "AGMA Members (Complete)", "AGMA", ["Industry", "Notes"]
    )

    # Supplementary data
    summary["Trade Shows"] = import_trade_shows(wb)
    summary["Competitors"] = import_competitors(wb)
    summary["Key Contacts"] = import_contacts(wb)

    wb.close()

    # Print summary
    print("\n=== Import Summary ===")
    total = 0
    for name, count in summary.items():
        print(f"  {name}: {count} records")
        total += count
    print(f"  TOTAL: {total} records")
    print("\nOutput files:")
    print(f"  data/raw/SOCMA/records.jsonl  ({summary['SOCMA']} records)")
    print(f"  data/raw/AIA/records.jsonl    ({summary['AIA']} records)")
    print(f"  data/raw/AGMA/records.jsonl   ({summary['AGMA']} records)")
    print(f"  data/raw/events/trade_shows.jsonl         ({summary['Trade Shows']} records)")
    print(f"  data/raw/intelligence/competitors.jsonl    ({summary['Competitors']} records)")
    print(f"  data/raw/contacts/association_contacts.jsonl ({summary['Key Contacts']} records)")


if __name__ == "__main__":
    main()
